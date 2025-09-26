import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, NamedStyle, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, font as tkfont
from tkinter.simpledialog import Dialog
import json
import os
import sys
import subprocess
from datetime import datetime
from typing import List, Dict, Optional, Tuple, Any
import traceback
from pathlib import Path
from openpyxl.cell.cell import MergedCell
import threading
import time
from watchdog.events import FileSystemEventHandler
from watchdog.observers import Observer
# Add this import at the top of the file
from datetime import timedelta
import re
from pathlib import Path

# Añadir estas constantes al inicio del archivo, después de los imports
CACHE_CONFIG = {
    'max_registry_size': 10000,  # Aumentar capacidad máxima de despachos
    'cache_enabled': True,
    'bultos_cache_size': 1000,   # Máximo de entradas en caché de bultos
    'auto_save_interval': 300,   # Guardar automáticamente cada 5 minutos
    'max_memory_usage_mb': 512,  # Límite de uso de memoria
}

class CacheManager:
    """Gestor de caché para optimizar el acceso a datos"""
    
    def __init__(self):
        self.bultos_cache = {}
        self.articulos_cache = {}
        self.clientes_cache = {}
        self.hit_count = 0
        self.miss_count = 0
        self.last_cleanup = time.time()
        
    def get_bultos(self, codigo_articulo):
        """Obtiene datos de bultos desde la caché"""
        if codigo_articulo in self.bultos_cache:
            self.hit_count += 1
            return self.bultos_cache[codigo_articulo].copy()
        self.miss_count += 1
        return None
        
    def set_bultos(self, codigo_articulo, datos_bultos):
        """Almacena datos de bultos en la caché"""
        if len(self.bultos_cache) >= CACHE_CONFIG['bultos_cache_size']:
            self._cleanup_cache()
        self.bultos_cache[codigo_articulo] = datos_bultos.copy()
        
    def clear_bultos(self, codigo_articulo):
        """Elimina datos de bultos de la caché"""
        if codigo_articulo in self.bultos_cache:
            del self.bultos_cache[codigo_articulo]
            
    def _cleanup_cache(self):
        """Limpia la caché usando estrategia LRU (Least Recently Used)"""
        # Implementación básica - eliminar el 10% de entradas más antiguas
        if self.bultos_cache:
            # Para una implementación real, necesitaríamos tracking de uso
            keys = list(self.bultos_cache.keys())
            remove_count = max(1, len(keys) // 10)
            for key in keys[:remove_count]:
                del self.bultos_cache[key]
                
        self.last_cleanup = time.time()
        
    def get_stats(self):
        """Obtiene estadísticas de la caché"""
        return {
            'bultos_cache_size': len(self.bultos_cache),
            'hit_count': self.hit_count,
            'miss_count': self.miss_count,
            'hit_rate': self.hit_count / (self.hit_count + self.miss_count) 
                    if (self.hit_count + self.miss_count) > 0 else 0
        }

class RegistroChangeHandler(FileSystemEventHandler):
    """Maneja los cambios en los archivos de registro"""
    
    def __init__(self, app):
        self.app = app
    
    def on_modified(self, event):
        """Se llama cuando un archivo es modificado"""
        if not event.is_directory:
            # Verificar si es un archivo de registro
            if any(name in event.src_path for name in ["registro_despachos", "registro_despachosV"]):
                print(f"Archivo modificado detectado: {event.src_path}")
                # Programar la actualización para evitar problemas de threading
                self.app.root.after(1000, self.app.actualizar_registros_si_cambiados)
    
    def on_created(self, event):
        """Se llama cuando se crea un nuevo archivo"""
        if not event.is_directory:
            if any(name in event.src_path for name in ["registro_despachos", "registro_despachosV"]):
                print(f"Nuevo archivo detectado: {event.src_path}")
                self.app.root.after(1000, self.app.actualizar_registros_si_cambiados)

class RegistroDespachos:
    def _obtener_ruta_registro(self, nombre_archivo):
        """Obtiene la ruta completa del archivo de registro"""
        # Obtener ruta del escritorio
        desktop_path = Path.home() / "Desktop"
        if not desktop_path.exists():
            desktop_path = Path.home() / "Escritorio"
        
        return desktop_path / nombre_archivo
    
    def __init__(self, nombre_archivo="registro_despachos.json"):
        self.archivo_registro = self._obtener_ruta_registro(nombre_archivo)
        # Crear directorio si no existe
        self.archivo_registro.parent.mkdir(parents=True, exist_ok=True)
        self.cache_manager = CacheManager()
        self.datos_completos = self._cargar_registro_completo()
        self.auto_save_timer = None
        # ELIMINAR: Llamada al auto-guardado que requiere root
        # self._iniciar_auto_guardado()
        
        print(f"Registro inicializado en: {self.archivo_registro}")
        print(f"Número de despachos en historial: {self.datos_completos['metadata']['total_despachos']}")
        
        # Inicializar datos_modificados
        self.datos_modificados = False

    def _iniciar_auto_guardado_app(self):
        """Inicia el guardado automático periódico desde la aplicación principal"""
        if hasattr(self, 'root') and CACHE_CONFIG['auto_save_interval'] > 0:
            self.auto_save_timer = self.root.after(
                CACHE_CONFIG['auto_save_interval'] * 1000, 
                self._guardar_automatico_app
            )

    def _guardar_automatico_app(self):
        """Guarda automáticamente los datos cada cierto tiempo desde la app"""
        try:
            # Verificar si hay datos modificados en el registro actual
            if hasattr(self, 'registro') and hasattr(self.registro, 'datos_modificados'):
                if self.registro.datos_modificados:
                    self.registro._guardar_registro_completo()
                    self.registro.datos_modificados = False
                    print("Auto-guardado realizado")
        except Exception as e:
            print(f"Error en auto-guardado: {e}")
        finally:
            # Programar próximo auto-guardado
            self._iniciar_auto_guardado_app()
    
    def _cargar_registro_completo(self) -> dict:
        """Carga el registro completo con estructura de datos completa - Versión optimizada"""
        estructura_base = {
            'metadata': {
                'version_estructura': '2.1',  # Incrementar versión
                'fecha_creacion': datetime.now().isoformat(),
                'fecha_ultima_actualizacion': datetime.now().isoformat(),
                'total_despachos': 0,
                'max_despachos_por_dia': CACHE_CONFIG['max_registry_size'] // 365  # Distribución aproximada por día
            },
            'ultimo_id': 0,
            'despachos_por_dia': {},
            'bultos_data': {},
            'cache_info': {
                'compressed': False,
                'last_optimization': datetime.now().isoformat()
            }
        }
        
        try:
            if self.archivo_registro.exists():
                # Para archivos grandes, usar lectura optimizada
                file_size = self.archivo_registro.stat().st_size
                if file_size > 10 * 1024 * 1024:  # > 10MB
                    print(f"Cargando registro grande ({file_size/1024/1024:.2f} MB)...")
                
                with open(self.archivo_registro, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    
                # Verificar que tenga la estructura correcta
                if isinstance(data, dict) and 'despachos_por_dia' in data:
                    print(f"Datos cargados del registro: {data['metadata']['total_despachos']} despachos")
                    
                    # Asegurar que todos los campos necesarios existan
                    if 'bultos_data' not in data:
                        data['bultos_data'] = {}
                    if 'cache_info' not in data:
                        data['cache_info'] = estructura_base['cache_info']
                    
                    # Inicializar caché con datos existentes
                    self._inicializar_cache_desde_datos(data)
                    
                    return data
                
                print("Formato inválido, creando nuevo registro")
            else:
                print("Archivo no existe, creando nuevo registro")
            
            # Crear archivo nuevo
            self._guardar_registro_completo(estructura_base)
            return estructura_base
        
        except (json.JSONDecodeError, IOError) as e:
            print(f"Error al cargar registro: {e}")
            # Si hay error, crear archivo nuevo
            try:
                with open(self.archivo_registro, 'w', encoding='utf-8') as f:
                    json.dump(estructura_base, f, indent=4, ensure_ascii=False)
                return estructura_base
            except IOError:
                return estructura_base
    
    def _inicializar_cache_desde_datos(self, datos):
        """Inicializa la caché con los datos existentes en el registro"""
        # Precargar datos de bultos en caché para acceso rápido
        for codigo, bultos in datos.get('bultos_data', {}).items():
            if len(self.cache_manager.bultos_cache) < CACHE_CONFIG['bultos_cache_size']:
                self.cache_manager.bultos_cache[codigo] = bultos.copy()
        
        print(f"Cache inicializada con {len(self.cache_manager.bultos_cache)} entradas de bultos")
    
    def _guardar_registro_completo(self, datos=None):
        """Guarda el registro completo en el archivo - Versión optimizada"""
        if datos is None:
            datos = self.datos_completos
        
        try:
            # Optimizar datos antes de guardar
            datos_optimizados = self._optimizar_datos_para_almacenamiento(datos)
            
            with open(self.archivo_registro, 'w', encoding='utf-8') as f:
                json.dump(datos_optimizados, f, indent=4, ensure_ascii=False)
            
            print(f"Registro guardado en: {self.archivo_registro}")
            return True
            
        except IOError as e:
            print(f"No se pudo guardar el registro: {str(e)}")
            # Intentar guardar en ubicación alternativa
            try:
                alt_path = Path.home() / "Documents" / self.archivo_registro.name
                with open(alt_path, 'w', encoding='utf-8') as f:
                    json.dump(datos, f, indent=4, ensure_ascii=False)
                print(f"Registro guardado en ubicación alternativa: {alt_path}")
                return True
            except IOError:
                print("Error crítico: No se pudo guardar el registro en ninguna ubicación")
                return False
    
    def _optimizar_datos_para_almacenamiento(self, datos):
        """Optimiza los datos para reducir el tamaño de almacenamiento"""
        # Crear copia para no modificar los datos originales
        datos_optimizados = datos.copy()
        
        # Comprimir datos de bultos si son muy grandes
        bultos_data = datos_optimizados.get('bultos_data', {})
        if len(str(bultos_data)) > 1000000:  # > 1MB de datos de bultos
            print("Optimizando datos de bultos para almacenamiento...")
            # Podríamos añadir compresión aquí si es necesario
        
        # Actualizar metadata de optimización
        datos_optimizados['cache_info']['last_optimization'] = datetime.now().isoformat()
        
        return datos_optimizados
    
    def agregar_despacho(self, datos_despacho: dict) -> str:
        """Agrega un nuevo despacho al historial - Versión optimizada con caché"""
        # Verificar límite de capacidad
        if self.datos_completos['metadata']['total_despachos'] >= CACHE_CONFIG['max_registry_size']:
            self._limpiar_despachos_antiguos()

        # Generar ID único basado en timestamp y contador
        timestamp = datetime.now().strftime('%Y%m%d-%H%M%S-%f')[:-3]
        self.datos_completos['ultimo_id'] += 1
        despacho_id = f"DESP-{timestamp}-{self.datos_completos['ultimo_id']:04d}"
        
        fecha_actual = datetime.now()
        fecha_str = fecha_actual.strftime("%Y-%m-%d")
        fecha_legible = fecha_actual.strftime("%d/%m/%Y %H:%M:%S")
        
        # Procesar artículos para ambos métodos
        articulos_procesados = []
        for articulo in datos_despacho.get('articulos', []):
            articulo_procesado = articulo.copy()
            
            # CORRECCIÓN: No duplicar el código en la descripción
            descripcion = articulo.get('descripcion', '')
            # Si la descripción ya contiene el código, no duplicarlo
            if 'codigo' in articulo and articulo['codigo']:
                codigo = articulo['codigo']
                # Verificar si el código ya está al inicio de la descripción
                if not descripcion.startswith(codigo):
                    descripcion = f"{codigo} - {descripcion}"
            
            articulo_procesado['descripcion_completa'] = descripcion
                
            # Conservar información de bultos detallados si existe
            if 'bultos_detallados' in articulo:
                articulo_procesado['bultos_detallados'] = articulo['bultos_detallados']
                
                # Almacenar en caché para acceso rápido
                codigo = articulo.get('codigo', '')
                if codigo and codigo in datos_despacho.get('bultos_data', {}):  # MODIFICADO: Usar bultos_data de los datos recibidos
                    bultos_data = datos_despacho['bultos_data'][codigo]
                    self.cache_manager.set_bultos(codigo, bultos_data)
                
            # Conservar información de unidad si existe
            if 'unidad' in articulo:
                articulo_procesado['unidad'] = articulo['unidad']
            
            # Extraer información para agrupación (si está disponible)
            partes = [p.strip() for p in descripcion.split(' - ') if p.strip()]
            codigo = partes[0] if partes and ' ' not in partes[0] else ""
            nombre_base = ' '.join(partes[1:]) if len(partes) > 1 else descripcion
            
            # Identificar el color (opcional)
            colores_conocidos = ["BLANCO", "OSCURO", "PASTEL", "ESPECIAL", "MELANGE", 
                                "NEGRO", "ROJO", "AZUL", "VERDE", "GRIS", "BEIGE", 
                                "CREMA", "AMARILLO", "POPULAR", "COLORES"]
            color = "BLANCO"  # Valor por defecto
            
            for parte in reversed(partes):
                parte_upper = parte.upper()
                if any(c in parte_upper for c in colores_conocidos):
                    color = next((c for c in colores_conocidos if c in parte_upper), "BLANCO")
                    break
            
            articulo_procesado.update({
                'codigo': codigo,
                'nombre_base': nombre_base,
                'color': color
            })
            articulos_procesados.append(articulo_procesado)
        
        # NEW: Preserve bultos_data for detailed method
        bultos_data = datos_despacho.get('bultos_data', {})
        
        # Estructura de datos compatible con ambos métodos
        datos_completos = {
            'id': despacho_id,
            'fecha_creacion': fecha_legible,
            'timestamp': datetime.now().isoformat(),
            **datos_despacho,
            'articulos': articulos_procesados,
            'metodo_guardado': 'detallado' if 'bultos_data' in datos_despacho and datos_despacho['bultos_data'] else 'general'
        }
        
        # NEW: Add bultos_data to the structure for detailed dispatches
        if bultos_data:
            datos_completos['bultos_data'] = bultos_data
        
        # Agregar a la estructura por día - PERMITIR MÚLTIPLES DESPACHOS IGUALES
        if fecha_str not in self.datos_completos['despachos_por_dia']:
            self.datos_completos['despachos_por_dia'][fecha_str] = []
        
        # NO VERIFICAR DUPLICADOS - PERMITIR MÚLTIPLES DESPACHOS AL MISMO CLIENTE
        # CON LOS MISMOS ARTÍCULOS PERO DIFERENTES CANTIDADES/PESOS
        self.datos_completos['despachos_por_dia'][fecha_str].append(datos_completos)
        self.datos_completos['metadata']['total_despachos'] += 1
        self.datos_completos['metadata']['fecha_ultima_actualizacion'] = datetime.now().isoformat()
        
        self._guardar_registro_completo()
        # Actualizar caché después de agregar
        self._actualizar_cache_despacho(datos_completos)
        
        return despacho_id
    
    def eliminar_despacho(self, despacho_id: str) -> bool:
        """Elimina un despacho por su ID"""
        for fecha, despachos in self.datos_completos['despachos_por_dia'].items():
            for i, despacho in enumerate(despachos):
                if despacho.get('id') == despacho_id:
                    del self.datos_completos['despachos_por_dia'][fecha][i]
                    
                    # Si no quedan despachos en esa fecha, eliminar la fecha
                    if not self.datos_completos['despachos_por_dia'][fecha]:
                        del self.datos_completos['despachos_por_dia'][fecha]
                    
                    self.datos_completos['metadata']['total_despachos'] -= 1
                    self.datos_completos['metadata']['fecha_ultima_actualizacion'] = datetime.now().isoformat()
                    
                    # Marcar datos como modificados y guardar inmediatamente
                    self.datos_modificados = True
                    self._guardar_registro_completo()
                    return True
        return False
    
    def listar_despachos(self, busqueda: str = None) -> list:
        """Lista todos los despachos, opcionalmente filtrados por búsqueda"""
        todos_despachos = []
        for fecha in sorted(self.datos_completos['despachos_por_dia'].keys(), reverse=True):
            todos_despachos.extend(self.datos_completos['despachos_por_dia'][fecha])
        
        if busqueda:
            busqueda = busqueda.lower()
            despachos_filtrados = []
            for despacho in todos_despachos:
                # Buscar en cliente
                cliente = despacho.get('cliente', {})
                if (busqueda in cliente.get('nombre', '').lower() or 
                    busqueda in cliente.get('rif', '').lower() or
                    busqueda in cliente.get('telefono', '').lower()):
                    despachos_filtrados.append(despacho)
                    continue
                
                # Buscar en artículos
                for articulo in despacho.get('articulos', []):
                    if (busqueda in articulo.get('descripcion', '').lower() or
                        busqueda in articulo.get('codigo', '').lower()):
                        despachos_filtrados.append(despacho)
                        break
                
                # Buscar en fecha
                if busqueda in despacho.get('fecha_creacion', '').lower():
                    despachos_filtrados.append(despacho)
            
            return despachos_filtrados
        
        return todos_despachos
    
    def _actualizar_cache_despacho(self, datos_despacho):
        """Actualiza la caché con los datos del nuevo despacho"""
        for articulo in datos_despacho.get('articulos', []):
            codigo = articulo.get('codigo', '')
            if codigo and 'bultos_data' in datos_despacho and codigo in datos_despacho['bultos_data']:  # MODIFICADO
                bultos_data = datos_despacho['bultos_data'][codigo]
                self.cache_manager.set_bultos(codigo, bultos_data)
    
    def _limpiar_despachos_antiguos(self):
        """Limpia despachos antiguos cuando se alcanza el límite de capacidad"""
        print("Límite de capacidad alcanzado, limpiando despachos antiguos...")
        
        # Encontrar la fecha más antigua
        fechas = sorted(self.datos_completos['despachos_por_dia'].keys())
        if not fechas:
            return
            
        fecha_mas_antigua = fechas[0]
        despachos_a_eliminar = self.datos_completos['despachos_por_dia'][fecha_mas_antigua]
        
        # Eliminar despachos de la fecha más antigua
        del self.datos_completos['despachos_por_dia'][fecha_mas_antigua]
        self.datos_completos['metadata']['total_despachos'] -= len(despachos_a_eliminar)
        
        # Limpiar caché de bultos de los despachos eliminados
        for despacho in despachos_a_eliminar:
            for articulo in despacho.get('articulos', []):
                codigo = articulo.get('codigo', '')
                if codigo:
                    self.cache_manager.clear_bultos(codigo)
        
        print(f"Eliminados {len(despachos_a_eliminar)} despachos de la fecha {fecha_mas_antigua}")
    
    def obtener_despacho(self, despacho_id: str) -> dict:
        """Obtiene un despacho específico por su ID - Versión con caché"""
        # Primero buscar en caché si está disponible (implementación futura)
        # Por ahora, buscar en los datos como antes
        
        for fecha, despachos in self.datos_completos['despachos_por_dia'].items():
            for despacho in despachos:
                if despacho.get('id') == despacho_id:
                    # Cargar datos de bultos desde caché si es posible
                    despacho_completo = despacho.copy()
                    self._cargar_bultos_desde_cache(despacho_completo)
                    return despacho_completo
        return None
    
    def _cargar_bultos_desde_cache(self, despacho):
        """Carga datos de bultos desde la caché para un despacho"""
        if 'articulos' not in despacho:
            return
            
        for articulo in despacho['articulos']:
            codigo = articulo.get('codigo', '')
            if codigo:
                # Intentar obtener de caché primero
                bultos_cache = self.cache_manager.get_bultos(codigo)
                if bultos_cache is not None:
                    articulo['bultos_detallados'] = bultos_cache
                elif 'bultos_data' in despacho and codigo in despacho.get('bultos_data', {}):  # MODIFICADO
                    # Cargar desde datos principales y almacenar en caché
                    bultos_data = despacho['bultos_data'][codigo]
                    articulo['bultos_detallados'] = bultos_data
                    self.cache_manager.set_bultos(codigo, bultos_data)
    
    def verificar_registro(self):
        """Verifica que el registro se esté guardando correctamente"""
        print(f"Ruta del registro: {self.archivo_registro}")
        print(f"¿Existe el archivo? {self.archivo_registro.exists()}")
        
        if self.archivo_registro.exists():
            print(f"Tamaño del archivo: {self.archivo_registro.stat().st_size} bytes")
            print(f"Número de días con despachos: {len(self.datos_completos.get('despachos_por_dia', {}))}")
            print(f"Total de despachos: {self.datos_completos.get('metadata', {}).get('total_despachos', 0)}")

class RegistroDespachosCombinado:
    """Clase que combina múltiples registros de despachos para lectura"""
    
    def __init__(self, nombres_archivos=None):
        if nombres_archivos is None:
            # Nombres de archivos para las 4 máquinas
            nombres_archivos = ["registro_despachos.json", "registro_despachosV2.json", 
                            "registro_despachosV3.json", "registro_despachosV4.json"]
        
        self.registros = []
        self.archivos_existentes = []
        
        # Obtener ruta del escritorio
        desktop_path = Path.home() / "Desktop"
        if not desktop_path.exists():
            desktop_path = Path.home() / "Escritorio"
        
        for nombre in nombres_archivos:
            archivo_path = desktop_path / nombre
            if archivo_path.exists():
                try:
                    print(f"Intentando cargar registro: {nombre}")
                    registro = RegistroDespachos(nombre)
                    self.registros.append(registro)
                    self.archivos_existentes.append(nombre)
                    print(f"✓ Registro {nombre} cargado exitosamente")
                except Exception as e:
                    print(f"✗ Error al cargar registro {nombre}: {str(e)}")
            else:
                print(f"✗ Archivo no encontrado: {nombre}")
        
        print(f"Registros cargados exitosamente: {len(self.registros)}")
        print(f"Archivos encontrados: {self.archivos_existentes}")
        
        # Combinar todos los datos
        self.datos_combinados = self._combinar_registros()
    
    def _combinar_registros(self):
        """Combina los datos de todos los registros existentes"""
        datos_combinados = {
            'metadata': {
                'version_estructura': '2.0',
                'fecha_creacion': datetime.now().isoformat(),
                'fecha_ultima_actualizacion': datetime.now().isoformat(),
                'total_despachos': 0,
                'registros_combinados': self.archivos_existentes,
                'fecha_combinacion': datetime.now().isoformat()
            },
            'ultimo_id': 0,
            'despachos_por_dia': {}
        }
        
        # Combinar despachos de todos los registros existentes
        for registro in self.registros:
            datos = registro.datos_completos
            for fecha, despachos in datos.get('despachos_por_dia', {}).items():
                if fecha not in datos_combinados['despachos_por_dia']:
                    datos_combinados['despachos_por_dia'][fecha] = []
                
                # Añadir información de origen a cada despacho
                for despacho in despachos:
                    despacho_con_origen = despacho.copy()
                    despacho_con_origen['origen_registro'] = registro.archivo_registro.name
                    datos_combinados['despachos_por_dia'][fecha].append(despacho_con_origen)
            
            # Actualizar el último ID
            datos_combinados['ultimo_id'] = max(datos_combinados['ultimo_id'], datos.get('ultimo_id', 0))
        
        # Actualizar el total de despachos
        datos_combinados['metadata']['total_despachos'] = sum(
            len(despachos) for despachos in datos_combinados['despachos_por_dia'].values())
        
        return datos_combinados
    
    def listar_despachos(self, busqueda: str = None) -> list:
        """Lista todos los despachos de todos los registros combinados"""
        todos_despachos = []
        for fecha in sorted(self.datos_combinados['despachos_por_dia'].keys(), reverse=True):
            todos_despachos.extend(self.datos_combinados['despachos_por_dia'][fecha])
        
        if busqueda:
            busqueda = busqueda.lower()
            despachos_filtrados = []
            for despacho in todos_despachos:
                # Buscar en cliente
                cliente = despacho.get('cliente', {})
                if (busqueda in cliente.get('nombre', '').lower() or 
                    busqueda in cliente.get('rif', '').lower() or
                    busqueda in cliente.get('telefono', '').lower()):
                    despachos_filtrados.append(despacho)
                    continue
                
                # Buscar en artículos
                for articulo in despacho.get('articulos', []):
                    if (busqueda in articulo.get('descripcion', '').lower() or
                        busqueda in articulo.get('codigo', '').lower()):
                        despachos_filtrados.append(despacho)
                        break
                
                # Buscar en fecha
                if busqueda in despacho.get('fecha_creacion', '').lower():
                    despachos_filtrados.append(despacho)
                
                # Buscar en origen del registro
                if busqueda in despacho.get('origen_registro', '').lower():
                    despachos_filtrados.append(despacho)
            
            return despachos_filtrados
        
        return todos_despachos
    
    def obtener_despachos_por_fecha(self, fecha: str = None) -> list:
        """Obtiene todos los despachos de una fecha específica de todos los registros"""
        if fecha is None:
            fecha = datetime.now().strftime("%Y-%m-%d")
        
        return self.datos_combinados['despachos_por_dia'].get(fecha, [])
    
    def obtener_todas_fechas(self) -> list:
        """Obtiene todas las fechas que tienen despachos registrados en todos los registros"""
        return sorted(self.datos_combinados['despachos_por_dia'].keys(), reverse=True)
    
    def obtener_despacho(self, despacho_id: str) -> dict:
        """Obtiene un despacho específico por su ID de cualquier registro"""
        for fecha, despachos in self.datos_combinados['despachos_por_dia'].items():
            for despacho in despachos:
                if despacho.get('id') == despacho_id:
                    return despacho.copy()
        return None
    
    def actualizar_registros(self):
        """Actualiza los datos combinados volviendo a cargar todos los registros"""
        for registro in self.registros:
            registro.datos_completos = registro._cargar_registro_completo()
        self.datos_combinados = self._combinar_registros()
    
    def verificar_registros(self):
        """Verifica que todos los registros se estén cargando correctamente"""
        print(f"Registros combinados cargados: {len(self.registros)}")
        for registro in self.registros:
            print(f"Registro: {registro.archivo_registro.name}")
            registro.verificar_registro()
            print("---")

class SeleccionDespachosDialog(tk.Toplevel):
    """Diálogo para seleccionar despachos para exportar en el registro diario"""
    
    def __init__(self, parent, registro: RegistroDespachos, fecha: str = None):
        super().__init__(parent)
        self.title("Seleccionar Despachos para Registro Diario")
        self.geometry("1200x800")
        self.registro = registro
        self.fecha = fecha or datetime.now().strftime("%Y-%m-%d")
        self.despachos_seleccionados = []
        self.result = None
        
        # Frame principal
        main_frame = ttk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Título
        ttk.Label(
            main_frame,
            text=f"Seleccione los despachos para el registro del día {datetime.now().strftime('%d/%m/%Y')}",
            font=('Segoe UI', 12, 'bold')
        ).pack(pady=(0, 10))
        
        # Búsqueda
        search_frame = ttk.Frame(main_frame)
        search_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(search_frame, text="Buscar (cliente, RIF, artículo):").pack(side=tk.LEFT, padx=5)
        self.search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var, width=50)
        search_entry.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)
        search_entry.bind('<KeyRelease>', self._filtrar_despachos)
        
        # Treeview con checkbox para selección
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        columns = ["Seleccionar", "Fecha y Hora", "ID", "Cliente", "RIF", "Teléfono", "Artículos", "Peso Total"]
        self.tree = ttk.Treeview(
            tree_frame, 
            columns=columns,
            show='headings',
            selectmode='extended',
            height=20
        )
        
        # Configurar columnas
        col_widths = {
            "Seleccionar": 80,
            "Fecha y Hora": 150,
            "ID": 180,
            "Cliente": 200, 
            "RIF": 100, 
            "Teléfono": 100,
            "Artículos": 80, 
            "Peso Total": 100
        }
        
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=col_widths.get(col, 100), 
                        anchor=tk.CENTER if col in ["Seleccionar", "Artículos", "Peso Total"] else tk.W)
        
        # Scrollbars
        y_scroll = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        x_scroll = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        
        # Posicionamiento
        self.tree.grid(row=0, column=0, sticky='nsew')
        y_scroll.grid(row=0, column=1, sticky='ns')
        x_scroll.grid(row=1, column=0, sticky='ew')
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # Checkbutton para seleccionar/deseleccionar todos
        select_all_frame = ttk.Frame(main_frame)
        select_all_frame.pack(fill=tk.X, pady=5)
        
        self.select_all_var = tk.BooleanVar()
        select_all_cb = ttk.Checkbutton(
            select_all_frame, 
            text="Seleccionar todos", 
            variable=self.select_all_var,
            command=self._toggle_select_all
        )
        select_all_cb.pack(side=tk.LEFT, padx=5)
        
        # Contador de seleccionados
        self.contador_var = tk.StringVar(value="0 despachos seleccionados")
        ttk.Label(select_all_frame, textvariable=self.contador_var).pack(side=tk.RIGHT, padx=5)
        
        # Botones
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(
            button_frame,
            text="Exportar Seleccionados",
            command=self._exportar_seleccionados,
            style='Accent.TButton'
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            button_frame,
            text="Cancelar",
            command=self._cancelar,
            style='Secondary.TButton'
        ).pack(side=tk.RIGHT, padx=5)
        
        # Cargar datos iniciales
        self._cargar_datos()
        
        # Configurar evento para checkboxes
        self.tree.bind('<Button-1>', self._on_click)
    
    def _cargar_datos(self):
        """Carga todos los despachos del registro"""
        self.tree.delete(*self.tree.get_children())
        self.despachos = self.registro.listar_despachos()
        self.despachos_data = {}  # Diccionario para almacenar datos completos
        
        for despacho in self.despachos:
            fecha = despacho.get('fecha_creacion', '')
            despacho_id = despacho.get('id', '')
            cliente = despacho.get('cliente', {})
            num_articulos = len(despacho.get('articulos', []))
            peso_total = sum(a.get('peso_total', 0) for a in despacho.get('articulos', []))
            
            # Insertar en el treeview con checkbox no seleccionado
            item = self.tree.insert('', tk.END, values=(
                "❌",  # Emoji para no seleccionado
                fecha,
                despacho_id,
                cliente.get('nombre', ''),
                cliente.get('rif', ''),
                cliente.get('telefono', ''),
                num_articulos,
                f"{peso_total:.2f} kg"
            ))
            
            # Guardar los datos completos
            self.despachos_data[item] = despacho
    
    def _filtrar_despachos(self, event=None):
        """Filtra los despachos según el texto de búsqueda"""
        busqueda = self.search_var.get().strip().lower()
        despachos_filtrados = self.registro.listar_despachos(busqueda) if busqueda else self.registro.listar_despachos()
        
        self.tree.delete(*self.tree.get_children())
        self.despachos_data = {}
        
        for despacho in despachos_filtrados:
            fecha = despacho.get('fecha_creacion', '')
            despacho_id = despacho.get('id', '')
            cliente = despacho.get('cliente', {})
            num_articulos = len(despacho.get('articulos', []))
            peso_total = sum(a.get('peso_total', 0) for a in despacho.get('articulos', []))
            
            # Mantener el estado de selección si ya estaba seleccionado
            seleccionado = "✅" if despacho in self.despachos_seleccionados else "❌"
            
            item = self.tree.insert('', tk.END, values=(
                seleccionado,
                fecha,
                despacho_id,
                cliente.get('nombre', ''),
                cliente.get('rif', ''),
                cliente.get('telefono', ''),
                num_articulos,
                f"{peso_total:.2f} kg"
            ))
            
            self.despachos_data[item] = despacho
        
        self._actualizar_contador()
    
    def _on_click(self, event):
        """Maneja el clic en los checkboxes"""
        region = self.tree.identify("region", event.x, event.y)
        if region == "cell":
            column = self.tree.identify_column(event.x)
            item = self.tree.identify_row(event.y)
            
            # Si se hizo clic en la columna de selección (columna 0)
            if column == "#1":
                valores = list(self.tree.item(item, 'values'))
                if valores[0] == "❌":
                    valores[0] = "✅"
                    self.despachos_seleccionados.append(self.despachos_data[item])
                else:
                    valores[0] = "❌"
                    # Remover de seleccionados
                    for i, despacho in enumerate(self.despachos_seleccionados):
                        if despacho.get('id') == self.despachos_data[item].get('id'):
                            del self.despachos_seleccionados[i]
                            break
                
                self.tree.item(item, values=valores)
                self._actualizar_contador()
    
    def _toggle_select_all(self):
        """Selecciona o deselecciona todos los despachos"""
        seleccionar = self.select_all_var.get()
        
        for item in self.tree.get_children():
            valores = list(self.tree.item(item, 'values'))
            
            if seleccionar:
                valores[0] = "✅"
                # Agregar a seleccionados si no está ya
                if self.despachos_data[item] not in self.despachos_seleccionados:
                    self.despachos_seleccionados.append(self.despachos_data[item])
            else:
                valores[0] = "❌"
                # Remover de seleccionados
                self.despachos_seleccionados = [
                    d for d in self.despachos_seleccionados 
                    if d.get('id') != self.despachos_data[item].get('id')
                ]
            
            self.tree.item(item, values=valores)
        
        self._actualizar_contador()
    
    def _actualizar_contador(self):
        """Actualiza el contador de despachos seleccionados"""
        self.contador_var.set(f"{len(self.despachos_seleccionados)} despachos seleccionados")
    
    def _exportar_seleccionados(self):
        """Exporta los despachos seleccionados"""
        if not self.despachos_seleccionados:
            messagebox.showwarning("Advertencia", "Seleccione al menos un despacho para exportar", parent=self)
            return
        
        self.result = self.despachos_seleccionados
        self.destroy()
    
    def _cancelar(self):
        """Cancela la operación"""
        self.result = None
        self.destroy()


class GestionRegistrosDialog(tk.Toplevel):
    """Diálogo para gestionar múltiples registros en modo combinado con capacidad de edición"""
    
    def __init__(self, parent, registro_combinado: RegistroDespachosCombinado):
        super().__init__(parent)
        self.title("Gestión de Registros Combinados - Edición Avanzada")
        self.geometry("1400x800")
        self.registro_combinado = registro_combinado
        self.registro_seleccionado = None
        self.despachos_seleccionados = []
        
        # Frame principal
        main_frame = ttk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Título
        ttk.Label(main_frame, 
                text="Gestión de Registros de Despachos Combinados",
                font=('Segoe UI', 12, 'bold')).pack(pady=(0, 15))
        
        # Información de estado
        info_frame = ttk.Frame(main_frame)
        info_frame.pack(fill=tk.X, pady=(0, 10))
        
        info_text = f"Registros cargados: {len(self.registro_combinado.registros)}/{len(self.registro_combinado.archivos_existentes)}"
        ttk.Label(info_frame, text=info_text, font=('Segoe UI', 10)).pack(side=tk.LEFT)
        
        # Búsqueda
        search_frame = ttk.Frame(info_frame)
        search_frame.pack(side=tk.RIGHT)
        
        ttk.Label(search_frame, text="Buscar:").pack(side=tk.LEFT, padx=(10, 5))
        self.search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var, width=30)
        search_entry.pack(side=tk.LEFT, padx=5)
        search_entry.bind('<KeyRelease>', self._filtrar_despachos)
        
        # Frame para dividir la interfaz
        paned_window = ttk.PanedWindow(main_frame, orient=tk.HORIZONTAL)
        paned_window.pack(fill=tk.BOTH, expand=True, pady=10)
        
        # Panel izquierdo - Lista de registros
        left_frame = ttk.Frame(paned_window)
        paned_window.add(left_frame, weight=1)
        
        # Lista de registros
        ttk.Label(left_frame, text="Registros Disponibles", font=('Segoe UI', 10, 'bold')).pack(pady=(0, 5))
        
        tree_frame = ttk.Frame(left_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        columns = ["Archivo", "Estado", "Despachos", "Última Actualización"]
        self.registros_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=10)
        
        # Configurar columnas
        col_widths = {"Archivo": 180, "Estado": 100, "Despachos": 80, "Última Actualización": 150}
        for col in columns:
            self.registros_tree.heading(col, text=col)
            self.registros_tree.column(col, width=col_widths.get(col, 100), anchor=tk.W)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.registros_tree.yview)
        self.registros_tree.configure(yscrollcommand=scrollbar.set)
        
        # Empaquetar
        self.registros_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Panel derecho - Despachos del registro seleccionado
        right_frame = ttk.Frame(paned_window)
        paned_window.add(right_frame, weight=2)
        
        # Frame para controles de selección
        selection_frame = ttk.Frame(right_frame)
        selection_frame.pack(fill=tk.X, pady=(0, 5))
        
        ttk.Label(selection_frame, text="Despachos del Registro", font=('Segoe UI', 10, 'bold')).pack(side=tk.LEFT)
        
        # Checkbox para seleccionar/deseleccionar todos
        self.select_all_var = tk.BooleanVar()
        select_all_cb = ttk.Checkbutton(
            selection_frame, 
            text="Seleccionar todos", 
            variable=self.select_all_var,
            command=self._toggle_select_all_despachos
        )
        select_all_cb.pack(side=tk.RIGHT, padx=5)
        
        # Contador de seleccionados
        self.contador_seleccionados = tk.StringVar(value="0 despachos seleccionados")
        ttk.Label(selection_frame, textvariable=self.contador_seleccionados).pack(side=tk.RIGHT, padx=10)
        
        # Treeview para despachos
        despachos_frame = ttk.Frame(right_frame)
        despachos_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # AGREGAR COLUMNA DE SELECCIÓN
        despachos_columns = ["Seleccionar", "Fecha", "ID", "Cliente", "RIF", "Artículos", "Peso Total"]
        self.despachos_tree = ttk.Treeview(
            despachos_frame, 
            columns=despachos_columns,
            show='headings',
            height=15,
            selectmode='extended'
        )
        
        # Configurar columnas
        col_widths = {
            "Seleccionar": 80, 
            "Fecha": 120, 
            "ID": 150, 
            "Cliente": 200, 
            "RIF": 100, 
            "Artículos": 80, 
            "Peso Total": 100
        }
        
        for col in despachos_columns:
            self.despachos_tree.heading(col, text=col)
            self.despachos_tree.column(col, width=col_widths.get(col, 100), 
                                    anchor=tk.CENTER if col in ["Seleccionar", "Artículos", "Peso Total"] else tk.W)
        
        # Scrollbars
        y_scroll = ttk.Scrollbar(despachos_frame, orient=tk.VERTICAL, command=self.despachos_tree.yview)
        x_scroll = ttk.Scrollbar(despachos_frame, orient=tk.HORIZONTAL, command=self.despachos_tree.xview)
        self.despachos_tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        
        # Posicionamiento
        self.despachos_tree.grid(row=0, column=0, sticky='nsew')
        y_scroll.grid(row=0, column=1, sticky='ns')
        x_scroll.grid(row=1, column=0, sticky='ew')
        
        despachos_frame.grid_rowconfigure(0, weight=1)
        despachos_frame.grid_columnconfigure(0, weight=1)
        
        # Bind para manejar clicks en checkboxes
        self.despachos_tree.bind('<Button-1>', self._on_despacho_click)
        
        # Botones de acción para despachos
        despachos_buttons = ttk.Frame(right_frame)
        despachos_buttons.pack(fill=tk.X, pady=5)
        
        ttk.Button(
            despachos_buttons,
            text="Eliminar Despacho(s) Seleccionado(s)",
            command=self._eliminar_despachos_seleccionados,
            style='Secondary.TButton'
        ).pack(side=tk.LEFT, padx=5)
        
        # NUEVO BOTÓN: Eliminar todos los despachos del registro
        ttk.Button(
            despachos_buttons,
            text="Eliminar Todos los Despachos",
            command=self._eliminar_todos_despachos,
            style='Danger.TButton'
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            despachos_buttons,
            text="Ver Detalles",
            command=self._ver_detalles_despacho,
            style='Accent.TButton'
        ).pack(side=tk.LEFT, padx=5)
        
        # Botones generales
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(button_frame, text="Actualizar Todo", 
                command=self._actualizar_todo).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Forzar Recarga", 
                command=self._forzar_recarga).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(button_frame, text="Eliminar Registro", 
                command=self._eliminar_registro,
                style='Secondary.TButton').pack(side=tk.LEFT, padx=5)
        
        ttk.Button(button_frame, text="Cerrar", 
                command=self.destroy).pack(side=tk.RIGHT, padx=5)
        
        # Cargar datos
        self._cargar_registros()
        self.registros_tree.bind('<<TreeviewSelect>>', self._on_registro_seleccionado)
        
        # Definir estilo para botón de peligro (rojo)
        style = ttk.Style()
        style.configure('Danger.TButton', foreground='white', background='#dc3545')
    
    def _filtrar_despachos(self, event=None):
        """Filtra los despachos según el texto de búsqueda"""
        if not self.registro_seleccionado:
            return
            
        busqueda = self.search_var.get().strip().lower()
        self._cargar_despachos_registro(self.registro_seleccionado)
    
    def _cargar_registros(self):
        """Carga la lista de registros disponibles en el Treeview"""
        self.registros_tree.delete(*self.registros_tree.get_children())
        
        for registro in self.registro_combinado.registros:
            # Obtener información del registro
            archivo = registro.archivo_registro.name
            estado = "✓ Activo" if registro.archivo_registro.exists() else "✗ No encontrado"
            
            # Contar despachos
            total_despachos = registro.datos_completos.get('metadata', {}).get('total_despachos', 0)
            
            # Obtener última actualización
            ultima_actualizacion = registro.datos_completos.get('metadata', {}).get('fecha_ultima_actualizacion', '')
            if ultima_actualizacion:
                try:
                    fecha_obj = datetime.fromisoformat(ultima_actualizacion)
                    ultima_actualizacion = fecha_obj.strftime("%d/%m/%Y %H:%M")
                except (ValueError, TypeError):
                    pass
            
            # Insertar en el Treeview
            self.registros_tree.insert('', tk.END, values=(
                archivo, estado, total_despachos, ultima_actualizacion
            ))
    
    def _on_registro_seleccionado(self, event):
        """Maneja la selección de un registro"""
        seleccion = self.registros_tree.selection()
        if not seleccion:
            return
            
        item = seleccion[0]
        valores = self.registros_tree.item(item, 'values')
        archivo_seleccionado = valores[0]
        
        # Encontrar el registro correspondiente
        for registro in self.registro_combinado.registros:
            if registro.archivo_registro.name == archivo_seleccionado:
                self.registro_seleccionado = registro
                self._cargar_despachos_registro(registro)
                break
    
    def _cargar_despachos_registro(self, registro):
        """Carga los despachos de un registro específico con columna de selección"""
        self.despachos_tree.delete(*self.despachos_tree.get_children())
        
        # Reiniciar lista de seleccionados
        self.despachos_seleccionados = []
        self.select_all_var.set(False)
        self._actualizar_contador_seleccionados()
        
        # Diccionario para almacenar referencia a los despachos
        self.despachos_data = {}
        
        # Obtener todos los despachos del registro
        todos_despachos = []
        for fecha in sorted(registro.datos_completos.get('despachos_por_dia', {}).keys(), reverse=True):
            todos_despachos.extend(registro.datos_completos['despachos_por_dia'][fecha])
        
        # Aplicar filtro si existe
        busqueda = self.search_var.get().strip().lower()
        if busqueda:
            todos_despachos = [d for d in todos_despachos if 
                            busqueda in d.get('id', '').lower() or
                            busqueda in d.get('cliente', {}).get('nombre', '').lower() or
                            busqueda in d.get('cliente', {}).get('rif', '').lower() or
                            any(busqueda in art.get('descripcion', '').lower() 
                                for art in d.get('articulos', []))]
        
        # Mostrar en el treeview con checkbox
        for despacho in todos_despachos:
            cliente = despacho.get('cliente', {})
            num_articulos = len(despacho.get('articulos', []))
            peso_total = sum(a.get('peso_total', 0) for a in despacho.get('articulos', []))
            
            # Formatear fecha
            fecha_creacion = despacho.get('fecha_creacion', '')
            if ' ' in fecha_creacion:
                fecha = fecha_creacion.split(' ')[0]
            else:
                fecha = fecha_creacion
            
            item = self.despachos_tree.insert('', tk.END, values=(
                "❌",  # Emoji para no seleccionado
                fecha,
                despacho.get('id', ''),
                cliente.get('nombre', ''),
                cliente.get('rif', ''),
                num_articulos,
                f"{peso_total:.2f} kg"
            ))
            
            # Guardar referencia al despacho completo usando el ID del item como clave
            self.despachos_data[item] = despacho
    
    def _on_despacho_click(self, event):
        """Maneja el clic en los checkboxes de despachos"""
        region = self.despachos_tree.identify("region", event.x, event.y)
        if region == "cell":
            column = self.despachos_tree.identify_column(event.x)
            item = self.despachos_tree.identify_row(event.y)
            
            # Si se hizo clic en la columna de selección (columna 0)
            if column == "#1":
                valores = list(self.despachos_tree.item(item, 'values'))
                despacho_id = valores[2]  # ID está en la tercera columna
                
                if valores[0] == "❌":
                    valores[0] = "✅"
                    self.despachos_seleccionados.append(despacho_id)
                else:
                    valores[0] = "❌"
                    if despacho_id in self.despachos_seleccionados:
                        self.despachos_seleccionados.remove(despacho_id)
                
                self.despachos_tree.item(item, values=valores)
                self._actualizar_contador_seleccionados()
                
                # Actualizar checkbox de "Seleccionar todos"
                total_items = len(self.despachos_tree.get_children())
                if len(self.despachos_seleccionados) == total_items:
                    self.select_all_var.set(True)
                else:
                    self.select_all_var.set(False)
    
    def _toggle_select_all_despachos(self):
        """Selecciona o deselecciona todos los despachos"""
        seleccionar = self.select_all_var.get()
        self.despachos_seleccionados = []
        
        for item in self.despachos_tree.get_children():
            valores = list(self.despachos_tree.item(item, 'values'))
            
            if seleccionar:
                valores[0] = "✅"
                despacho_id = valores[2]
                self.despachos_seleccionados.append(despacho_id)
            else:
                valores[0] = "❌"
            
            self.despachos_tree.item(item, values=valores)
        
        self._actualizar_contador_seleccionados()
    
    def _actualizar_contador_seleccionados(self):
        """Actualiza el contador de despachos seleccionados"""
        self.contador_seleccionados.set(f"{len(self.despachos_seleccionados)} despachos seleccionados")
    
    def _eliminar_despachos_seleccionados(self):
        """Elimina los despachos seleccionados del registro"""
        if not self.registro_seleccionado:
            messagebox.showwarning("Advertencia", "Seleccione un registro primero.", parent=self)
            return
        
        if not self.despachos_seleccionados:
            messagebox.showwarning("Advertencia", "Seleccione al menos un despacho para eliminar.", parent=self)
            return
        
        # Confirmar eliminación
        confirmacion = messagebox.askyesno(
            "Confirmar Eliminación",
            f"¿Está seguro de que desea eliminar {len(self.despachos_seleccionados)} despacho(s) seleccionado(s)?\n\n"
            "Esta acción no se puede deshacer.",
            parent=self
        )
        
        if not confirmacion:
            return
        
        # Eliminar cada despacho seleccionado
        despachos_eliminados = 0
        for despacho_id in self.despachos_seleccionados:
            if self.registro_seleccionado.eliminar_despacho(despacho_id):
                despachos_eliminados += 1
        
        # Actualizar la lista
        if despachos_eliminados > 0:
            messagebox.showinfo(
                "Eliminación Exitosa", 
                f"Se eliminaron {despachos_eliminados} despacho(s) correctamente.",
                parent=self
            )
            self._cargar_despachos_registro(self.registro_seleccionado)
            self._cargar_registros()  # Actualizar contadores
    
    def _eliminar_todos_despachos(self):
        """Elimina todos los despachos del registro seleccionado"""
        if not self.registro_seleccionado:
            messagebox.showwarning("Advertencia", "Seleccione un registro primero.", parent=self)
            return
        
        # Obtener todos los despachos del registro
        todos_despachos = []
        for fecha in self.registro_seleccionado.datos_completos.get('despachos_por_dia', {}).keys():
            todos_despachos.extend(self.registro_seleccionado.datos_completos['despachos_por_dia'][fecha])
        
        if not todos_despachos:
            messagebox.showinfo("Información", "El registro seleccionado no tiene despachos para eliminar.", parent=self)
            return
        
        # Confirmar eliminación
        confirmacion = messagebox.askyesno(
            "Confirmar Eliminación Total",
            f"¿Está seguro de que desea eliminar TODOS los {len(todos_despachos)} despachos del registro?\n\n"
            "Esta acción es irreversible y eliminará todos los despachos del registro seleccionado.",
            parent=self
        )
        
        if not confirmacion:
            return
        
        # Eliminar todos los despachos
        try:
            # Crear una copia de las fechas para evitar problemas de modificación durante iteración
            fechas = list(self.registro_seleccionado.datos_completos['despachos_por_dia'].keys())
            
            for fecha in fechas:
                # Eliminar todos los despachos de esta fecha
                self.registro_seleccionado.datos_completos['despachos_por_dia'][fecha] = []
            
            # Eliminar fechas vacías
            self.registro_seleccionado.datos_completos['despachos_por_dia'] = {
                k: v for k, v in self.registro_seleccionado.datos_completos['despachos_por_dia'].items() 
                if v  # Solo mantener fechas con despachos
            }
            
            # Actualizar metadatos
            self.registro_seleccionado.datos_completos['metadata']['total_despachos'] = 0
            self.registro_seleccionado.datos_completos['metadata']['fecha_ultima_actualizacion'] = datetime.now().isoformat()
            
            # Guardar cambios
            self.registro_seleccionado._guardar_registro_completo()
            
            messagebox.showinfo(
                "Eliminación Exitosa", 
                f"Se eliminaron todos los {len(todos_despachos)} despachos del registro.",
                parent=self
            )
            
            # Actualizar la interfaz
            self._cargar_despachos_registro(self.registro_seleccionado)
            self._cargar_registros()
            
        except Exception as e:
            messagebox.showerror(
                "Error", 
                f"No se pudieron eliminar todos los despachos:\n{str(e)}",
                parent=self
            )
    
    def _ver_detalles_despacho(self):
        """Muestra los detalles del despacho seleccionado"""
        seleccion = self.despachos_tree.selection()
        if not seleccion:
            messagebox.showwarning("Advertencia", "Seleccione un despacho para ver sus detalles.", parent=self)
            return
        
        item = seleccion[0]
        despacho = self.despachos_data.get(item)
        
        if not despacho:
            messagebox.showerror("Error", "No se pudieron cargar los detalles del despacho.", parent=self)
            return
        
        # Crear ventana de detalles
        detalles_window = tk.Toplevel(self)
        detalles_window.title(f"Detalles del Despacho - {despacho.get('id', '')}")
        detalles_window.geometry("800x600")
        
        # Frame principal
        main_frame = ttk.Frame(detalles_window, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Información básica
        info_frame = ttk.LabelFrame(main_frame, text="Información del Despacho", padding=10)
        info_frame.pack(fill=tk.X, pady=(0, 10))
        
        campos_info = [
            ("ID:", despacho.get('id', '')),
            ("Fecha:", despacho.get('fecha_creacion', '')),
            ("Cliente:", despacho.get('cliente', {}).get('nombre', '')),
            ("RIF:", despacho.get('cliente', {}).get('rif', '')),
            ("Teléfono:", despacho.get('cliente', {}).get('telefono', '')),
            ("Dirección:", despacho.get('cliente', {}).get('direccion', ''))
        ]
        
        for i, (label, value) in enumerate(campos_info):
            ttk.Label(info_frame, text=label, font=('Segoe UI', 10, 'bold')).grid(row=i, column=0, sticky=tk.W, padx=5, pady=2)
            ttk.Label(info_frame, text=value, font=('Segoe UI', 10)).grid(row=i, column=1, sticky=tk.W, padx=5, pady=2)
        
        # Artículos
        articulos_frame = ttk.LabelFrame(main_frame, text="Artículos", padding=10)
        articulos_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        # Treeview para artículos
        columns = ["Descripción", "Cantidad", "Peso Total"]
        tree = ttk.Treeview(articulos_frame, columns=columns, show='headings', height=10)
        
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=200, anchor=tk.W)
        
        scrollbar = ttk.Scrollbar(articulos_frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Llenar artículos
        for articulo in despacho.get('articulos', []):
            tree.insert('', tk.END, values=(
                articulo.get('descripcion', ''),
                articulo.get('cantidad', ''),
                f"{articulo.get('peso_total', 0):.2f} kg"
            ))
        
        # Botón cerrar
        ttk.Button(main_frame, text="Cerrar", command=detalles_window.destroy).pack(pady=10)
    
    def _actualizar_todo(self):
        """Actualiza todos los registros"""
        self.registro_combinado.actualizar_registros()
        self._cargar_registros()
        messagebox.showinfo("Actualizado", "Todos los registros han sido actualizados.", parent=self)
    
    def _forzar_recarga(self):
        """Fuerza la recarga de todos los registros"""
        nombres_archivos = ["registro_despachos.json", "registro_despachosV2.json", 
                        "registro_despachosV3.json", "registro_despachosV4.json"]
        self.registro_combinado = RegistroDespachosCombinado(nombres_archivos)
        self._cargar_registros()
        messagebox.showinfo("Recargado", "Todos los registros han sido recargados.", parent=self)
    
    def _eliminar_registro(self):
        """Elimina un registro completo"""
        seleccion = self.registros_tree.selection()
        if not seleccion:
            messagebox.showwarning("Advertencia", "Seleccione un registro para eliminar.", parent=self)
            return
        
        item = seleccion[0]
        valores = self.registros_tree.item(item, 'values')
        archivo = valores[0]
        
        confirmacion = messagebox.askyesno(
            "Confirmar Eliminación",
            f"¿Está seguro de que desea eliminar el registro '{archivo}'?\n\n"
            "Esta acción eliminará permanentemente el archivo y todos sus despachos.",
            parent=self
        )
        
        if not confirmacion:
            return
        
        # Encontrar y eliminar el registro
        for registro in self.registro_combinado.registros[:]:
            if registro.archivo_registro.name == archivo:
                try:
                    registro.archivo_registro.unlink()  # Eliminar archivo
                    self.registro_combinado.registros.remove(registro)
                    messagebox.showinfo("Eliminado", f"Registro '{archivo}' eliminado correctamente.", parent=self)
                    self._cargar_registros()
                except Exception as e:
                    messagebox.showerror("Error", f"No se pudo eliminar el registro:\n{str(e)}", parent=self)
                break


class CargarRegistroDialog(tk.Toplevel):
    """Diálogo para cargar despachos desde el registro con información completa"""
    
    def __init__(self, parent, registro: RegistroDespachos):
        super().__init__(parent)
        self.title("Historial de Despachos")
        self.geometry("1300x800")
        self.registro = registro
        self.result = None
        self.despachos_data = {}  # Diccionario para almacenar los datos completos
        
        # Frame principal
        main_frame = ttk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Búsqueda
        search_frame = ttk.Frame(main_frame)
        search_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(search_frame, text="Buscar (cliente, RIF, fecha, artículo):").pack(side=tk.LEFT, padx=5)
        self.search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var, width=50)
        search_entry.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)
        search_entry.bind('<KeyRelease>', self._filtrar_despachos)
        
        # Treeview para mostrar despachos con más columnas
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True)
        
        columns = ["Fecha y Hora", "ID", "Cliente", "RIF", "Teléfono", "Artículos", "Peso Total"]
        self.tree = ttk.Treeview(
            tree_frame, 
            columns=columns,
            show='headings',
            selectmode='browse',
            height=15
        )
        
        # Configurar columnas
        col_widths = {
            "Fecha y Hora": 150,
            "ID": 180,
            "Cliente": 200, 
            "RIF": 100, 
            "Teléfono": 100,
            "Artículos": 80, 
            "Peso Total": 100
        }
        
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=col_widths.get(col, 100), 
                        anchor=tk.CENTER if col in ["Artículos", "Peso Total"] else tk.W)
        
        # Scrollbars
        y_scroll = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        x_scroll = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        
        # Posicionamiento
        self.tree.grid(row=0, column=0, sticky='nsew')
        y_scroll.grid(row=0, column=1, sticky='ns')
        x_scroll.grid(row=1, column=0, sticky='ew')
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # Detalles del despacho seleccionado
        detail_frame = ttk.LabelFrame(main_frame, text="Detalles Completos del Despacho", padding=10)
        detail_frame.pack(fill=tk.X, pady=10)
        
        self.detail_text = tk.Text(
            detail_frame,
            wrap=tk.WORD,
            height=10,
            state=tk.DISABLED,
            font=('Segoe UI', 9)
        )
        self.detail_text.pack(fill=tk.BOTH, expand=True)
        
        # Botones
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(
            button_frame,
            text="Cargar Despacho Seleccionado",
            command=self._cargar_seleccionado,
            style='Accent.TButton'
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            button_frame,
            text="Cancelar",
            command=self._cancelar,
            style='Secondary.TButton'
        ).pack(side=tk.RIGHT, padx=5)
        
        # Cargar datos iniciales
        self._cargar_datos()
        self.tree.bind('<<TreeviewSelect>>', self._mostrar_detalle)
    
    def _cargar_datos(self):
        """Carga todos los despachos del registro con información completa"""
        self.tree.delete(*self.tree.get_children())
        self.despachos = self.registro.listar_despachos()
        self.despachos_data = {}  # Reiniciamos el diccionario
        
        for despacho in self.despachos:
            fecha = despacho.get('fecha_creacion', '')
            despacho_id = despacho.get('id', '')
            cliente = despacho.get('cliente', {})
            num_articulos = len(despacho.get('articulos', []))
            peso_total = sum(a.get('peso_total', 0) for a in despacho.get('articulos', []))
            
            # Insertamos en el treeview
            item = self.tree.insert('', tk.END, values=(
                fecha,
                despacho_id,
                cliente.get('nombre', ''),
                cliente.get('rif', ''),
                cliente.get('telefono', ''),
                num_articulos,
                f"{peso_total:.2f} kg"
            ))
            
            # Guardamos los datos completos en nuestro diccionario usando el item ID como clave
            self.despachos_data[item] = despacho
    
    def _filtrar_despachos(self, event=None):
        """Filtra los despachos según el texto de búsqueda"""
        busqueda = self.search_var.get().strip().lower()
        despachos_filtrados = self.registro.listar_despachos(busqueda) if busqueda else self.registro.listar_despachos()
        
        self.tree.delete(*self.tree.get_children())
        self.despachos_data = {}  # Reiniciamos el diccionario
        
        for despacho in despachos_filtrados:
            fecha = despacho.get('fecha_creacion', '')
            despacho_id = despacho.get('id', '')
            cliente = despacho.get('cliente', {})
            num_articulos = len(despacho.get('articulos', []))
            peso_total = sum(a.get('peso_total', 0) for a in despacho.get('articulos', []))
            
            item = self.tree.insert('', tk.END, values=(
                fecha,
                despacho_id,
                cliente.get('nombre', ''),
                cliente.get('rif', ''),
                cliente.get('telefono', ''),
                num_articulos,
                f"{peso_total:.2f} kg"
            ))
            
            # Guardamos los datos completos en nuestro diccionario
            self.despachos_data[item] = despacho
    
    def _mostrar_detalle(self, event):
        """Muestra los detalles del despacho seleccionado"""
        seleccion = self.tree.selection()
        if not seleccion:
            return
        
        item = seleccion[0]
        despacho = self.despachos_data.get(item)
        
        if not despacho:
            return
        
        self.detail_text.config(state=tk.NORMAL)
        self.detail_text.delete(1.0, tk.END)
        
        # Información básica
        cliente = despacho.get('cliente', {})
        texto = f"ID Despacho: {despacho.get('id', '')}\n"
        texto += f"Fecha Creación: {despacho.get('fecha_creacion', '')}\n"
        if 'fecha_modificacion' in despacho:
            texto += f"Última Modificación: {despacho['fecha_modificacion']}\n"
        texto += "\nDatos del Cliente:\n"
        texto += f"• Nombre: {cliente.get('nombre', '')}\n"
        texto += f"• RIF: {cliente.get('rif', '')}\n"
        texto += f"• Teléfono: {cliente.get('telefono', '')}\n"
        texto += f"• Dirección: {cliente.get('direccion', '')}\n"
        
        # Artículos
        texto += "\nArtículos:\n"
        for i, articulo in enumerate(despacho.get('articulos', []), 1):
            texto += f"{i}. {articulo.get('descripcion', '')}\n"
            texto += f"   - Cantidad: {articulo.get('cantidad', 1)} bultos\n"
            texto += f"   - Peso Total: {articulo.get('peso_total', 0):.2f} kg\n"
            
            # Mostrar pesos individuales si existen
            codigo = articulo.get('codigo', '')
            if codigo and 'bultos_data' in despacho and codigo in despacho['bultos_data']:
                bultos = despacho['bultos_data'][codigo]
                if bultos:
                    texto += "   - Pesos Individuales:\n"
                    for bulto, peso in sorted(bultos.items(), key=lambda x: int(x[0])):
                        texto += f"     Bulto {bulto}: {peso:.2f} kg\n"
        
        texto += f"\nPeso Total del Despacho: {sum(a.get('peso_total', 0) for a in despacho.get('articulos', [])):.2f} kg"
        
        self.detail_text.insert(tk.END, texto)
        self.detail_text.config(state=tk.DISABLED)
    
    def _cargar_seleccionado(self):
        """Carga el despacho seleccionado"""
        seleccion = self.tree.selection()
        if not seleccion:
            messagebox.showwarning("Advertencia", "Seleccione un despacho para cargar", parent=self)
            return
        
        item = seleccion[0]
        self.result = self.despachos_data.get(item)
        
        if self.result:
            self.destroy()
        else:
            messagebox.showerror("Error", "No se pudo cargar el despacho seleccionado", parent=self)
    
    def _cancelar(self):
        """Cancela la operación"""
        self.result = None
        self.destroy() 

def excepthook(exc_type, exc_value, exc_traceback):
    error_msg = "".join(traceback.format_exception(exc_type, exc_value, exc_traceback))
    print("Error detallado:", error_msg)  # Esto aparecerá en la consola
    with open("error_log.txt", "a") as f:
        f.write(error_msg)
    messagebox.showerror("Error crítico", f"Ocurrió un error:\n\n{error_msg}")

# Constantes para configuración
DEFAULT_CONFIG = {
    "ultimo_archivo": "",
    "tipo_articulos": ["Importado", "Nacional"],
    "columnas_clientes": ["Nombre", "RIF", "Teléfono", "Dirección"],
    "columnas_articulos": ["Código Importado", "Descripción Importado", 
                        "Código Nacional", "Descripción Nacional"],
    "formato_fecha": "%d/%m/%Y",
    "mostrar_pesos_individuales": True,
    "color_primario": "#0765a3",
    "color_secundario": "#1f924f",
    "color_fondo": "#f8f9fa"
}

# Modificaciones en la clase CalculadoraPesoDialog
class CalculadoraPesoDialog(tk.Toplevel):
    """Diálogo mejorado para calcular peso de bultos con manejo robusto de datos"""
    
    def __init__(self, parent, pesos_existentes=None):
        super().__init__(parent)
        self.title("Cálculo de Peso de Bultos")
        self.geometry("600x600")
        self.resizable(True, True)
        
        # Inicializar con datos existentes o vacíos
        self.pesos = {}
        if pesos_existentes:
            # Filtrar solo datos válidos (números positivos)
            for k, v in pesos_existentes.items():
                if isinstance(v, (int, float)) and v > 0 and not k.startswith('_'):
                    self.pesos[k] = float(v)
        
        self.total = sum(self.pesos.values()) if self.pesos else 0.0
        self.result = None
        self.entries = []
        self.max_bultos = 200
        
        # Configurar el diálogo para no cerrarse con Enter
        self.bind('<Return>', lambda e: None)
        
        # Crear la interfaz
        self._crear_interfaz()
        
        # Configurar comportamiento al cerrar
        self.protocol("WM_DELETE_WINDOW", self._on_cancel)
        self.grab_set()
    
    def _validar_entrada_peso(self, valor: str) -> bool:
        """Valida que la entrada sea un número decimal positivo o vacío"""
        if not valor:  # Permitir campo vacío
            return True
        try:
            # Verificar que sea un número positivo
            num = float(valor)
            return num >= 0
        except ValueError:
            return False
    
    def _crear_interfaz(self):
        """Crea todos los elementos de la interfaz"""
        # Frame principal con configuración de grid
        main_frame = ttk.Frame(self)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Frame para el contenido con scroll
        content_frame = ttk.Frame(main_frame)
        content_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Frame para botones (fijo a la derecha)
        button_frame = ttk.Frame(main_frame, width=120)
        button_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=(10, 0))
        
        # Canvas para scroll
        canvas = tk.Canvas(content_frame, highlightthickness=0)
        scrollbar = ttk.Scrollbar(content_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Empaquetar widgets de scroll
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Título e instrucciones
        ttk.Label(
            scrollable_frame,
            text="Ingrese pesos de bultos (kg)",
            font=('Segoe UI', 12, 'bold')
        ).pack(pady=(0, 10))
        
        ttk.Label(
            scrollable_frame,
            text="Complete los pesos necesarios (máximo 200 bultos)",
            font=('Segoe UI', 9),
            foreground='#666666'
        ).pack(pady=(0, 15))
        
        # Contenedor para los controles de peso
        entry_frame = ttk.Frame(scrollable_frame)
        entry_frame.pack(fill=tk.BOTH, expand=True)
        
        # Crear campos para hasta max_bultos bultos
        for i in range(1, 201):
            frame = ttk.Frame(entry_frame, padding=2)
            frame.pack(fill=tk.X, pady=1)
            
            ttk.Label(frame, text=f"Bulto {i}:", width=8).pack(side=tk.LEFT)
            
            entry_var = tk.StringVar()
            entry = ttk.Entry(
                frame,
                width=10,
                textvariable=entry_var,
                justify=tk.RIGHT,
                validate="key",
                validatecommand=(frame.register(self._validar_entrada_peso), '%P'))
            entry.pack(side=tk.LEFT, padx=5)
            
            ttk.Label(frame, text="kg").pack(side=tk.LEFT)
            
            # Configurar eventos
            entry.bind("<KeyRelease>", lambda e, idx=i: self._actualizar_peso(idx))
            entry.bind("<Return>", self._manejar_enter)
            entry.bind("<Down>", self._manejar_flecha_abajo)
            entry.bind("<Up>", self._manejar_flecha_arriba)
            
            self.entries.append((entry, entry_var))
        
        # Llenar campos con pesos existentes
        for bulto_num, peso in self.pesos.items():
            try:
                bulto_idx = int(bulto_num) - 1
                if 0 <= bulto_idx < len(self.entries):
                    entry, entry_var = self.entries[bulto_idx]
                    entry_var.set(f"{peso:.2f}")
            except (ValueError, IndexError):
                continue
        
        # Sección de total
        ttk.Label(
            button_frame,
            text="Peso Total:",
            font=('Segoe UI', 10, 'bold')
        ).pack(pady=(10, 5))
        
        self.total_label = ttk.Label(
            button_frame,
            text="0.00 kg",
            font=('Segoe UI', 10)
        )
        self.total_label.pack(pady=(0, 20))
        
        # Botones
        ttk.Button(
            button_frame,
            text="Aceptar",
            command=self._on_accept,
            style='Accent.TButton',
            width=15
        ).pack(pady=5)
        
        ttk.Button(
            button_frame,
            text="Cancelar",
            command=self._on_cancel,
            style='Secondary.TButton',
            width=15
        ).pack(pady=5)
        
        # Ajustar canvas
        scrollable_frame.update_idletasks()
        canvas.config(width=scrollable_frame.winfo_reqwidth())
        
        # Enfocar primer campo
        if self.entries:
            self.entries[0][0].focus_set()
        
        self._calcular_total()
    
    def _actualizar_peso(self, bulto_num: int):
        """Actualiza el peso de un bulto específico, solo si tiene valor"""
        entry, entry_var = self.entries[bulto_num - 1]
        valor = entry_var.get().strip()
        
        if valor:  # Solo actualizar si hay un valor
            try:
                peso = float(valor)
                if peso >= 0:
                    self.pesos[str(bulto_num)] = peso
                else:
                    self.pesos.pop(str(bulto_num), None)  # Eliminar si es negativo
            except ValueError:
                self.pesos.pop(str(bulto_num), None)  # Eliminar si no es número válido
        else:
            self.pesos.pop(str(bulto_num), None)  # Eliminar si está vacío
            
        self._calcular_total()
    
    def _calcular_total(self):
        """Calcula el peso total de todos los bultos y la cantidad real de bultos con peso"""
        self.total = sum(self.pesos.values())
        cantidad_bultos = len(self.pesos)
        self.total_label.config(text=f"{self.total:.2f} kg\nBultos: {cantidad_bultos}")
    
    def _manejar_enter(self, event):
        """Maneja la tecla Enter moviendo el foco al siguiente campo"""
        current = self.focus_get()
        for i, (entry, _) in enumerate(self.entries):
            if entry == current and i < len(self.entries) - 1:
                self.entries[i + 1][0].focus_set()
                break
    
    def _manejar_flecha_abajo(self, event):
        """Maneja la flecha hacia abajo"""
        self._manejar_enter(event)
    
    def _manejar_flecha_arriba(self, event):
        """Maneja la flecha hacia arriba"""
        current = self.focus_get()
        for i, (entry, _) in enumerate(self.entries):
            if entry == current and i > 0:
                self.entries[i - 1][0].focus_set()
                break
    
    def _on_accept(self):
        """Maneja el botón Aceptar, solo incluyendo bultos con peso"""
        # Filtrar solo bultos con peso > 0
        pesos_validos = {}
        for k, v in self.pesos.items():
            if isinstance(v, (int, float)) and v > 0:
                pesos_validos[k] = float(v)
        
        # Si no quedan bultos, devolver None
        if not pesos_validos:
            self.result = None
        else:
            self.result = (pesos_validos, sum(pesos_validos.values()))
        
        self.destroy()
    
    def _on_cancel(self):
        """Maneja el botón Cancelar"""
        self.result = None
        self.destroy()


class CalculadoraMetrosDialog(CalculadoraPesoDialog):
    """Diálogo para calcular metros de bultos con botones fijos"""
    
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Cálculo de Metros de Bultos")
        
        # Cambiar las etiquetas de kg a mts en todos los campos existentes
        for frame in self.winfo_children():
            if isinstance(frame, ttk.Frame):
                for widget in frame.winfo_children():
                    if isinstance(widget, ttk.Label) and widget.cget("text") == "kg":
                        widget.config(text="mts")
        
        # Actualizar la etiqueta de total
        self.total_label.config(text="0.00 mts\nBultos: 0")
    def _on_accept(self):
        """Maneja el botón Aceptar, solo incluyendo bultos con metros"""
        # Filtrar solo bultos con metros > 0 y que sean numéricos
        metros_validos = {}
        for k, v in self.pesos.items():
            if (isinstance(v, (int, float)) and v > 0 and k != '_unidad'):
                metros_validos[k] = float(v)
        
        # Si no quedan bultos, devolver None
        if not metros_validos:
            self.result = None
        else:
            self.result = (metros_validos, sum(metros_validos.values()))
        
        self.destroy()
        

class AplicacionDespachos:
    def __init__(self, root: tk.Tk, nombre_registro="registro_despachos.json", modo_combinado=False):
        self.root = root
        self.config = {}
        self.modo_combinado = modo_combinado
        self.nombre_registro_individual = nombre_registro
        
        # Inicializar cache_manager para todos los modos
        self.cache_manager = CacheManager()
        
        if modo_combinado:
            # Para modo combinado, usar nombres específicos de las 4 máquinas
            nombres_archivos = [
                "registro_despachos.json",      # Máquina 1
                "registro_despachosV2.json",    # Máquina 2  
                "registro_despachosV3.json",    # Máquina 3
                "registro_despachosV4.json"     # Máquina 4
            ]
            self.registro = RegistroDespachosCombinado(nombres_archivos)
            print("Modo combinado activado - Leyendo de múltiples registros")
            
            # Para modo combinado, usar el cache_manager de la aplicación
            for registro in self.registro.registros:
                registro.cache_manager = self.cache_manager
        else:
            self.registro = RegistroDespachos(nombre_registro)
            print(f"Registro individual: {nombre_registro}")
            
            # Para modo individual, usar el cache_manager del registro
            self.cache_manager = self.registro.cache_manager

        # Si estás en modo combinado, usa verificar_registros, sino verificar_registro
        if self.modo_combinado:
            self.registro.verificar_registros()
        else:
            self.registro.verificar_registro()
        
        if not modo_combinado:
            print(f"Ruta del archivo de registro: {self.registro.archivo_registro}")
            print(f"¿Existe el archivo? {self.registro.archivo_registro.exists()}")  # Debug

        self.despacho_actual_id = None
        self._configurar_ventana_principal()
        self._inicializar_datos()
        self._inicializar_estilos()
        
        # Variables Tkinter
        self.peso_total_var = tk.StringVar(value="Peso Total: 0.00 kg")
        self.tipo_var = tk.StringVar()
        
        self._crear_interfaz()
        self._cargar_ultimo_archivo()

        self.datos_sin_guardar = False
        self.cantidad_temp = ""
        self.peso_temp = ""
        self.obs_temp = ""
        self.articulo_seleccionado_temp = None


        # Variables para el monitoreo de cambios
        self.ultima_modificacion = {}
        self.monitor_activo = True
        self.observer = None
        
        #Funcion nueva
        self.despacho_guardado = False  # Add this line
        
        # Iniciar monitoreo de cambios
        self._iniciar_monitoreo_cambios()
        
        # Añadir opción de menú para actualizar registros combinados
        if modo_combinado:
            self._agregar_menu_actualizacion()

        self.bultos_data: Dict[str, Dict[str, float]] = {}

    def _iniciar_monitoreo_cambios(self):
        """Inicia el monitoreo de cambios en los archivos de registro"""
        try:
            # Obtener rutas de todos los archivos a monitorear
            archivos_a_monitorear = []
            
            if self.modo_combinado:
                nombres_archivos = ["registro_despachos.json", "registro_despachosV2.json", 
                                "registro_despachosV3.json", "registro_despachosV4.json"]
                for nombre in nombres_archivos:
                    archivo_path = Path.home() / "Desktop" / nombre
                    if not archivo_path.exists():
                        archivo_path = Path.home() / "Escritorio" / nombre
                    if archivo_path.exists():
                        archivos_a_monitorear.append(str(archivo_path))
            else:
                archivos_a_monitorear.append(str(self.registro.archivo_registro))
            
            # Guardar timestamps iniciales
            for archivo in archivos_a_monitorear:
                if os.path.exists(archivo):
                    self.ultima_modificacion[archivo] = os.path.getmtime(archivo)
            
            # Iniciar observer de watchdog si está disponible
            try:
                from watchdog.observers import Observer
                from watchdog.events import FileSystemEventHandler
                
                class RegistroChangeHandler(FileSystemEventHandler):
                    def __init__(self, app):
                        self.app = app
                    
                    def on_modified(self, event):
                        if not event.is_directory:
                            if any(name in event.src_path for name in ["registro_despachos", ".json"]):
                                print(f"Archivo modificado: {event.src_path}")
                                self.app.root.after(100, self.app.actualizar_registros_si_cambiados)
                
                self.observer = Observer()
                handler = RegistroChangeHandler(self)
                
                # Monitorear directorios que contienen los archivos
                directorios = set(os.path.dirname(archivo) for archivo in archivos_a_monitorear)
                for directorio in directorios:
                    self.observer.schedule(handler, directorio, recursive=False)
                
                self.observer.start()
                print("Monitoreo de cambios activado")
            except ImportError:
                print("Watchdog no disponible, usando verificación periódica")
                # Fallback: verificación periódica cada 5 segundos
                self.root.after(5000, self.verificar_cambios_externos)
                
        except Exception as e:
            print(f"Error al iniciar monitoreo: {e}")
            # Fallback a verificación periódica
            self.root.after(5000, self.verificar_cambios_externos)
    
    def _iniciar_verificacion_periodica(self):
        """Inicia verificación periódica si watchdog no está disponible"""
        def verificar_cambios():
            while self.monitor_activo:
                try:
                    self.actualizar_registros_si_cambiados()
                    time.sleep(5)  # Verificar cada 5 segundos
                except Exception as e:
                    print(f"Error en verificación periódica: {e}")
                    time.sleep(10)
        
        thread = threading.Thread(target=verificar_cambios, daemon=True)
        thread.start()
    
    def actualizar_registros_si_cambiados(self):
        """Verifica si los archivos han cambiado y actualiza si es necesario"""
        try:
            archivos_actualizados = []
            
            # Verificar cada archivo monitoreado
            for archivo, ultima_mod_known in list(self.ultima_modificacion.items()):
                if os.path.exists(archivo):
                    ultima_mod_actual = os.path.getmtime(archivo)
                    if ultima_mod_actual > ultima_mod_known:
                        archivos_actualizados.append(archivo)
                        self.ultima_modificacion[archivo] = ultima_mod_actual
            
            if archivos_actualizados:
                print(f"Archivos actualizados: {archivos_actualizados}")
                self._actualizar_registros_desde_archivos()
                
        except Exception as e:
            print(f"Error al verificar cambios: {e}")
    
    def _actualizar_registros_desde_archivos(self):
        """Actualiza los registros desde los archivos físicos"""
        try:
            if self.modo_combinado:
                # Recargar el registro combinado
                nombres_archivos = ["registro_despachos.json", "registro_despachosV2.json", 
                                "registro_despachosV3.json", "registro_despachosV4.json"]
                self.registro = RegistroDespachosCombinado(nombres_archivos)
            else:
                # Recargar el registro individual
                self.registro.datos_completos = self.registro._cargar_registro_completo()
            
            # Mostrar notificación al usuario
            self._mostrar_notificacion("Registros actualizados", "Los datos de despachos se han actualizado automáticamente.")
            
        except Exception as e:
            print(f"Error al actualizar registros: {e}")
    
    def _mostrar_notificacion(self, titulo, mensaje):
        """Muestra una notificación no intrusiva"""
        # Puedes implementar un sistema de notificación toast
        # Por ahora usamos un mensaje en la barra de estado
        self._actualizar_estado(f"{titulo}: {mensaje}")
        
        # Opcional: mostrar un messagebox si se prefiere
        # self.root.after(100, lambda: messagebox.showinfo(titulo, mensaje, parent=self.root))
    
    def __del__(self):
        """Detener el monitoreo al cerrar la aplicación"""
        if hasattr(self, 'monitor_activo'):
            self.monitor_activo = False
        if hasattr(self, 'observer') and self.observer:
            self.observer.stop()
            if hasattr(self.observer, 'join'):
                self.observer.join()      

    def _agregar_menu_actualizacion(self):
        """Agrega opciones de menú específicas para el modo combinado"""
        # Obtener la barra de menú actual
        menubar = self.root.config('menu')
        if isinstance(menubar, tk.Menu):
            # Buscar el menú Archivo
            for i in range(menubar.index('end') + 1):
                try:
                    label = menubar.entrycget(i, 'label')
                    if label == 'Archivo':
                        # Obtener el submenú Archivo
                        file_menu = menubar.nametowidget(menubar.entrycget(i, 'menu'))
                        # Insertar opción de actualización después del separador
                        file_menu.insert_separator(3)
                        file_menu.insert_command(4, 
                            label="Actualizar Registros Combinados",
                            command=self._actualizar_registros_combinados)
                        break
                except:
                    continue

    def _actualizar_registros_combinados(self):
        """Actualiza los registros combinados recargando todos los archivos"""
        if hasattr(self.registro, 'actualizar_registros'):
            self.registro.actualizar_registros()
            self._mostrar_info("Registros Actualizados", 
                            "Los registros combinados se han actualizado correctamente.")
        else:
            self._mostrar_advertencia("Esta función solo está disponible en modo combinado")

    def _gestionar_registros(self):
        """Muestra el diálogo de gestión de registros combinados"""
        if hasattr(self, 'modo_combinado') and self.modo_combinado:
            dialog = GestionRegistrosDialog(self.root, self.registro)
            self.root.wait_window(dialog)
        else:
            messagebox.showinfo("Información", 
                            "Esta función solo está disponible en modo combinado.", 
                            parent=self.root)   

    # NUEVO MÉTODO EN LA CLASE APLICACIONDESPACHOS
    def _exportar_registro_diario(self):
        """Exporta el registro diario de despachos con selección manual"""
        try:
            # Mostrar diálogo de selección
            dialog = SeleccionDespachosDialog(self.root, self.registro)
            self.root.wait_window(dialog)
            
            if not dialog.result:
                return False  # Usuario canceló
            
            registros_diarios = dialog.result
            
            if not registros_diarios:
                self._mostrar_advertencia("No hay despachos seleccionados para exportar")
                return False
            
            # Obtener escritorio como ubicación por defecto
            desktop_path = Path.home() / "Desktop"
            if not desktop_path.exists():
                desktop_path = Path.home() / "Escritorio"
            
            # Configurar nombre del archivo
            fecha_str = datetime.now().strftime("%d-%m-%Y")
            default_filename = f"Registro_Despachos_Diario_{fecha_str}.xlsx"
            default_filepath = desktop_path / default_filename
            
            filepath = filedialog.asksaveasfilename(
                title="Guardar Registro Diario como...",
                defaultextension=".xlsx",
                filetypes=[("Archivo Excel", "*.xlsx")],
                initialdir=str(desktop_path),
                initialfile=default_filename
            )
            
            if not filepath:
                return False
            
            # Agrupar registros por cliente (RIF + Nombre)
            registros_agrupados = {}
            for registro in registros_diarios:
                cliente = registro.get('cliente', {})
                cliente_rif = cliente.get('rif', 'SIN_RIF')
                cliente_nombre = cliente.get('nombre', 'SIN_NOMBRE')
                clave = f"{cliente_rif}_{cliente_nombre}"
                
                if clave not in registros_agrupados:
                    registros_agrupados[clave] = {
                        'cliente_nombre': cliente_nombre,
                        'cliente_rif': cliente_rif,
                        'articulos': {},
                        'total_bultos': 0,
                        'total_peso': 0.0,
                        'despachos': []
                    }
                
                # Agregar artículos agrupados por código
                for articulo in registro.get('articulos', []):
                    codigo_articulo = articulo.get('codigo', 'SIN_CODIGO')
                    if codigo_articulo not in registros_agrupados[clave]['articulos']:
                        registros_agrupados[clave]['articulos'][codigo_articulo] = {
                            'descripcion': articulo.get('descripcion', ''),
                            'cantidad_bultos': 0,
                            'peso_total': 0.0,
                            'unidad': articulo.get('unidad', 'kg')
                        }
                    
                    registros_agrupados[clave]['articulos'][codigo_articulo]['cantidad_bultos'] += articulo.get('cantidad', 0)
                    registros_agrupados[clave]['articulos'][codigo_articulo]['peso_total'] += articulo.get('peso_total', 0.0)
                
                # Sumar totales
                registros_agrupados[clave]['total_bultos'] += sum(art.get('cantidad', 0) for art in registro.get('articulos', []))
                registros_agrupados[clave]['total_peso'] += sum(art.get('peso_total', 0.0) for art in registro.get('articulos', []))
                registros_agrupados[clave]['despachos'].append(registro.get('fecha_creacion', ''))
                
            # Crear libro de Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "REGISTRO DIARIO"

            # CONFIGURACIÓN DE FUENTE GLOBAL - TODAS LAS LETRAS A 36 PTS
            global_font_size = 36
            
            # CONFIGURACIÓN DE MÁRGENES ESPECÍFICOS (en centímetros convertidos a pulgadas)
            # 0.2 cm = 0.07874 pulgadas, 3 cm = 1.1811 pulgadas
            ws.page_margins.left = 0.07874    # 0.2 cm
            ws.page_margins.right = 1.1811    # 3 cm
            ws.page_margins.top = 0.07874     # 0.2 cm
            ws.page_margins.bottom = 0.07874  # 0.2 cm
            ws.page_margins.header = 0.3
            ws.page_margins.footer = 0.3

            # AMPLIAR ANCHO DE COLUMNAS (más ancho que antes)
            columnas = ['A', 'B', 'C', 'D', 'E', 'F']
            anchos = [45, 30, 25, 180, 25, 35]  # Columnas más anchas
            
            for col, ancho in zip(columnas, anchos):
                ws.column_dimensions[col].width = ancho

            # AUMENTAR ALTURA DE FILAS (mínimo 50 puntos para acomodar fuente de 36pts)
            for row in range(1, 100):
                ws.row_dimensions[row].height = 50

            # Estilo de borde para todas las celdas
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Cabecera de la empresa - FUENTE 36 PTS
            ws.merge_cells('A1:E1')
            ws['A1'] = "TEJIDOS EVERTEX CA - RIF J306528261"
            ws['A1'].font = Font(bold=True, size=global_font_size)
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Aplicar bordes a celdas individuales en lugar del rango combinado
            for col in range(1, 6):
                cell = ws.cell(row=1, column=col)
                cell.border = thin_border

            ws.merge_cells('A2:E2')
            ws['A2'] = "REGISTRO DIARIO DE DESPACHOS"
            ws['A2'].font = Font(bold=True, size=global_font_size)
            ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
            
            for col in range(1, 6):
                cell = ws.cell(row=2, column=col)
                cell.border = thin_border

            ws.merge_cells('A3:E3')
            ws['A3'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y')}"
            ws['A3'].font = Font(bold=True, size=global_font_size)
            ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
            
            for col in range(1, 6):
                cell = ws.cell(row=3, column=col)
                cell.border = thin_border

            # MODIFICACIÓN 1: ELIMINAR ESPACIO ENTRE CABECERA Y TABLA
            # (Se eliminó la fila 4 que antes era espacio en blanco)

            # Encabezados de tabla - FUENTE 36 PTS
            encabezados = ["Nombre Cliente", "Cotización", "Código", "Descripción Artículo", "Bultos", "Cantidad"]

            # MODIFICACIÓN 2: EMPEZAR DIRECTAMENTE EN LA FILA 4 (sin línea de separación)
            for col_num, header in enumerate(encabezados, 1):
                cell = ws.cell(row=4, column=col_num, value=header)
                cell.font = Font(bold=True, size=global_font_size)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill("solid", fgColor="D9E1F2")
                cell.border = thin_border

            # Escribir datos agrupados - FUENTE 36 PTS
            fila_actual = 5
            
            for cliente_key, datos_cliente in registros_agrupados.items():
                # Calcular cuántas filas ocupará este cliente
                num_filas_articulos = len(datos_cliente['articulos'])
                
                # SOLO LA CELDA A (NOMBRE CLIENTE) SE COMBINA VERTICALMENTE
                if num_filas_articulos > 0:
                    ws.merge_cells(start_row=fila_actual, start_column=1, 
                                end_row=fila_actual + num_filas_articulos, end_column=1)
                
                # ESCRIBIR NOMBRE DEL CLIENTE (SOLO EN COLUMNA A)
                cell_nombre = ws.cell(row=fila_actual, column=1, value=datos_cliente['cliente_nombre'])
                cell_nombre.font = Font(bold=True, size=global_font_size)
                cell_nombre.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell_nombre.border = thin_border
                
                # LAS DEMÁS COLUMNAS (B-F) SE LLENAN CON DATOS DESDE ESTA MISMA FILA
                # NO DEJAR CELDAS VACÍAS EN LA PRIMERA FILA
                
                # Columna B: Cotización (primer artículo si existe)
                if datos_cliente['articulos']:
                    primer_articulo = list(datos_cliente['articulos'].values())[0]
                    # Columna B: Cotización (vacío pero con borde)
                    cell = ws.cell(row=fila_actual, column=2, value="")
                    cell.border = thin_border
                    
                    # Columna C: Código del primer artículo
                    descripcion = primer_articulo['descripcion']
                    if ' - ' in descripcion:
                        codigo, _ = descripcion.split(' - ', 1)
                    else:
                        codigo = list(datos_cliente['articulos'].keys())[0]
                    
                    cell = ws.cell(row=fila_actual, column=3, value=codigo)
                    cell.font = Font(size=global_font_size)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                    
                    # Columna D: Descripción del primer artículo
                    descripcion_limpia = descripcion.split(' - ', 1)[1] if ' - ' in descripcion else descripcion
                    cell = ws.cell(row=fila_actual, column=4, value=descripcion_limpia)
                    cell.font = Font(size=global_font_size)
                    cell.alignment = Alignment(vertical='center')
                    cell.border = thin_border
                    
                    # Columna E: Bultos del primer artículo
                    cell = ws.cell(row=fila_actual, column=5, value=primer_articulo['cantidad_bultos'])
                    cell.font = Font(size=global_font_size)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                    
                    # Columna F: Cantidad del primer artículo
                    cantidad_text = f"{primer_articulo['peso_total']:.2f}"
                    if primer_articulo['unidad'] == 'mts':
                        cantidad_text += " mts"
                    else:
                        cantidad_text += " kg"
                    
                    cell = ws.cell(row=fila_actual, column=6, value=cantidad_text)
                    cell.font = Font(size=global_font_size)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                else:
                    # Si no hay artículos, poner celdas vacías pero con bordes
                    for col in range(2, 7):
                        cell = ws.cell(row=fila_actual, column=col, value="")
                        cell.border = thin_border
                
                fila_articulos = fila_actual
                fila_actual += 1
                
                # Escribir el resto de artículos del cliente (si hay más de uno)
                articulos_list = list(datos_cliente['articulos'].items())
                for articulo_idx in range(1, len(articulos_list)):
                    codigo_articulo, articulo = articulos_list[articulo_idx]
                    
                    # Columna A: Nombre Cliente (ya está combinada, dejar vacío pero con borde)
                    # Para celdas combinadas, solo debemos modificar la celda principal
                    if articulo_idx == 0:  # Solo la primera celda de la combinación
                        cell = ws.cell(row=fila_actual, column=1, value="")
                        cell.border = thin_border
                    else:
                        # Para filas posteriores en la combinación, no modificar la celda A
                        pass
                    
                    # Columna B: Cotización (vacío)
                    cell = ws.cell(row=fila_actual, column=2, value="")
                    cell.border = thin_border
                    
                    # Columna C: Código del artículo
                    descripcion = articulo['descripcion']
                    if ' - ' in descripcion:
                        codigo, _ = descripcion.split(' - ', 1)
                    else:
                        codigo = codigo_articulo
                    
                    cell = ws.cell(row=fila_actual, column=3, value=codigo)
                    cell.font = Font(size=global_font_size)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                    
                    # Columna D: Descripción del artículo
                    descripcion_limpia = descripcion.split(' - ', 1)[1] if ' - ' in descripcion else descripcion
                    cell = ws.cell(row=fila_actual, column=4, value=descripcion_limpia)
                    cell.font = Font(size=global_font_size)
                    cell.alignment = Alignment(vertical='center')
                    cell.border = thin_border
                    
                    # Columna E: Bultos
                    cell = ws.cell(row=fila_actual, column=5, value=articulo['cantidad_bultos'])
                    cell.font = Font(size=global_font_size)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                    
                    # Columna F: Cantidad con unidad
                    cantidad_text = f"{articulo['peso_total']:.2f}"
                    if articulo['unidad'] == 'mts':
                        cantidad_text += " mts"
                    else:
                        cantidad_text += " kg"
                    
                    cell = ws.cell(row=fila_actual, column=6, value=cantidad_text)
                    cell.font = Font(size=global_font_size)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                    
                    fila_actual += 1
                
                # Escribir totales del cliente - FUENTE 36 PTS
                # Columna A: Nombre Cliente (vacío)
                # Para celdas combinadas, solo modificar si no es parte de una combinación
                if num_filas_articulos > 0:
                    # Para celdas combinadas, no intentar modificar las celdas internas
                    pass
                else:
                    cell = ws.cell(row=fila_actual, column=1, value="")
                    cell.border = thin_border
                
                # Columna B: Cotización (vacío)
                cell = ws.cell(row=fila_actual, column=2, value="")
                cell.border = thin_border
                
                # Columna C: Código (vacío)
                cell = ws.cell(row=fila_actual, column=3, value="")
                cell.border = thin_border
                
                # Columna D: "TOTALES"
                cell = ws.cell(row=fila_actual, column=4, value="TOTALES")
                cell.font = Font(bold=True, size=global_font_size)
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.border = thin_border
                
                # Columna E: Total bultos
                cell = ws.cell(row=fila_actual, column=5, value=datos_cliente['total_bultos'])
                cell.font = Font(bold=True, size=global_font_size)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                
                # Columna F: Total cantidad con unidad
                misma_unidad = all(art['unidad'] == list(datos_cliente['articulos'].values())[0]['unidad'] 
                                for art in datos_cliente['articulos'].values() if datos_cliente['articulos'])
                
                if misma_unidad and datos_cliente['articulos']:
                    unidad = list(datos_cliente['articulos'].values())[0]['unidad']
                    total_text = f"{datos_cliente['total_peso']:.2f} {unidad}"
                else:
                    total_text = f"{datos_cliente['total_peso']:.2f}"
                
                cell = ws.cell(row=fila_actual, column=6, value=total_text)
                cell.font = Font(bold=True, size=global_font_size)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                
                fila_actual += 2  # Espacio entre clientes

            # Calcular totales generales
            total_general_bultos = sum(cliente['total_bultos'] for cliente in registros_agrupados.values())
            total_general_peso = sum(cliente['total_peso'] for cliente in registros_agrupados.values())
            
            # Escribir totales generales - FUENTE 36 PTS
            # Columna A: Nombre Cliente (vacío)
            cell = ws.cell(row=fila_actual, column=1, value="")
            cell.border = thin_border
            
            # Columna B: Cotización (vacío)
            cell = ws.cell(row=fila_actual, column=2, value="")
            cell.border = thin_border
            
            # Columna C: Código (vacío)
            cell = ws.cell(row=fila_actual, column=3, value="")
            cell.border = thin_border
            
            # Columna D: "TOTAL GENERAL"
            cell = ws.cell(row=fila_actual, column=4, value="TOTAL GENERAL")
            cell.font = Font(bold=True, size=global_font_size)
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = thin_border
            
            # Columna E: Total general bultos
            cell = ws.cell(row=fila_actual, column=5, value=total_general_bultos)
            cell.font = Font(bold=True, size=global_font_size)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            
            # Columna F: Total general cantidad
            cell = ws.cell(row=fila_actual, column=6, value=f"{total_general_peso:.2f}")
            cell.font = Font(bold=True, size=global_font_size)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

            # Asegurar que TODAS las celdas tengan bordes y fuente de 36pts
            for row in range(1, fila_actual + 1):
                for col in range(1, 7):
                    try:
                        cell = ws.cell(row=row, column=col)
                        # Verificar si la celda es parte de un rango combinado
                        if not isinstance(cell, MergedCell):
                            if cell.font.size != global_font_size:
                                cell.font = Font(size=global_font_size, bold=cell.font.bold)
                            if cell.border.left.style is None:
                                cell.border = thin_border
                            if cell.value is None:
                                cell.value = ""
                    except:
                        continue

            # Guardar archivo
            wb.save(filepath)
            
            self._mostrar_info(
                "Exportación exitosa", 
                f"Registro diario exportado correctamente:\n{os.path.basename(filepath)}\n"
                f"Total de clientes: {len(registros_agrupados)}\n"
                f"Total de bultos: {total_general_bultos}\n"
                f"Total peso/metros: {total_general_peso:.2f}"
            )
            
            # Preguntar si abrir el archivo
            respuesta = messagebox.askyesno(
                "Exportación completada",
                "¿Desea abrir el archivo ahora?"
            )
            
            if respuesta:
                self._abrir_archivo(filepath)
                
            return True

        except Exception as e:
            error_msg = f"No se pudo exportar el registro diario:\n{str(e)}"
            self._mostrar_error("Error al exportar", error_msg)
            self._registrar_error(e)
            return False     

    def _configurar_ventana_principal(self):
        """Configura los parámetros iniciales de la ventana principal"""
        self.root.title("Sistema de Gestión de Despachos")
        self.root.geometry("1280x800")
        self.root.minsize(1200, 750)
        self.root.protocol("WM_DELETE_WINDOW", self._confirmar_salida)
        
        # Configurar icono si existe
        icon_path = os.path.join(os.path.dirname(__file__), "icono.ico")
        if os.path.exists(icon_path):
            try:
                self.root.iconbitmap(icon_path)
            except Exception:
                pass

    def _inicializar_datos(self):
        """Inicializa las variables de datos y configuración"""
        self.config_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), 
            'config_despachos.json')
        self.config = self._cargar_configuracion()
        
        # Variables de datos
        self.excel_path = ""
        self.clientes_df = pd.DataFrame()
        self.articulos_df = pd.DataFrame()
        self.cliente_actual: Optional[Dict[str, str]] = None
        self.bultos_data: Dict[str, Dict[str, float]] = {}
        
        # Inicializar widgets como None
        self._inicializar_widgets()

    def _inicializar_estilos(self):
        """Configura los estilos visuales de la aplicación"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Configurar paleta de colores desde la configuración
        self.colores = {
            'primario': self._obtener_color_config('color_primario', '#0765a3'),
            'secundario': self._obtener_color_config('color_secundario', '#1f924f'),
            'fondo': self._obtener_color_config('color_fondo', '#f8f9fa'),
            'texto': '#2c3e50',
            'exito': '#27ae60',
            'error': '#e74c3c'
        }
        
        # Configurar estilos base
        style.configure('.', 
                    font=('Segoe UI', 10),
                    background=self.colores['fondo'])
        
        # Estilos para botones
        button_config = {
            'font': ('Segoe UI', 10, 'bold'),
            'padding': 6,
            'borderwidth': 1
        }
        
        style.configure('Accent.TButton',
                    foreground='white',
                    background=self.colores['primario'],
                      **button_config)
        
        style.configure('Secondary.TButton',
                    foreground='white',
                    background=self.colores['secundario'],
                      **button_config)
        
        # Estilo para campos inválidos
        style.configure('Error.TEntry',
                    foreground='black',
                    fieldbackground='#ffdddd',
                    borderwidth=1)
        
        # Estilo para Treeview
        style.configure('Custom.Treeview', 
                    rowheight=30,
                    fieldbackground='white',
                    background='white')
        
        style.configure('Custom.Treeview.Heading', 
                    font=('Segoe UI', 10, 'bold'),
                    background=self.colores['primario'],
                    foreground='white')
    
    def _confirmar_salida(self):
        """Cierra la aplicación con confirmación si hay datos sin guardar"""
        if hasattr(self, 'tree') and self.tree.get_children():
            respuesta = messagebox.askyesnocancel(
                "Confirmar Salida",
                "Tiene un despacho en progreso con artículos no guardados.\n\n"
                "¿Desea guardar antes de salir?",
                icon=messagebox.WARNING,
                default=messagebox.YES
            )
            
            if respuesta is None:  # Cancelar
                return
            elif respuesta:  # Sí, guardar
                if not self._guardar_despacho():
                    return  # No salir si el guardado falla
        
        # Guardar configuración antes de salir
        try:
            self._guardar_configuracion()
        except Exception as e:
            messagebox.showerror(
                "Error", 
                f"No se pudo guardar la configuración:\n{str(e)}\n"
                "La aplicación se cerrará de todos modos.")
        
        self.root.destroy()

    def _crear_seccion_tejido(self, parent):
        """Eliminada - La sección de tejido ya está incluida en la sección del cliente"""
        pass
        
    def _obtener_color_config(self, clave: str, default: str) -> str:
        """Obtiene un color de la configuración validando su formato"""
        if not hasattr(self, 'config'):
            return default
        color = self.config.get(clave, default)
        return color if self._es_color_valido(color) else default

    def _es_color_valido(self, color: Any) -> bool:
        """Valida que un valor sea un color hexadecimal válido"""
        if not hasattr(self, 'config'):
            return False
        if not isinstance(color, str):
            return False
        if color.startswith('#') and len(color) in (4, 7):
            try:
                int(color[1:], 16)
                return True
            except ValueError:
                pass
        return False

    def _validar_archivo_excel(self, filepath):
        """Valida que el archivo Excel tenga la estructura requerida con manejo de errores mejorado"""
        try:
            with pd.ExcelFile(filepath) as xls:
                # Verificar existencia de hojas requeridas
                hojas_requeridas = ['Clientes', 'ARTICULOS']
                hojas_faltantes = [hoja for hoja in hojas_requeridas if hoja not in xls.sheet_names]
                
                if hojas_faltantes:
                    return False, f"Faltan hojas requeridas: {', '.join(hojas_faltantes)}"
                
                # Verificar que la hoja ARTICULOS tenga al menos 4 columnas
                try:
                    df_articulos = pd.read_excel(xls, sheet_name='ARTICULOS', nrows=1)
                    if len(df_articulos.columns) < 4:
                        return False, "La hoja 'ARTICULOS' debe tener al menos 4 columnas"
                    
                    # Verificar que la hoja Clientes tenga columnas básicas
                    df_clientes = pd.read_excel(xls, sheet_name='Clientes', nrows=1)
                    if len(df_clientes.columns) < 2:
                        return False, "La hoja 'Clientes' debe tener al menos 2 columnas"
                    
                    return True, ""
                except Exception as e:
                    return False, f"Error al leer hojas: {str(e)}"
        except Exception as e:
            return False, f"Error al abrir archivo: {str(e)}"
    
    def _inicializar_widgets(self):
        """Inicializa todos los widgets como None"""
        widgets = [
            'cantidad_entry', 'tipo_combobox', 'peso_entry', 'obs_entry',
            'cliente_search', 'clientes_listbox', 'cliente_nombre',
            'cliente_rif', 'cliente_telefono', 'cliente_direccion',
            'tree', 'peso_total_var', 'btn_agregar', 'btn_guardar',
            'btn_exportar', 'tejido_entry', 'codigo_tejido_entry'
        ]
        for widget in widgets:
            setattr(self, widget, None) 
    def _cargar_configuracion(self) -> Dict[str, Any]:
        """Carga la configuración desde archivo JSON o crea una nueva con valores por defecto"""
        try:
            if os.path.exists(self.config_path):
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    # Asegurar que todas las claves necesarias existan
                    return {**DEFAULT_CONFIG, **config}
            
            # Crear archivo de configuración si no existe
            with open(self.config_path, 'w', encoding='utf-8') as f:
                json.dump(DEFAULT_CONFIG, f, indent=4, ensure_ascii=False)
            return DEFAULT_CONFIG.copy()
            
        except Exception as e:
            self._mostrar_error(
                "Error de Configuración",
                f"No se pudo cargar la configuración:\n{str(e)}\nSe usarán valores por defecto.")
            self._registrar_error(e)
            return DEFAULT_CONFIG.copy()

    def _guardar_configuracion(self) -> bool:
        """Guarda la configuración actual en el archivo JSON"""
        try:
            with open(self.config_path, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, indent=4, ensure_ascii=False)
            return True
        except Exception as e:
            self._mostrar_error("Error", f"No se pudo guardar la configuración:\n{str(e)}")
            self._registrar_error(e)
            return False

    def _crear_interfaz(self):
        """Construye la interfaz gráfica principal usando pack()"""
        self._crear_menu_principal()
        
        # Frame principal con pestañas
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        
        # Pestaña de Despacho
        self.despacho_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.despacho_frame, text="Gestión de Despachos")
        
        # Secciones usando pack()
        self._crear_seccion_cliente(self.despacho_frame)
        self._crear_seccion_articulos(self.despacho_frame)
        self._crear_botones_accion(self.despacho_frame)
        
        # Barra de estado
        self.statusbar = ttk.Label(self.root, text="Listo", relief=tk.SUNKEN, anchor=tk.W)
        self.statusbar.pack(side=tk.BOTTOM, fill=tk.X)

    def _nuevo_despacho(self):
        """Limpia el formulario para un nuevo despacho"""
        # Verificar si hay datos sin guardar
        if hasattr(self, 'tree') and self.tree.get_children():
            respuesta = messagebox.askyesnocancel(
                "Nuevo Despacho",
                "Tiene un despacho en progreso con artículos no guardados.\n\n"
                "¿Desea guardar antes de limpiar el formulario?",
                icon=messagebox.WARNING,
                default=messagebox.YES
            )
            
            if respuesta is None:  # Cancelar
                return
            elif respuesta:  # Sí, guardar
                if not self._guardar_despacho():
                    return  # No continuar si el guardado falla
        
        # Limpiar campos del cliente
        if hasattr(self, 'cliente_nombre'):
            self.cliente_nombre.delete(0, tk.END)
        if hasattr(self, 'cliente_rif'):
            self.cliente_rif.delete(0, tk.END)
        if hasattr(self, 'cliente_telefono'):
            self.cliente_telefono.delete(0, tk.END)
        if hasattr(self, 'cliente_direccion'):
            self.cliente_direccion.delete(0, tk.END)
            
        if hasattr(self, 'clientes_listbox'):
            self.clientes_listbox.selection_clear(0, tk.END)
            
        self.cliente_actual = None
        
        # Limpiar artículos
        if hasattr(self, 'tree'):
            for item in self.tree.get_children():
                self.tree.delete(item)
        
        # LIMPIAR COMPLETAMENTE los datos de bultos
        if hasattr(self, 'bultos_data'):
            self.bultos_data.clear()
            
        # Restablecer totales
        if hasattr(self, 'peso_total_var'):
            self.peso_total_var.set("Peso Total: 0.00 kg")
        
        # Enfocar búsqueda de cliente
        if hasattr(self, 'cliente_search'):
            self.cliente_search.focus()
        
        self._actualizar_estado("Listo para nuevo despacho")
        self.despacho_actual_id = None  # Resetear el ID al crear nuevo despacho

    def _cargar_datos_despacho(self, datos_despacho: dict):
        """Carga los datos de un despacho en la interfaz - compatible con ambos métodos"""
        try:
            # Limpiar el despacho actual
            self._nuevo_despacho()
            
            # Guardar el ID del despacho cargado
            self.despacho_actual_id = datos_despacho.get('id')
            
            # Cargar datos del cliente
            cliente_data = datos_despacho.get('cliente', {})
            if cliente_data:
                self.cliente_actual = {
                    'Nombre': cliente_data.get('nombre', ''),
                    'RIF': cliente_data.get('rif', ''),
                    'Teléfono': cliente_data.get('telefono', ''),
                    'Dirección': cliente_data.get('direccion', '')
                }
                
                # Actualizar campos de cliente
                for attr, key in [('cliente_nombre', 'Nombre'), ('cliente_rif', 'RIF'), 
                                ('cliente_telefono', 'Teléfono'), ('cliente_direccion', 'Dirección')]:
                    if hasattr(self, attr):
                        widget = getattr(self, attr)
                        widget.delete(0, tk.END)
                        widget.insert(0, self.cliente_actual[key])
            
            # Cargar datos de bultos (solo si existe y es método detallado)
            metodo = datos_despacho.get('metodo_guardado', 'general')
            if metodo == 'detallado' and 'bultos_data' in datos_despacho:
                self.bultos_data = datos_despacho.get('bultos_data', {}).copy()
            
            # Cargar artículos - CORRECCIÓN MEJORADA
            for articulo in datos_despacho.get('articulos', []):
                codigo = articulo.get('codigo', '')
                descripcion = articulo.get('descripcion', '')
                cantidad = articulo.get('cantidad', 1)
                peso_total = articulo.get('peso_total', 0.0)
                
                # Determinar la descripción a mostrar
                if 'descripcion_completa' in articulo:
                    descripcion_completa = articulo['descripcion_completa']
                elif codigo and not descripcion.startswith(codigo):
                    descripcion_completa = f"{codigo} - {descripcion}"
                else:
                    descripcion_completa = descripcion
                
                cantidad = articulo.get('cantidad', 1)
                peso_total = articulo.get('peso_total', 0.0)
                
                # Crear valores para el Treeview
                valores = [
                    str(cantidad),
                    descripcion_completa,  # Usar la descripción correcta
                    f"{peso_total:.2f}"
                ]
                
                # Agregar pesos de bultos si existen (método detallado)
                if metodo == 'detallado' and codigo in self.bultos_data:
                    for i in range(1, 201):  # Para hasta 200 bultos
                        peso = self.bultos_data[codigo].get(str(i), 0.0)
                        valores.append(f"{peso:.2f}" if peso > 0 else "")
                
                # Insertar en el Treeview
                if hasattr(self, 'tree'):
                    tags = ('evenrow',) if len(self.tree.get_children()) % 2 == 0 else ('oddrow',)
                    self.tree.insert('', tk.END, values=valores, tags=tags)
            
            # Actualizar peso total
            self._actualizar_peso_total()
            self._actualizar_estado(f"Despacho cargado: {self.cliente_actual.get('Nombre', '')} (ID: {self.despacho_actual_id}, Método: {metodo})")
            
        except Exception as e:
            self._mostrar_error("Error al cargar despacho", f"No se pudieron cargar los datos:\n{str(e)}")
            self._registrar_error(e)

    def _cargar_despacho_registro(self):
        """Muestra el diálogo para cargar un despacho desde el registro"""
        dialog = CargarRegistroDialog(self.root, self.registro)
        self.root.wait_window(dialog)
        
        if dialog.result:
            self._cargar_datos_despacho(dialog.result)

    def _crear_menu_principal(self):
        """Crea la barra de menú principal con la nueva opción"""
        menubar = tk.Menu(self.root)
        
        # Menú Archivo (modificado)
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="Nuevo Despacho", command=self._nuevo_despacho)
        file_menu.add_command(label="Abrir Archivo...", command=self._abrir_archivo_excel)
        file_menu.add_command(label="Cargar del Registro...", command=self._cargar_despacho_registro)
        file_menu.add_separator()
        file_menu.add_separator()
        menubar.add_cascade(label="Archivo", menu=file_menu)

        # Submenú para gestión de registros (solo en modo combinado)
        if hasattr(self, 'modo_combinado') and self.modo_combinado:
            registros_menu = tk.Menu(file_menu, tearoff=0)
            registros_menu.add_command(label="Gestionar Registros...", 
                                    command=self._gestionar_registros)
            registros_menu.add_command(label="Actualizar Registros", 
                                    command=self._actualizar_registros_combinados)
            file_menu.add_cascade(label="Registros Combinados", menu=registros_menu)
            file_menu.add_separator()
        
        file_menu.add_command(label="Guardar Despacho", command=self._guardar_despacho)
        file_menu.add_command(label="Exportar a Excel", command=self._exportar_excel)
        file_menu.add_command(label="Exportar Registro del Día", command=self._exportar_registro_diario)
        file_menu.add_separator()
        file_menu.add_command(label="Salir", command=self._confirmar_salida)
        
        # Menú Despacho
        despacho_menu = tk.Menu(menubar, tearoff=0)
        despacho_menu.add_command(label="Agregar Artículo", command=self._agregar_articulo)
        despacho_menu.add_command(label="Eliminar Artículo", command=self._eliminar_articulo)
        despacho_menu.add_separator()
        despacho_menu.add_command(label="Calcular Peso Total", command=self._actualizar_peso_total)
        menubar.add_cascade(label="Despacho", menu=despacho_menu)
        
        # Menú Herramientas
        tools_menu = tk.Menu(menubar, tearoff=0)
        tools_menu.add_command(label="Calculadora de Peso", command=self._mostrar_calculadora_peso)
        menubar.add_cascade(label="Herramientas", menu=tools_menu)
        
        # Menú Ayuda
        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="Documentación", command=self._mostrar_documentacion)
        help_menu.add_command(label="Acerca de...", command=self._mostrar_acerca_de)
        menubar.add_cascade(label="Ayuda", menu=help_menu)
        
        self.root.config(menu=menubar)

    def _crear_seccion_cliente(self, parent):
        """Crea la sección de datos del cliente (versión simplificada)"""
        cliente_frame = ttk.LabelFrame(
            parent,
            text=" Datos del Cliente ",
            padding=(15, 10))
        cliente_frame.pack(fill=tk.BOTH, pady=5)
        
        # Búsqueda de cliente
        search_frame = ttk.Frame(cliente_frame)
        search_frame.pack(fill=tk.X, pady=(0, 15))
        
        ttk.Label(
            search_frame,
            text="Buscar Cliente:",
            font=('Segoe UI', 10, 'bold')
        ).pack(side=tk.LEFT, padx=(0, 10))
        
        self.cliente_search = ttk.Entry(
            search_frame,
            width=40)
        self.cliente_search.pack(side=tk.LEFT, expand=True, fill=tk.X, padx=5)
        self.cliente_search.bind('<Return>', lambda e: self._buscar_cliente())
        
        ttk.Button(
            search_frame,
            text="Buscar",
            command=self._buscar_cliente,
            style='Accent.TButton',
            width=10
        ).pack(side=tk.LEFT, padx=5)
        
        # Contenedor principal
        content_frame = ttk.Frame(cliente_frame)
        content_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))
        
        # Lista de clientes
        list_frame = ttk.Frame(content_frame)
        list_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 15))
        
        scroll_y = ttk.Scrollbar(list_frame, orient=tk.VERTICAL)
        self.clientes_listbox = tk.Listbox(
            list_frame,
            height=6,
            width=50,
            yscrollcommand=scroll_y.set,
            font=('Segoe UI', 10),
            selectbackground=self.colores['primario'])
        scroll_y.config(command=self.clientes_listbox.yview)
        
        self.clientes_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        self.clientes_listbox.bind('<<ListboxSelect>>', self._seleccionar_cliente)
        
        # Datos del cliente seleccionado
        data_frame = ttk.LabelFrame(
            content_frame,
            text=" Información del Cliente Seleccionado ",
            padding=15)
        data_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        
        # Campos de información del cliente
        campos = [
            ("Nombre:", 'cliente_nombre', 40),
            ("RIF:", 'cliente_rif', 20),
            ("Teléfono:", 'cliente_telefono', 20),
            ("Dirección:", 'cliente_direccion', 40)
        ]
        
        for i, (texto, attr, ancho) in enumerate(campos):
            ttk.Label(
                data_frame,
                text=texto,
                font=('Segoe UI', 10, 'bold')
            ).grid(row=i, column=0, sticky=tk.W, pady=5, padx=5)
            
            entry = ttk.Entry(
                data_frame,
                width=ancho)
            entry.grid(row=i, column=1, sticky=tk.W, pady=5, padx=5)
            setattr(self, attr, entry)
    def _crear_seccion_articulos(self, parent):
        """Crea la sección de artículos usando pack() consistentemente"""
        articulos_frame = ttk.LabelFrame(
            parent,
            text=" Artículos del Despacho ",
            padding=(15, 10))
        articulos_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        # Treeview y controles - todos con pack()
        self._crear_treeview_articulos(articulos_frame)
        self._crear_controles_articulos(articulos_frame)

    def _crear_treeview_articulos(self, parent):
        """Crea el Treeview con columnas para los bultos (hasta 200 bultos) usando pack()"""
        # Frame contenedor principal
        tree_frame = ttk.Frame(parent)
        tree_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 15))

        # Frame interno para el treeview y scrollbar vertical
        tree_container = ttk.Frame(tree_frame)
        tree_container.pack(fill=tk.BOTH, expand=True)
        
        # Scrollbar vertical
        scroll_y = ttk.Scrollbar(tree_container, orient=tk.VERTICAL)
        
        # En _crear_treeview_articulos() creas 200 columnas siempre
        columns = ['Cantidad', 'Descripción', 'Peso Total (kg)'] + [f'Bulto {i} (kg)' for i in range(1, 201)]
        
        # Treeview
        self.tree = ttk.Treeview(
            tree_container,
            columns=columns,
            show='headings',
            height=8,
            selectmode='extended',
            yscrollcommand=scroll_y.set,
            style='Custom.Treeview'
        )
        
        # Configurar columnas
        self.tree.heading('Cantidad', text='Cantidad', anchor=tk.CENTER)
        self.tree.column('Cantidad', width=80, anchor=tk.CENTER, stretch=False)
        
        self.tree.heading('Descripción', text='Descripción', anchor=tk.W)
        self.tree.column('Descripción', width=400, anchor=tk.W, stretch=False)
        
        self.tree.heading('Peso Total (kg)', text='Peso Total (kg)', anchor=tk.CENTER)
        self.tree.column('Peso Total (kg)', width=120, anchor=tk.CENTER, stretch=False)
        
        # Columnas para bultos
        for i in range(1, 201):
            col_name = f'Bulto {i} (kg)'
            self.tree.heading(col_name, text=col_name, anchor=tk.CENTER)
            self.tree.column(col_name, width=110, anchor=tk.CENTER, stretch=False)
        
        # Scrollbar horizontal
        scroll_x = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(xscrollcommand=scroll_x.set)
        
        # Posicionamiento con pack()
        scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll_x.pack(fill=tk.X)

    def _editar_articulo_seleccionado(self):
        """Edita artículo seleccionado con gestión mejorada de caché"""
        seleccion = self.tree.selection()
        if not seleccion:
            self._mostrar_advertencia("Seleccione un artículo para editar")
            return
        
        item = seleccion[0]
        valores = self.tree.item(item, 'values')
        
        if not valores or len(valores) < 2:
            self._mostrar_advertencia("No se pueden obtener los datos del artículo seleccionado")
            return
        
        # Extraer información del artículo
        descripcion = valores[1]
        codigo = descripcion.split(' - ')[0] if ' - ' in descripcion else ""
        
        if not codigo:
            self._mostrar_advertencia("No se puede identificar el código del artículo")
            return
        
        # OBTENER PESOS DESDE CACHÉ PRIMERO (más rápido)
        # Usar el cache_manager correcto según el modo
        if self.modo_combinado:
            pesos_existentes = self.cache_manager.get_bultos(codigo)
        else:
            pesos_existentes = self.registro.cache_manager.get_bultos(codigo)
            
        if pesos_existentes is None and codigo in self.bultos_data:
            # Si no está en caché, cargar desde datos principales
            pesos_existentes = self.bultos_data[codigo].copy()
            # Y almacenar en caché para próximas veces
            if self.modo_combinado:
                self.cache_manager.set_bultos(codigo, pesos_existentes)
            else:
                self.registro.cache_manager.set_bultos(codigo, pesos_existentes)

        # Crear diálogo de calculadora con los pesos existentes
        dialog = CalculadoraPesoDialog(self.root, pesos_existentes)
        self.root.wait_window(dialog)
        
        if dialog.result is not None:
            nuevos_pesos, nuevo_total = dialog.result
            
            # Actualizar los datos de bultos
            if nuevos_pesos:  # Si hay nuevos pesos
                self.bultos_data[codigo] = nuevos_pesos
            else:  # Si se eliminaron todos los pesos, eliminar la entrada
                self.bultos_data.pop(codigo, None)
            
            # Actualizar la fila en el treeview manteniendo la posición
            self._actualizar_fila_articulo(item, codigo, descripcion, nuevos_pesos)
            
            # Actualizar el peso total
            self._actualizar_peso_total()
            
            self._mostrar_info("Artículo actualizado", "Los pesos del artículo se han actualizado correctamente.")

    def _actualizar_fila_articulo(self, item, codigo, descripcion, pesos_bultos):
        """Actualiza una fila del artículo con los nuevos pesos manteniendo la posición"""
        # Calcular cantidad de bultos y peso total
        cantidad_bultos = len(pesos_bultos)
        peso_total = sum(pesos_bultos.values())
        
        # Crear lista de valores para el treeview
        valores = [
            str(cantidad_bultos),
            descripcion,
            f"{peso_total:.2f}"
        ]
        
        # Agregar pesos individuales de bultos
        for i in range(1, 201):  # Para hasta 200 bultos
            peso = pesos_bultos.get(str(i), 0.0)
            valores.append(f"{peso:.2f}" if peso > 0 else "")
        
        # Actualizar el item en el treeview manteniendo tags y posición
        tags = self.tree.item(item, 'tags')
        self.tree.item(item, values=valores, tags=tags)
    
    

    def _actualizar_fila_completa(self, item, codigo, values):
        """Actualiza toda la fila con la nueva estructura de bultos"""
        # Limpiar todos los bultos en la visualización
        for i in range(4, len(values)):
            values[i] = ""
        
        if codigo in self.bultos_data and self.bultos_data[codigo]:
            # Calcular nuevos totales
            pesos_validos = list(self.bultos_data[codigo].values())
            total_peso = sum(pesos_validos)
            cantidad_bultos = len(pesos_validos)
            
            values[0] = str(cantidad_bultos)
            values[2] = f"{total_peso:.2f}"
            
            # Mostrar bultos en sus nuevas posiciones (1, 2, 3...)
            for bulto, peso in sorted(self.bultos_data[codigo].items(), key=lambda x: int(x[0])):
                col_pos = 3 + int(bulto)
                if col_pos < len(values):
                    values[col_pos] = f"{peso:.2f}"
        else:
            # Si no quedan bultos, resetear valores
            values[0] = "0"
            values[2] = "0.00"

    

    

    def _validar_entrada_peso(self, valor: str) -> bool:
        """Valida que la entrada sea un número decimal positivo o vacío"""
        if not valor:  # Permitir campo vacío
            return True
        try:
            # Verificar que sea un número positivo
            num = float(valor)
            return num >= 0
        except ValueError:
            return False
        
    

    


    

    def _crear_controles_articulos(self, parent):
        """Crea los controles usando pack()"""
        control_frame = ttk.Frame(parent, padding=(10, 15, 10, 10))
        control_frame.pack(fill=tk.X, pady=(0, 5))
        
        # Frame para tipo de artículo
        tipo_frame = ttk.Frame(control_frame)
        tipo_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
            
        ttk.Label(
            tipo_frame,
            text="Tipo de Artículo:",
            font=('Segoe UI', 15, 'bold')
        ).pack(side=tk.LEFT, padx=(0, 10))
            
        self.tipo_var = tk.StringVar()
        self.tipo_combobox = ttk.Combobox(
            tipo_frame,
            width=20,
            textvariable=self.tipo_var,
            values=["Importado", "Nacional"],  # Valores fijos según la imagen
            state="readonly")
        self.tipo_combobox.pack(side=tk.LEFT, padx=6)
        self.tipo_combobox.current(0)
            
        # Frame para botones
        button_frame = ttk.Frame(control_frame)
        button_frame.pack(side=tk.RIGHT)
            
        ttk.Button(
            button_frame,
            text="Agregar Artículo",
            command=self._agregar_articulo,
            style='Accent.TButton',
            width=16
        ).pack(side=tk.LEFT, padx=5)
            
        ttk.Button(
            button_frame,
            text="Editar Seleccionado",
            command=self._editar_articulo_seleccionado,  # NUEVO BOTÓN
            style='Secondary.TButton',
            width=18
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            button_frame,
            text="Eliminar Seleccionado",
            command=self._eliminar_articulo,
            style='Secondary.TButton',
            width=22
        ).pack(side=tk.LEFT, padx=5)
        
        # Peso total (simplificado)
        self.peso_total_var = tk.StringVar(value="Peso Total: 0.00 kg")
        ttk.Label(
            control_frame,
            textvariable=self.peso_total_var,
            font=('Segoe UI', 11, 'bold'),
            foreground=self.colores['primario']
        ).pack(side=tk.RIGHT, padx=10)

    def _crear_botones_accion(self, parent):
        """Crea los botones de acción principales con opciones de guardado"""
        button_frame = ttk.Frame(parent)
        button_frame.pack(fill=tk.X, pady=(10, 0))
        
        # Frame para opciones de guardado
        guardado_frame = ttk.Frame(button_frame)
        guardado_frame.pack(side=tk.LEFT, expand=True)
        
        ttk.Label(guardado_frame, text="Guardar como:").pack(side=tk.LEFT, padx=5)
        
        # Botones de guardado
        ttk.Button(
            guardado_frame,
            text="General",
            command=lambda: self._guardar_despacho('general'),
            style='Accent.TButton',
            width=10
        ).pack(side=tk.LEFT, padx=2)
        
        ttk.Button(
            guardado_frame,
            text="Detallado",
            command=lambda: self._guardar_despacho('detallado'),
            style='Accent.TButton',
            width=10
        ).pack(side=tk.LEFT, padx=2)
        
        # Resto de botones...
        ttk.Button(
            guardado_frame,
            text="Nuevo Despacho",
            command=self._nuevo_despacho,
            style='Secondary.TButton',
            width=15
        ).pack(side=tk.LEFT, padx=5)
        
        # Botones de exportación - CORREGIDO: Solo un botón para exportar normal
        export_frame = ttk.Frame(button_frame)
        export_frame.pack(side=tk.RIGHT)
        
        # Solo un botón para exportación normal
        ttk.Button(
            export_frame,
            text="Exportar a Excel",
            command=self._exportar_excel,  # Este es el método principal
            style='Accent.TButton',
            width=16
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            export_frame,
            text="Exportar Detallado",
            command=self._exportar_despacho_detallado,
            style='Accent.TButton',
            width=20
        ).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(
            export_frame,
            text="Salir",
            command=self._confirmar_salida,
            style='Secondary.TButton',
            width=15
        ).pack(side=tk.LEFT, padx=5)


    def _limpiar_despacho(self):
        """Elimina todos los artículos del despacho actual"""
        if not hasattr(self, 'tree') or not self.tree.get_children():
            return
        
        respuesta = messagebox.askyesno(
            "Confirmar Limpieza",
            "¿Está seguro que desea eliminar todos los artículos del despacho actual?",
            icon=messagebox.WARNING
        )
        
        if respuesta:
            # Eliminar todos los artículos del Treeview
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            # Limpiar datos de bultos
            self.bultos_data.clear()
            
            # Restablecer totales
            self.peso_total_var.set("Peso Total: 0.00 kg")
            
            self._actualizar_estado("Despacho limpiado, listo para nuevos artículos")

    def _habilitar_interfaz(self, habilitar: bool):
        """Habilita o deshabilita los controles según el estado"""
        widgets = [
            self.tipo_combobox, self.btn_agregar,
            self.btn_guardar, self.btn_exportar,
            self.tejido_entry, self.codigo_tejido_entry
        ]
        
        state = tk.NORMAL if habilitar else tk.DISABLED
        for widget in widgets:
            if widget:
                widget['state'] = state
    
    def _cargar_ultimo_archivo(self):
        """Carga el último archivo Excel usado o busca uno predeterminado en ubicaciones comunes"""
        # 1. Primero intenta cargar el último archivo usado de la configuración
        ultimo_archivo = self.config.get("ultimo_archivo", "")
        
        # Verificar si el archivo existe y es válido
        if ultimo_archivo and os.path.exists(ultimo_archivo):
            try:
                valido, mensaje = self._validar_archivo_excel(ultimo_archivo)
                if valido:
                    self.excel_path = ultimo_archivo
                    self._cargar_datos_excel()
                    self._actualizar_estado(f"Archivo cargado: {os.path.basename(ultimo_archivo)}")
                    return
                else:
                    print(f"Archivo configurado no válido: {mensaje}")
            except Exception as e:
                print(f"Error al cargar último archivo: {str(e)}")
        
        # 2. Si no hay último archivo o no es válido, buscar en ubicaciones comunes
        ubicaciones_comunes = [
            Path.home() / "Desktop",  # Escritorio
            Path.home() / "Documents",  # Documentos
            Path.cwd(),  # Directorio actual
            Path(__file__).parent  # Directorio de la aplicación
        ]
        
        nombres_posibles = [
            "BaseDatosDespachos.xlsx",
            "DatosDespachos.xlsx",
            "DespachosDB.xlsx",
            "DB_Despachos.xlsx"
        ]
        
        for ubicacion in ubicaciones_comunes:
            for nombre in nombres_posibles:
                ruta_archivo = ubicacion / nombre
                if ruta_archivo.exists():
                    try:
                        valido, mensaje = self._validar_archivo_excel(str(ruta_archivo))
                        if valido:
                            self.excel_path = str(ruta_archivo)
                            self.config["ultimo_archivo"] = self.excel_path
                            self._guardar_configuracion()
                            self._cargar_datos_excel()
                            self._actualizar_estado(f"Archivo encontrado: {nombre}")
                            return
                    except Exception as e:
                        continue
        
        # 3. Si no se encontró ningún archivo válido, mostrar diálogo para seleccionar
        self._abrir_archivo_excel()

    def _cargar_datos_excel(self):
        """Carga los datos de clientes y artículos desde el archivo Excel con un rango ampliado"""
        if not self.excel_path:
            return
            
        try:
            # Cargar hoja de clientes (sin cambios)
            self.clientes_df = pd.read_excel(
                self.excel_path,
                sheet_name='Clientes',
                dtype=str,
                usecols=lambda x: x in ['Nombre', 'RIF', 'Teléfono', 'Dirección']
            ).fillna('')
            
            if self.clientes_df.empty:
                raise ValueError("La hoja 'Clientes' está vacía o no contiene datos válidos")
                
            # Cargar hoja de artículos con rango ampliado (hasta 800 filas)
            articulos_df = pd.read_excel(
                self.excel_path,
                sheet_name='ARTICULOS',
                dtype=str,
                header=1,  # Saltar la primera fila que son encabezados generales
                nrows=800  # Ampliar el rango a 800 filas
            ).fillna('')
                
            if articulos_df.empty:
                raise ValueError("La hoja 'ARTICULOS' está vacía o no contiene datos válidos")
                    
            # Procesar los artículos para crear un DataFrame unificado con más columnas
            articulos_list = []
            
            # Mapeo de columnas a tipos de artículos
            column_mapping = [
                # (columnas, tipo)
                ([0, 1], 'Hilado Nacional'),
                ([2, 3], 'Hilado Importado'),
                ([4, 5], 'Hilado Especial'),  # Nueva categoría si existe
                ([6, 7], 'Tejido Importado'),
                ([8, 9], 'Tejido Nacional'),
                ([10, 11], 'Tejido Especial')  # Nueva categoría si existe
            ]
            
            for cols, tipo in column_mapping:
                # Verificar que las columnas existan en el DataFrame
                if max(cols) < len(articulos_df.columns):
                    df_temp = articulos_df.iloc[:, cols].copy()
                    df_temp.columns = ['Codigo', 'Descripcion']
                    df_temp['Tipo'] = tipo
                    df_temp = df_temp[df_temp['Codigo'].str.strip().astype(bool)]  # Filtrar filas vacías
                    articulos_list.append(df_temp)
            
            # Combinar todos los artículos en un solo DataFrame
            self.articulos_df = pd.concat(articulos_list, ignore_index=True).drop_duplicates()
            
            # Actualizar interfaz
            self._actualizar_lista_clientes()
            self._habilitar_interfaz(True)
            
        except Exception as e:
            self._mostrar_error("Error al cargar datos", f"No se pudieron cargar los datos:\n{str(e)}")
            self._registrar_error(e)
            self._habilitar_interfaz(False)
            raise

    def _filtrar_articulos(self, busqueda: str, search_tree=None):
        """Filtra artículos basado en el texto de búsqueda y tipo seleccionado"""
        if search_tree is None:
            search_tree = self.search_tree
        if not search_tree or self.articulos_df.empty:
            return
        
        search_tree.delete(*search_tree.get_children())
        
        tipo = self.tipo_var.get()
        busqueda = busqueda.lower().strip()
        
        # Filtrar por tipo (Nacional o Importado)
        if tipo == "Nacional":
            mask = self.articulos_df['Tipo'].isin(['Hilado Nacional', 'Tejido Nacional'])
        else:  # "Importado"
            mask = self.articulos_df['Tipo'].isin(['Hilado Importado', 'Tejido Importado'])
        
        df_filtrado = self.articulos_df[mask].copy()
        
        # Aplicar filtro de búsqueda si existe
        if busqueda:
            mask = (
                df_filtrado['Codigo'].str.lower().str.contains(busqueda) | 
                df_filtrado['Descripcion'].str.lower().str.contains(busqueda)
            )
            df_filtrado = df_filtrado[mask]
        
        # Limitar resultados para mejor rendimiento
        df_filtrado = df_filtrado.head(200)
        
        # Insertar datos en el Treeview
        for i, (_, row) in enumerate(df_filtrado.iterrows(), 1):
            self.search_tree.insert(
                '', 
                tk.END,
                values=(row['Codigo'], row['Descripcion'], ""),
                tags=('oddrow' if i % 2 == 1 else 'evenrow')
            )

    def _actualizar_lista_clientes(self):
        """Actualiza la lista de clientes en el Listbox"""
        if not hasattr(self, 'clientes_listbox') or self.clientes_df.empty:
            return
            
        self.clientes_listbox.delete(0, tk.END)
        
        nombre_col = self.config["columnas_clientes"][0]
        rif_col = self.config["columnas_clientes"][1]
        
        # Ordenar y mostrar clientes
        clientes_ordenados = self.clientes_df.sort_values(by=[nombre_col, rif_col])
        
        for _, row in clientes_ordenados.iterrows():
            nombre = str(row[nombre_col]).strip()
            rif = str(row[rif_col]).strip()
            
            if nombre and rif:
                self.clientes_listbox.insert(tk.END, f"{nombre} - {rif}")
        
        self._actualizar_estado(f"Clientes cargados: {len(clientes_ordenados)}")

    def _buscar_cliente(self):
        """Busca clientes según el texto ingresado"""
        if not hasattr(self, 'cliente_search'):
            return
            
        busqueda = self.cliente_search.get().strip().lower()
        self.clientes_listbox.delete(0, tk.END)
        
        if not busqueda:
            self._actualizar_lista_clientes()
            return
            
        if not self.clientes_df.empty:
            nombre_col = self.config["columnas_clientes"][0]
            rif_col = self.config["columnas_clientes"][1]
            
            # Búsqueda flexible (coincidencias parciales)
            mask = (
                self.clientes_df[nombre_col].astype(str).str.lower().str.contains(busqueda) | 
                self.clientes_df[rif_col].astype(str).str.lower().str.contains(busqueda)
            )  # <-- Paréntesis de cierre correctamente ubicado

            resultados = self.clientes_df[mask].sort_values(by=[nombre_col, rif_col])

            for _, row in resultados.iterrows():
                nombre = str(row[nombre_col]).strip()
                rif = str(row[rif_col]).strip()
                self.clientes_listbox.insert(tk.END, f"{nombre} - {rif}")
                self._actualizar_estado(f"Clientes encontrados: {len(resultados)}")

    def _seleccionar_cliente(self, event):
        """Selecciona un cliente de la lista y muestra sus datos (versión simplificada)"""
        if not self.clientes_listbox.curselection():
            return
            
        seleccion = self.clientes_listbox.get(self.clientes_listbox.curselection())
        rif = seleccion.split(" - ")[-1].strip()
        
        if not self.clientes_df.empty:
            rif_col = self.config["columnas_clientes"][1]
            
            try:
                cliente = self.clientes_df[self.clientes_df[rif_col] == rif].iloc[0].to_dict()
                
                self.cliente_actual = {
                    "Nombre": cliente.get(self.config["columnas_clientes"][0], ""),
                    "RIF": cliente.get(self.config["columnas_clientes"][1], ""),
                    "Teléfono": cliente.get(self.config["columnas_clientes"][2], ""),
                    "Dirección": cliente.get(self.config["columnas_clientes"][3], "")
                }
                
                # Actualizar campos
                for attr, key in [
                    ('cliente_nombre', 'Nombre'),
                    ('cliente_rif', 'RIF'),
                    ('cliente_telefono', 'Teléfono'),
                    ('cliente_direccion', 'Dirección')
                ]:
                    widget = getattr(self, attr)
                    widget.delete(0, tk.END)
                    widget.insert(0, self.cliente_actual[key])
                
                self._habilitar_interfaz(True)
                self._actualizar_estado(f"Cliente seleccionado: {self.cliente_actual['Nombre']}")
                
            except IndexError:
                self._mostrar_error("Error", "No se encontraron datos para el cliente seleccionado")
            except Exception as e:
                self._mostrar_error("Error", f"Error al cargar datos del cliente:\n{str(e)}")
                self._registrar_error(e)

    
    def _seleccion_articulo_busqueda(self, peso_var: tk.StringVar):
        """Actualiza el campo de peso cuando se selecciona un artículo en la búsqueda"""
        try:
            selected = self._obtener_articulo_seleccionado_busqueda()
            if selected:
                # Aquí puedes agregar lógica para actualizar el peso_var si es necesario
                # Por ejemplo, si los artículos tienen pesos predefinidos:
                # peso_var.set("1.00")  # Peso por defecto
                pass
        except Exception as e:
            self._mostrar_error("Error", f"No se pudo procesar la selección:\n{str(e)}")
            
    def _agregar_articulo(self):
        """Muestra diálogo para buscar y agregar artículos con interfaz mejorada"""
        if not self.cliente_actual:
            self._mostrar_advertencia("Seleccione un cliente primero")
            return
            
        if self.articulos_df.empty:
            self._mostrar_advertencia("No hay datos de artículos cargados. Verifique el archivo Excel.")
            return
        
        # Variables para controlar si hay datos sin guardar
        self.datos_sin_guardar = False
        self.cantidad_temp = "1"
        self.peso_temp = ""
        self.articulo_seleccionado_temp = None
        
        # Crear diálogo de búsqueda
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Agregar Artículo ({self.tipo_var.get()})")
        dialog.geometry("1000x700")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Configurar grid principal
        dialog.grid_rowconfigure(1, weight=1)
        dialog.grid_columnconfigure(0, weight=1)

        # Frame de búsqueda
        search_frame = ttk.Frame(dialog, padding=10)
        search_frame.grid(row=0, column=0, sticky='ew')
        search_frame.grid_columnconfigure(1, weight=1)

        ttk.Label(
            search_frame,
            text="Buscar Artículo:",
            font=('Segoe UI', 10, 'bold')
        ).grid(row=0, column=0, padx=(0, 10), sticky='w')

        search_var = tk.StringVar()
        search_entry = ttk.Entry(
            search_frame,
            textvariable=search_var,
            width=50
        )
        search_entry.grid(row=0, column=1, padx=5, sticky='ew')
        search_entry.focus()
        
        # Frame para resultados
        results_frame = ttk.Frame(dialog)
        results_frame.grid(row=1, column=0, sticky='nsew', padx=10, pady=10)
        results_frame.grid_rowconfigure(0, weight=1)
        results_frame.grid_columnconfigure(0, weight=1)

        # Treeview para resultados
        columns = ["Código", "Descripción", "Peso Unitario"]
        self.search_tree = ttk.Treeview(
            results_frame,
            columns=columns,
            show='headings',
            height=20,
            selectmode='browse',
            style='Custom.Treeview'
        )
        self.search_tree.tag_configure('selected', background='#0078d7', foreground='white')

        # Configurar columnas
        col_widths = {'Código': 200, 'Descripción': 500, 'Peso Unitario': 100}
        for col in columns:
            self.search_tree.heading(col, text=col)
            self.search_tree.column(col, width=col_widths.get(col, 150), 
                                anchor=tk.W if col != 'Peso Unitario' else tk.CENTER)

        # Scrollbars
        y_scroll = ttk.Scrollbar(results_frame, orient=tk.VERTICAL, command=self.search_tree.yview)
        x_scroll = ttk.Scrollbar(results_frame, orient=tk.HORIZONTAL, command=self.search_tree.xview)
        self.search_tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

        # Posicionamiento
        self.search_tree.grid(row=0, column=0, sticky='nsew')
        y_scroll.grid(row=0, column=1, sticky='ns')
        x_scroll.grid(row=1, column=0, sticky='ew')

        # Frame de controles
        control_frame = ttk.Frame(dialog, padding=10)
        control_frame.grid(row=2, column=0, sticky='ew')
        control_frame.grid_columnconfigure(1, weight=1)
        control_frame.grid_columnconfigure(3, weight=1)
        control_frame.grid_columnconfigure(5, weight=1)

        # Campos de entrada
        cantidad_var = tk.StringVar(value="1")
        peso_var = tk.StringVar()

        # Configuración de widgets
        ttk.Label(control_frame, text="Cantidad:").grid(row=0, column=0, padx=(0, 5), sticky='e')
        cantidad_entry = ttk.Entry(
            control_frame,
            textvariable=cantidad_var,
            width=8,
            validate="key",
            validatecommand=(dialog.register(self._validar_entero_positivo), '%P')
        )
        cantidad_entry.grid(row=0, column=1, padx=5, sticky='w')

        ttk.Label(control_frame, text="Peso Total (kg):").grid(row=0, column=2, padx=(10, 5), sticky='e')
        peso_entry = ttk.Entry(
            control_frame,
            textvariable=peso_var,
            width=10,
            validate="key",
            validatecommand=(dialog.register(self._validar_decimal_positivo), '%P')
        )
        peso_entry.grid(row=0, column=3, padx=5, sticky='w')

        ttk.Button(
            control_frame,
            text="Calcular Peso",
            command=lambda: self._calcular_peso_bultos(cantidad_var, peso_var),
            style='Secondary.TButton',
            width=12
        ).grid(row=0, column=4, padx=5)

        # En el frame de controles, agregar el nuevo botón junto al de calcular peso
        ttk.Button(
            control_frame,
            text="Calcular Metro",
            command=lambda: self._calcular_metros_bultos(cantidad_var, peso_var),
            style='Secondary.TButton',
            width=15
        ).grid(row=0, column=5, padx=5)  # Ajustar posición según diseño

        # Frame de botones del diálogo
        dialog_button_frame = ttk.Frame(dialog, padding=10)
        dialog_button_frame.grid(row=3, column=0, sticky='e')

        def agregar_y_limpiar():
            self._agregar_articulo_desde_busqueda(  # ← CORREGIDO
                cantidad_var.get(),
                peso_var.get())
            # Limpiar campos después de agregar
            cantidad_var.set("1")
            peso_var.set("")
            for item in self.search_tree.get_children():
                self.search_tree.item(item, tags=())
            self.datos_sin_guardar = False

        def marcar_datos_sin_guardar(*args):
            self.datos_sin_guardar = True
            self.cantidad_temp = cantidad_var.get()
            self.peso_temp = peso_var.get()
            selected = self.search_tree.selection()
            if selected:
                self.articulo_seleccionado_temp = self.search_tree.item(selected[0])['values'][:2]

        ttk.Button(
            dialog_button_frame,
            text="Agregar al Despacho",
            command=agregar_y_limpiar,
            style='Accent.TButton'
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(
            dialog_button_frame,
            text="Cancelar",
            command=dialog.destroy,
            style='Secondary.TButton'
        ).pack(side=tk.LEFT)

        # Configuración de eventos - MODIFICACIÓN PRINCIPAL AQUÍ
        def on_tree_select(event=None):
            selected = self.search_tree.selection()
            if selected:
                item = self.search_tree.item(selected[0])
                self.search_tree.tag_configure('selected', background='#0078d7', foreground='white')
                for i in self.search_tree.get_children():
                    self.search_tree.item(i, tags=())
                self.search_tree.item(selected[0], tags=('selected',))
                cantidad_entry.focus()

        self.search_tree.bind('<<TreeviewSelect>>', on_tree_select)
        # Solo permitir selección con click simple, no agregar automáticamente
        self.search_tree.bind('<Button-1>', lambda e: on_tree_select(e))
        
        search_entry.bind('<Return>', lambda e: self._filtrar_articulos(search_var.get(), self.search_tree))
        cantidad_entry.bind('<Return>', lambda e: peso_entry.focus())
        peso_entry.bind('<Return>', lambda e: agregar_y_limpiar())

        # Cargar datos iniciales
        self._filtrar_articulos("", self.search_tree)

        def on_close():
            if self.datos_sin_guardar:
                respuesta = messagebox.askyesno(
                    "Artículo no agregado",
                    "Tiene un artículo listo para agregar pero no lo ha guardado.\n\n"
                    "¿Desea agregarlo ahora?",
                    parent=dialog)
                if respuesta:
                    agregar_y_limpiar()
            dialog.destroy()
        
        dialog.protocol("WM_DELETE_WINDOW", on_close)

        # Configurar seguimiento de cambios
        cantidad_var.trace_add('write', marcar_datos_sin_guardar)
        peso_var.trace_add('write', marcar_datos_sin_guardar)
    
    def _validar_entero_positivo(self, valor: str) -> bool:
        """Valida que la entrada sea un entero positivo"""
        if not valor:
            return True
        return valor.isdigit() and int(valor) > 0

    def _validar_decimal_positivo(self, valor: str) -> bool:
        """Valida que la entrada sea un decimal positivo"""
        if not valor:
            return True
        try:
            return float(valor) > 0
        except ValueError:
            return False

    def _calcular_peso_bultos(self, cantidad_var: tk.StringVar, peso_var: tk.StringVar):
        """Calcula el peso total para múltiples bultos (versión optimizada)"""
        try:
            dialog = CalculadoraPesoDialog(self.root)
            self.root.wait_window(dialog)
            
            if dialog.result is not None:
                bultos_data, total = dialog.result
                
                # Actualizar el campo de peso total
                peso_var.set(f"{total:.2f}")
                
                # Actualizar la cantidad con el número REAL de bultos con peso
                cantidad_var.set(str(len(bultos_data)))
                
                # Guardar los pesos de bultos
                selected = self._obtener_articulo_seleccionado_busqueda()
                if selected:
                    codigo_articulo = selected[0]
                    self.bultos_data[codigo_articulo] = {str(k): v for k, v in bultos_data.items()}
                    
        except Exception as e:
            self._mostrar_error("Error", f"No se pudo calcular pesos:\n{str(e)}")

    def _calcular_metros_bultos(self, cantidad_var, peso_var):
        """Calcula los metros totales para múltiples bultos"""
        try:
            dialog = CalculadoraMetrosDialog(self.root)
            self.root.wait_window(dialog)
            
            if dialog.result is not None:
                bultos_data, total = dialog.result
                peso_var.set(f"{total:.2f}")
                cantidad_var.set(str(len(bultos_data)))
                
                # Marcar que este artículo usa metros
                selected = self._obtener_articulo_seleccionado_busqueda()
                if selected:
                    codigo_articulo = selected[0]
                    self.bultos_data[codigo_articulo] = {str(k): v for k, v in bultos_data.items()}
                    self.bultos_data[codigo_articulo]['_unidad'] = 'mts'  # Marcador especial
                    
        except Exception as e:
            self._mostrar_error("Error", f"No se pudo calcular metros:\n{str(e)}")

    def _obtener_articulo_seleccionado_busqueda(self) -> Optional[Tuple[str, str]]:
        """Obtiene el artículo seleccionado en el diálogo de búsqueda"""
        if not hasattr(self, 'search_tree'):
            return None
        
        seleccion = self.search_tree.selection()
        if not seleccion:
            self._mostrar_advertencia("Seleccione un artículo de la lista")
            return None
            
        item = self.search_tree.item(seleccion[0])
        return tuple(item['values'][:2])  # (Código, Descripción)

    def _agregar_articulo_desde_busqueda(self, cantidad: str, peso_total: str):
        """Agrega un artículo desde el diálogo de búsqueda - Versión corregida"""
        try:
            selected = self._obtener_articulo_seleccionado_busqueda()
            if not selected:
                return
                
            codigo_articulo, descripcion = selected
            
            # Validar y convertir cantidad
            try:
                cantidad_int = int(cantidad) if cantidad else 1
            except ValueError:
                cantidad_int = 1
            
            # Validar y convertir peso
            try:
                peso_float = float(peso_total) if peso_total else 0.0
            except ValueError:
                peso_float = 0.0
            
            # CALCULAR CANTIDAD REAL DE BULTOS CON PESO
            cantidad_real = 0
            if codigo_articulo in self.bultos_data and self.bultos_data[codigo_articulo]:
                # Contar solo bultos con peso > 0 (excluyendo marcadores especiales)
                bultos_data = self.bultos_data[codigo_articulo]
                cantidad_real = len([p for p in bultos_data.values() 
                                if isinstance(p, (int, float)) and p > 0 and not isinstance(p, str)])
            else:
                # Si no hay bultos detallados, usar la cantidad ingresada
                cantidad_real = cantidad_int
                # Crear entrada básica en bultos_data si no existe
                if codigo_articulo not in self.bultos_data:
                    self.bultos_data[codigo_articulo] = {'1': peso_float}
            
            # Crear valores para el Treeview
            valores = [
                str(cantidad_real),  # Usar la cantidad real de bultos
                f"{codigo_articulo} - {descripcion}",
                f"{peso_float:.2f}"
            ]
            
            # Agregar pesos individuales de bultos
            if codigo_articulo in self.bultos_data:
                for i in range(1, 201):
                    peso = self.bultos_data[codigo_articulo].get(str(i), 0.0)
                    valores.append(f"{peso:.2f}" if isinstance(peso, (int, float)) and peso > 0 else "")
            else:
                for i in range(1, 201):
                    valores.append("")
            
            # Insertar en el Treeview
            tags = ('evenrow',) if len(self.tree.get_children()) % 2 == 0 else ('oddrow',)
            self.tree.insert('', tk.END, values=valores, tags=tags)
            
            self._actualizar_peso_total()

            # ALMACENAR EN CACHÉ LOS DATOS DE BULTOS
            if codigo_articulo in self.bultos_data:
                # Usar el cache_manager correcto según el modo
                if self.modo_combinado:
                    self.cache_manager.set_bultos(codigo_articulo, self.bultos_data[codigo_articulo])
                else:
                    self.registro.cache_manager.set_bultos(codigo_articulo, self.bultos_data[codigo_articulo])
                    
        except Exception as e:
            self._mostrar_error("Error", f"No se pudo agregar el artículo:\n{str(e)}")
            self._registrar_error(e)

    def verificar_cambios_externos(self):
        """Verifica cambios externos en los archivos de registro"""
        try:
            self.actualizar_registros_si_cambiados()
            # Programar la siguiente verificación
            self.root.after(5000, self.verificar_cambios_externos)
        except Exception as e:
            print(f"Error en verificación de cambios: {e}")
            self.root.after(10000, self.verificar_cambios_externos)

    def _verificar_integridad_datos(self):
        """Verifica la integridad de los datos antes de operaciones críticas"""
        try:
            # Verificar cliente
            if not self.cliente_actual or not self.cliente_actual.get('Nombre'):
                return False, "No hay cliente seleccionado"
            
            # Verificar artículos
            if not hasattr(self, 'tree') or not self.tree.get_children():
                return False, "No hay artículos en el despacho"
            
            # Verificar pesos válidos
            for item in self.tree.get_children():
                valores = self.tree.item(item)['values']
                if len(valores) > 2:
                    try:
                        peso = float(valores[2].replace(' kg', '')) if isinstance(valores[2], str) else float(valores[2])
                        if peso <= 0:
                            return False, f"Artículo '{valores[1]}' tiene peso inválido"
                    except (ValueError, TypeError):
                        return False, f"Artículo '{valores[1]}' tiene peso inválido"
            
            return True, "Datos válidos"
        
        except Exception as e:
            return False, f"Error en verificación: {str(e)}"

    def _limpiar_datos_corruptos(self):
        """Limpia datos corruptos o inconsistentes"""
        try:
            # Limpiar bultos_data de entradas inválidas
            claves_a_eliminar = []
            for codigo, bultos in self.bultos_data.items():
                if not isinstance(bultos, dict):
                    claves_a_eliminar.append(codigo)
                    continue
                    
                # Eliminar bultos con pesos inválidos
                bultos_validos = {}
                for bulto_num, peso in bultos.items():
                    if (isinstance(peso, (int, float)) and peso >= 0 and 
                        bulto_num != '_unidad' and str(bulto_num).isdigit()):
                        bultos_validos[bulto_num] = peso
                
                if bultos_validos:
                    self.bultos_data[codigo] = bultos_validos
                else:
                    claves_a_eliminar.append(codigo)
            
            for codigo in claves_a_eliminar:
                del self.bultos_data[codigo]
                
        except Exception as e:
            self._registrar_error(e)


            
    def _mostrar_detalle_articulo(self, event):
        """Muestra un diálogo con los detalles del artículo seleccionado"""
        seleccion = self.tree.selection()
        if not seleccion:
            return
            
        item = self.tree.item(seleccion[0])
        valores = item['values']
        
        dialog = tk.Toplevel(self.root)
        dialog.title(f"Detalles del Artículo: {valores[2]}")
        dialog.geometry("500x400")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Frame principal
        main_frame = ttk.Frame(dialog, padding=15)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Mostrar información básica
        campos = [
            ("Cantidad:", valores[0]),
            ("Tipo:", valores[1]),
            ("Código:", valores[2]),
            ("Descripción:", valores[3]),
            ("Peso Total (kg):", valores[4]),
            ("Observaciones:", valores[5])
        ]
        
        for i, (label, value) in enumerate(campos):
            ttk.Label(main_frame, text=label, font=('Segoe UI', 10, 'bold')).grid(
                row=i, column=0, sticky=tk.W, pady=5, padx=5)
            ttk.Label(main_frame, text=value, font=('Segoe UI', 10)).grid(
                row=i, column=1, sticky=tk.W, pady=5, padx=5)
        
        # Mostrar detalles de bultos si existen
        codigo_articulo = valores[2]
        if codigo_articulo in self.bultos_data:
            ttk.Separator(main_frame).grid(
                row=len(campos), column=0, columnspan=2, pady=10, sticky=tk.EW)
            
            ttk.Label(main_frame, text="Detalles de Bultos:", 
                    font=('Segoe UI', 10, 'bold')).grid(
                row=len(campos)+1, column=0, columnspan=2, sticky=tk.W, pady=5)
            
            for i, (bulto, peso) in enumerate(self.bultos_data[codigo_articulo].items(), start=len(campos)+2):
                ttk.Label(main_frame, text=f"Bulto {bulto}:", font=('Segoe UI', 9)).grid(
                    row=i, column=0, sticky=tk.E, padx=5)
                ttk.Label(main_frame, text=f"{peso} kg", font=('Segoe UI', 9)).grid(
                    row=i, column=1, sticky=tk.W, padx=5)
        
        # Botón para cerrar
        ttk.Button(main_frame, text="Cerrar", command=dialog.destroy,
                style='Accent.TButton').grid(
            row=len(campos)+4, column=0, columnspan=2, pady=10)

    def _mostrar_calculadora_peso(self):
        """Muestra el diálogo para calcular peso de múltiples bultos"""
        try:
            # Verificar si tenemos los widgets necesarios
            if not hasattr(self, 'cantidad_entry'):
                self._mostrar_advertencia("No se puede acceder al campo de cantidad")
                return
                
            cantidad_str = self.cantidad_entry.get().strip()
            
            if not cantidad_str or not cantidad_str.isdigit():
                self._mostrar_advertencia("Ingrese una cantidad válida de bultos primero")
                return
                
            cantidad = int(cantidad_str)
            
            if cantidad <= 0:
                self._mostrar_advertencia("La cantidad debe ser mayor a cero")
                return
                
            # Mostrar el diálogo de cálculo de peso
            dialog = CalculadoraPesoDialog(self.root, cantidad)
            
            # Si se ingresaron valores, actualizar el campo de peso
            if dialog.result:
                pesos, total = dialog.result
                if hasattr(self, 'peso_entry'):
                    self.peso_entry.delete(0, tk.END)
                    self.peso_entry.insert(0, f"{total:.2f}")
                    
        except Exception as e:
            self._mostrar_error("Error en calculadora", f"Ocurrió un error:\n{str(e)}")
            self._registrar_error(e)

    def _validar_entero_positivo(self, valor: str) -> bool:
        """Valida que la entrada sea un entero positivo"""
        if not valor:
            return True
        return valor.isdigit() and int(valor) > 0

    def _actualizar_peso_total(self):
        """Calcula y muestra el peso total de todos los artículos"""
        try:
            total = 0.0
            
            for item in self.tree.get_children():
                try:
                    # El peso total está en la tercera columna (índice 2)
                    peso_str = self.tree.item(item)['values'][2]
                    peso = float(peso_str)
                    total += peso
                except (ValueError, IndexError):
                    continue
            
            self.peso_total_var.set(f"Peso Total: {total:.2f} kg")
            
        except Exception as e:
            self._mostrar_error("Error", f"No se pudo calcular el peso total:\n{str(e)}")
            self._registrar_error(e)

    def _eliminar_articulo(self, event=None):
        """Elimina el artículo seleccionado del despacho"""
        seleccion = self.tree.selection()
        if not seleccion:
            self._mostrar_advertencia("Seleccione al menos un artículo para eliminar")
            return
            
        # Obtener códigos de los artículos a eliminar
        articulos_a_eliminar = []
        for item in seleccion:
            valores = self.tree.item(item)['values']
            if len(valores) > 1:
                descripcion = valores[1]
                codigo = descripcion.split(' - ')[0] if ' - ' in descripcion else ""
                if codigo:
                    articulos_a_eliminar.append(codigo)
        
        if not articulos_a_eliminar:
            self._mostrar_advertencia("No se pudieron identificar los artículos a eliminar")
            return
        
        mensaje = f"¿Eliminar {len(seleccion)} artículo(s) seleccionado(s)?"
        
        if not messagebox.askyesno("Confirmar Eliminación", mensaje):
            return
            
        # Eliminar los items seleccionados y sus datos de bultos
        for item in seleccion:
            valores = self.tree.item(item)['values']
            if len(valores) > 1:
                descripcion = valores[1]
                codigo = descripcion.split(' - ')[0] if ' - ' in descripcion else ""
                # ELIMINACIÓN COMPLETA: Remover completamente de bultos_data
                if codigo in self.bultos_data:
                    del self.bultos_data[codigo]
            self.tree.delete(item)
        
        # Reaplicar colores alternados
        for i, item in enumerate(self.tree.get_children()):
            tags = ('evenrow',) if i % 2 == 0 else ('oddrow',)
            self.tree.item(item, tags=tags)
        
        self._actualizar_peso_total()
        self._mostrar_info("Éxito", f"Se eliminaron {len(seleccion)} artículos")

    def _abrir_archivo_excel(self):
        """Permite al usuario seleccionar un archivo Excel con datos y guarda la ruta"""
        filetypes = [
            ("Archivos Excel", "*.xlsx *.xls"),
            ("Todos los archivos", "*.*")
        ]
        
        filepath = filedialog.askopenfilename(
            title="Seleccionar archivo Excel de datos",
            initialdir=os.path.dirname(self.excel_path) if self.excel_path else os.path.expanduser("~"),
            filetypes=filetypes
        )
        
        if filepath:
            try:
                # Validar el archivo antes de cargarlo
                valido, mensaje = self._validar_archivo_excel(filepath)
                if not valido:
                    raise ValueError(mensaje)
                
                # Si pasa la validación, proceder con la carga
                self.excel_path = filepath
                self.config["ultimo_archivo"] = filepath
                    
                if not self._guardar_configuracion():
                    raise ValueError("No se pudo guardar la configuración")
                    
                self._cargar_datos_excel()
                self._habilitar_interfaz(True)
                    
                self._mostrar_info(
                    "Archivo cargado", 
                    f"Se cargó correctamente:\n{os.path.basename(filepath)}\n\n"
                    "Esta ruta se guardará para futuras sesiones.")
                    
            except Exception as e:
                self._mostrar_error("Error", f"No se pudo cargar el archivo:\n{str(e)}")
                self._registrar_error(e)
                # Intentar cargar cualquier archivo válido disponible
                self._cargar_ultimo_archivo()

    def _validar_despacho(self) -> bool:
        """Valida que el despacho esté completo antes de guardar"""
        errores = []
        
        if not self.cliente_actual:
            errores.append("- Seleccione un cliente primero")
            
        if not hasattr(self, 'tree') or not self.tree.get_children():
            errores.append("- Agregue al menos un artículo al despacho")
        
        # Validar pesos de artículos (manejar tanto strings como números)
        articulos_invalidos = []
        for item in self.tree.get_children():
            try:
                valores = self.tree.item(item)['values']
                peso_valor = valores[2] if len(valores) > 2 else 0
                
                # Convertir a string si es número, luego eliminar 'kg' si existe
                if isinstance(peso_valor, (int, float)):
                    peso_str = str(peso_valor)
                else:
                    peso_str = str(peso_valor).replace(' kg', '')
                
                peso = float(peso_str)
                if peso <= 0:
                    articulos_invalidos.append(valores[1] if len(valores) > 1 else "Artículo sin descripción")
            except (ValueError, IndexError, TypeError):
                articulos_invalidos.append(valores[1] if len(valores) > 1 else "Artículo sin descripción")
        
        if articulos_invalidos:
            errores.append(
                f"- Los siguientes artículos tienen pesos inválidos:\n  "
                f"{', '.join(articulos_invalidos)}")
        
        if errores:
            self._mostrar_error(
                "Error en el despacho",
                "Corrija los siguientes errores antes de guardar:\n\n" +
                "\n".join(errores))
            return False
                
        return True
    

    def _verificar_estado_registro(self):
        """Verifica el estado actual del registro"""
        if self.modo_combinado:
            print("=== MODO COMBINADO ===")
            print(f"Registros cargados: {len(self.registro.registros)}")
            for i, reg in enumerate(self.registro.registros):
                print(f"Registro {i+1}: {reg.archivo_registro}")
                print(f"  Existe: {reg.archivo_registro.exists()}")
                if reg.archivo_registro.exists():
                    print(f"  Tamaño: {reg.archivo_registro.stat().st_size} bytes")
        else:
            print("=== MODO INDIVIDUAL ===")
            print(f"Registro: {self.registro.archivo_registro}")
            print(f"Existe: {self.registro.archivo_registro.exists()}")
            if self.registro.archivo_registro.exists():
                print(f"Tamaño: {self.registro.archivo_registro.stat().st_size} bytes")
                
        # Verificar permisos de escritura
        try:
            if self.modo_combinado:
                # Usar el primer registro para verificar permisos
                test_file = self.registro.registros[0].archivo_registro
            else:
                test_file = self.registro.archivo_registro
                
            with open(test_file, 'a') as f:
                f.write("\n")
            print("✓ Permisos de escritura OK")
        except Exception as e:
            print(f"✗ Error de permisos: {e}")

    def _guardar_despacho(self, metodo='general') -> bool:
        """Guarda el despacho usando el método especificado con manejo robusto de errores"""
        print(f"Iniciando guardado de despacho (método: {metodo})...")
        
        if not self._validar_despacho():
            print("Validación de despacho falló")
            return False
        
        self.despacho_guardado = True
        
        try:
            # PREPARAR DATOS PARA EL REGISTRO JSON - compatible con ambos métodos
            datos_despacho = {
                'cliente': {
                    'nombre': self.cliente_actual.get('Nombre', ''),
                    'rif': self.cliente_actual.get('RIF', ''),
                    'telefono': self.cliente_actual.get('Teléfono', ''),
                    'direccion': self.cliente_actual.get('Dirección', '')
                },
                'articulos': [],
                'peso_total': 0.0,
                'metodo_guardado': metodo  # Añadir información del método usado
            }
            
            # NEW: Include bultos_data for detailed method
            if metodo == 'detallado':
                datos_despacho['bultos_data'] = self.bultos_data.copy()
            
            # Procesar artículos según el método seleccionado
            total_peso = 0.0
            
            for item in self.tree.get_children():
                valores = self.tree.item(item)['values']
                descripcion = valores[1] if len(valores) > 1 else ""
                codigo = descripcion.split(' - ')[0] if ' - ' in descripcion else ""
                
                # Obtener peso
                peso_valor = valores[2] if len(valores) > 2 else 0
                if isinstance(peso_valor, (int, float)):
                    peso = float(peso_valor)
                else:
                    peso = float(str(peso_valor).replace(' kg', ''))
                
                # Obtener cantidad REAL de bultos (CORRECCIÓN)
                try:
                    cantidad = int(valores[0]) if valores[0] else 1
                except (ValueError, TypeError):
                    cantidad = 1
                
                # Estructura base del artículo
                articulo_data = {
                    'codigo': codigo,
                    'descripcion': descripcion,
                    'cantidad': cantidad,  # Cantidad real de bultos
                    'peso_total': peso
                }
                
                # Método detallado: incluir información de bultos individuales
                if metodo == 'detallado' and codigo in self.bultos_data:
                    # Filtrar solo bultos con peso válido
                    bultos_con_peso = {}
                    for bulto_num, peso_bulto in self.bultos_data[codigo].items():
                        if (isinstance(peso_bulto, (int, float)) and 
                            peso_bulto > 0 and bulto_num != '_unidad'):
                            bultos_con_peso[bulto_num] = peso_bulto
                    
                    articulo_data['bultos_detallados'] = bultos_con_peso
                    
                    # También incluir información de unidad si existe
                    if '_unidad' in self.bultos_data[codigo]:
                        articulo_data['unidad'] = self.bultos_data[codigo]['_unidad']
                
                datos_despacho['articulos'].append(articulo_data)
                total_peso += peso
            
            datos_despacho['peso_total'] = total_peso
            
            # 2. GUARDAR EN EL REGISTRO JSON CORRESPONDIENTE
            if self.modo_combinado:
                registro_individual = RegistroDespachos(self.nombre_registro_individual)
            else:
                registro_individual = self.registro
            
            # ELIMINAR VERIFICACIÓN DE DESPACHOS DUPLICADOS - PERMITIR MÚLTIPLES DESPACHOS
            despacho_id = registro_individual.agregar_despacho(datos_despacho)
            
            self.despacho_actual_id = despacho_id
            mensaje = f"Nuevo despacho registrado (ID: {despacho_id})"
            
            # 3. Mostrar confirmación
            self._mostrar_info("Despacho guardado", 
                            f"{mensaje}\n"
                            f"Método: {metodo}\n"
                            "Los datos se han guardado correctamente en el registro.")
            
            # Marcar como guardado para evitar duplicados
            self.despacho_guardado = True
            
            return True

        except Exception as e:
            error_msg = f"No se pudo guardar el despacho:\n{str(e)}"
            print(f"Error completo: {traceback.format_exc()}")
            
            self._mostrar_error("Error al guardar", error_msg)
            
            self._registrar_error(e)
            return False
        
    def _diagnosticar_registro(self):
        """Método para diagnosticar problemas con el registro"""
        print("=== DIAGNÓSTICO DEL REGISTRO ===")
        
        if self.modo_combinado:
            print("MODO COMBINADO")
            for i, registro in enumerate(self.registro.registros):
                print(f"Registro {i+1}: {registro.archivo_registro}")
                print(f"  Existe: {registro.archivo_registro.exists()}")
                if registro.archivo_registro.exists():
                    print(f"  Tamaño: {registro.archivo_registro.stat().st_size} bytes")
                    print(f"  Modificable: {os.access(registro.archivo_registro, os.W_OK)}")
            
            # Verificar directorio usando el primer registro como referencia
            if self.registro.registros:
                directorio = self.registro.registros[0].archivo_registro.parent
                print(f"Directorio: {directorio}")
                print(f"Directorio existe: {directorio.exists()}")
                print(f"Directorio escribible: {os.access(directorio, os.W_OK)}")
        else:
            print("MODO INDIVIDUAL")
            print(f"Registro: {self.registro.archivo_registro}")
            print(f"Existe: {self.registro.archivo_registro.exists()}")
            if self.registro.archivo_registro.exists():
                print(f"Tamaño: {self.registro.archivo_registro.stat().st_size} bytes")
                print(f"Modificable: {os.access(self.registro.archivo_registro, os.W_OK)}")
            
            # Verificar directorio
            directorio = self.registro.archivo_registro.parent
            print(f"Directorio: {directorio}")
            print(f"Directorio existe: {directorio.exists()}")
            print(f"Directorio escribible: {os.access(directorio, os.W_OK)}")
        
    def _preparar_datos_para_registro(self) -> dict:
        """Prepara los datos del despacho para guardar en el registro JSON"""
        datos = {
            'cliente': self.cliente_actual.copy() if self.cliente_actual else {},
            'articulos': [],
            'bultos_data': self.bultos_data.copy(),
            'peso_total': 0.0,
            'fecha_creacion': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        
        # Procesar artículos
        total_peso = 0.0
        for item in self.tree.get_children():
            valores = self.tree.item(item)['values']
            descripcion = valores[1] if len(valores) > 1 else ""
            codigo = descripcion.split(' - ')[0] if ' - ' in descripcion else ""
            
            try:
                peso = float(valores[2].replace(' kg', '')) if len(valores) > 2 else 0.0
            except (ValueError, TypeError):
                peso = 0.0
            
            datos['articulos'].append({
                'codigo': codigo,
                'descripcion': descripcion,
                'cantidad': int(valores[0]) if valores[0] else 1,
                'peso_total': peso
            })
            
            total_peso += peso
        
        datos['peso_total'] = total_peso
        return datos

    def _encontrar_fila_vacia(self, sheet) -> int:
        """Encuentra la primera fila vacía en la hoja"""
        fila = 1
        while sheet.cell(row=fila, column=1).value is not None:
            fila += 1
            if fila > 1000:  # Límite para evitar bucles infinitos
                raise ValueError("No se encontró fila vacía en las primeras 1000 filas")
        return fila

    def _escribir_info_cliente(self, sheet, fila_inicial: int):
        """Escribe la información del cliente en la hoja Excel"""
        if not self.cliente_actual:
            return
            
        # Escribir encabezado
        sheet.cell(row=fila_inicial, column=1, value="INFORMACIÓN DEL CLIENTE").font = Font(
            bold=True, size=12, color="FFFFFF")
        sheet.cell(row=fila_inicial, column=1).fill = PatternFill(
            "solid", fgColor=self.colores['primario'])
        
        # Combinar celdas
        sheet.merge_cells(start_row=fila_inicial, start_column=1, end_row=fila_inicial, end_column=2)
        
        # Escribir datos del cliente
        campos = [
            ("Nombre", self.cliente_actual.get("Nombre", "")),
            ("RIF", self.cliente_actual.get("RIF", "")),
            ("Teléfono", self.cliente_actual.get("Teléfono", "")),
            ("Dirección", self.cliente_actual.get("Dirección", ""))
        ]
        
        for i, (campo, valor) in enumerate(campos, start=1):
            # Etiqueta
            sheet.cell(row=fila_inicial+i, column=1, value=f"{campo}:").font = Font(
                bold=True, color="44546A")
            
            # Valor
            sheet.cell(row=fila_inicial+i, column=2, value=valor).border = Border(
                bottom=Side(style='thin', color="D9D9D9"))
            
            # Aplicar bordes
            for col in [1, 2]:
                sheet.cell(row=fila_inicial+i, column=col).border = Border(
                    left=Side(style='thin', color="D9D9D9"),
                    right=Side(style='thin', color="D9D9D9"),
                    bottom=Side(style='thin', color="D9D9D9"))
                
    def _procesar_articulos(self, mantener_orden_original=False):
        """Procesa artículos agrupándolos por nombre base para exportación detallada"""
        productos = {}
        
        # Recorrer todos los items del treeview
        for item in self.tree.get_children():
            valores = self.tree.item(item)['values']
            
            if not valores or len(valores) < 3:
                continue
                
            descripcion = valores[1]
            peso = float(valores[2].replace(' kg', '')) if isinstance(valores[2], str) else float(valores[2])
            
            # Extraer nombre base (sin código)
            if ' - ' in descripcion:
                partes = descripcion.split(' - ', 1)
                nombre_base = partes[1].strip()
            else:
                nombre_base = descripcion
                
            if nombre_base not in productos:
                productos[nombre_base] = {
                    'descripciones': [],
                    'pesos': []
                }
                
            productos[nombre_base]['descripciones'].append(descripcion)
            productos[nombre_base]['pesos'].append(peso)
        
        # Ordenar por nombre base o mantener orden original
        if mantener_orden_original:
            # Crear lista ordenada según aparición en treeview
            ordenados = []
            seen_bases = set()
            
            for item in self.tree.get_children():
                valores = self.tree.item(item)['values']
                if valores and len(valores) > 1:
                    descripcion = valores[1]
                    if ' - ' in descripcion:
                        nombre_base = descripcion.split(' - ', 1)[1].strip()
                    else:
                        nombre_base = descripcion
                        
                    if nombre_base not in seen_bases:
                        seen_bases.add(nombre_base)
                        ordenados.append((nombre_base, productos.get(nombre_base, {'descripciones': [], 'pesos': []})))
            
            return ordenados
        else:
            # Orden alfabético
            return sorted(productos.items())


    def _procesar_articulos_para_exportacion(self, metodo='detallado'):
        """Procesa artículos de manera consistente para cualquier método de exportación - CORREGIDO"""
        productos = {}
        
        # Recorrer TODOS los items del treeview UNA SOLA VEZ
        for item in self.tree.get_children():
            valores = self.tree.item(item)['values']
            
            # Validación exhaustiva de datos
            if not valores or len(valores) < 3:
                print(f"ADVERTENCIA: Artículo con datos incompletos: {valores}")
                continue
                
            descripcion = valores[1]
            cantidad = valores[0]
            
            try:
                peso = float(valores[2].replace(' kg', '')) if isinstance(valores[2], str) else float(valores[2])
            except (ValueError, TypeError):
                print(f"ADVERTENCIA: Peso inválido en artículo: {descripcion}")
                peso = 0.0
            
            # Clave única que garantice que ningún artículo se pierda
            clave = f"{descripcion}_{item}"  # Incluir el ID del item para unicidad
            
            if clave not in productos:
                productos[clave] = {
                    'descripcion': descripcion,
                    'cantidad': cantidad,
                    'peso': peso,
                    'item_id': item,  # Guardar referencia al item original
                    'bultos_data': None
                }
            
            # Para método detallado, obtener datos de bultos
            if metodo == 'detallado':
                codigo = self._extraer_codigo_seguro(descripcion)
                if codigo and codigo in self.bultos_data:
                    productos[clave]['bultos_data'] = self.bultos_data[codigo].copy()
        
        return productos
    

    def _extraer_codigo_seguro(self, descripcion):
        """Extrae código de manera robusta, con múltiples estrategias"""
        if not descripcion:
            return None
        
        # Estrategia 1: Buscar patrón "código - descripción"
        if ' - ' in descripcion:
            partes = descripcion.split(' - ')
            posible_codigo = partes[0].strip()
            # Validar que sea un código probable (sin espacios, alfanumérico)
            if len(posible_codigo.split()) == 1 and len(posible_codigo) > 0:
                return posible_codigo
        
        # Estrategia 2: Buscar al inicio de la descripción
        palabras = descripcion.split()
        if palabras and len(palabras[0]) <= 20:  # Códigos normalmente cortos
            return palabras[0]
        
        # Estrategia 3: Buscar en bultos_data por coincidencia parcial
        for codigo_existente in self.bultos_data.keys():
            if codigo_existente in descripcion:
                return codigo_existente
        
        return None

    def _validar_integridad_exportacion(self):
        """Valida que todos los artículos visibles sean procesados correctamente"""
        articulos_treeview = []
        articulos_procesados = []
        
        # Recoger todos los artículos del treeview
        for item in self.tree.get_children():
            valores = self.tree.item(item)['values']
            if valores and len(valores) > 1:
                articulos_treeview.append(valores[1])  # Descripción
        
        # Procesar artículos
        productos = self._procesar_articulos_para_exportacion('detallado')
        for producto in productos.values():
            articulos_procesados.append(producto['descripcion'])
        
        # Verificar coincidencia
        faltantes = set(articulos_treeview) - set(articulos_procesados)
        
        if faltantes:
            print(f"ERROR: Artículos faltantes en exportación: {faltantes}")
            return False, list(faltantes)
        
        return True, []

    def _validar_datos_exportacion(self):
        """Valida que haya datos para exportar"""
        if not hasattr(self, 'tree') or not self.tree.get_children():
            self._mostrar_advertencia("No hay artículos para exportar")
            return False
        
        # Verificar que al menos un artículo tenga peso válido
        for item in self.tree.get_children():
            valores = self.tree.item(item)['values']
            if len(valores) > 2:
                try:
                    peso = float(valores[2].replace(' kg', '')) if isinstance(valores[2], str) else float(valores[2])
                    if peso > 0:
                        return True
                except (ValueError, TypeError):
                    continue
        
        self._mostrar_advertencia("Todos los artículos tienen peso cero o inválido")
        return False

                
    def _procesar_articulos_para_excel_normal(self):
        """Procesa artículos para Excel normal con misma lógica que el detallado"""
        productos = {}
        
        # Recorrer artículos en orden inverso (de arriba hacia abajo)
        items = list(self.tree.get_children())
        for item in reversed(items):  # INVERTIR EL ORDEN
            valores = self.tree.item(item)['values']
            
            if not valores or len(valores) < 3:
                continue
                
            descripcion = valores[1]
            cantidad = valores[0]
            
            try:
                peso = float(valores[2].replace(' kg', '')) if isinstance(valores[2], str) else float(valores[2])
            except (ValueError, TypeError):
                peso = 0.0
            
            # Extraer nombre base y color (misma lógica que Excel detallado)
            if ' - ' in descripcion:
                partes = descripcion.split(' - ', 1)
                nombre_base = partes[1].strip()
            else:
                nombre_base = descripcion
            
            # Identificar color
            colores_conocidos = ["BLANCO", "OSCURO", "PASTEL", "ESPECIAL", "MELANGE", 
                                "NEGRO", "ROJO", "AZUL", "VERDE", "GRIS", "BEIGE", 
                                "CREMA", "AMARILLO", "POPULAR", "COLORES"]
            color = "BLANCO"
            
            for parte in reversed(partes):
                parte_upper = parte.upper()
                if any(c in parte_upper for c in colores_conocidos):
                    color = next((c for c in colores_conocidos if c in parte_upper), "BLANCO")
                    break
            
            # Clave única para agrupación
            clave = f"{nombre_base}_{color}"
            
            if clave not in productos:
                productos[clave] = {
                    'nombre_base': nombre_base,
                    'color': color,
                    'descripciones': [],
                    'cantidades': [],
                    'pesos': []
                }
            
            productos[clave]['descripciones'].append(descripcion)
            productos[clave]['cantidades'].append(cantidad)
            productos[clave]['pesos'].append(peso)
        
        return productos

    # Modificaciones en el método _exportar_excel
    def _exportar_excel(self):
        """Exporta el despacho en formato profesional - Versión corregida"""
        try:
            # Primero validar y guardar el despacho
            if not self.despacho_guardado:
                if not self._guardar_despacho('general'):
                    return False

            # Configurar nombre del archivo con verificación de duplicados
            hoy = datetime.now()
            fecha_str = hoy.strftime("%d-%m-%Y")
            cliente_nombre = self.cliente_actual.get('Nombre', 'DESCONOCIDO').replace(' ', '_')
            default_filename = f"DESPACHO_{cliente_nombre}_{fecha_str}.xlsx"
            
            # Obtener directorio del escritorio para el diálogo de guardado
            desktop_path = Path.home() / "Desktop"
            if not desktop_path.exists():
                desktop_path = Path.home() / "Escritorio"
            
            # Generar nombre único que evite sobrescritura
            nombre_unico = self._generar_nombre_archivo_unico(cliente_nombre, Path(default_filename))
            
            filepath = filedialog.asksaveasfilename(
                title="Guardar Despacho como...",
                defaultextension=".xlsx",
                filetypes=[("Archivo Excel", "*.xlsx")],
                initialdir=str(desktop_path),  # Directorio inicial
                initialfile=nombre_unico.name  # Solo el nombre del archivo, no la ruta completa
            )

            if not filepath:
                return False

            # Determinar método de guardado basado en si hay datos de bultos
            metodo_guardado = 'detallado' if self.bultos_data else 'general'
            
            # Obtener información del método de guardado
            metodo_guardado = 'general'  # Por defecto
            if self.despacho_actual_id:
                # Cargar el despacho para verificar cómo fue guardado
                despacho = self.registro.obtener_despacho(self.despacho_actual_id)
                if despacho and 'metodo_guardado' in despacho:
                    metodo_guardado = despacho['metodo_guardado']

            # Determinar si todos los artículos son en metros
            solo_metros = True
            for item in self.tree.get_children():
                valores = self.tree.item(item, 'values')
                descripcion = valores[1] if len(valores) > 1 else ""
                codigo = descripcion.split(' - ')[0] if ' - ' in descripcion else ""
                if codigo in self.bultos_data and '_unidad' in self.bultos_data[codigo]:
                    if self.bultos_data[codigo]['_unidad'] != 'mts':
                        solo_metros = False
                        break
                else:
                    solo_metros = False
                    break

            # Configurar texto según unidad de medida
            unidad_texto = "METROS" if solo_metros else "KILOS"
            total_texto = f"TOTAL {unidad_texto}" if solo_metros else "TOTAL KILOS"

            # USAR LA MISMA LÓGICA DE ORGANIZACIÓN QUE EL EXCEL DETALLADO
            colores_conocidos = ["BLANCO", "OSCURO", "PASTEL", "ESPECIAL", "MELANGE", 
                                "NEGRO", "ROJO", "AZUL", "VERDE", "GRIS", "BEIGE", 
                                "CREMA", "AMARILLO", "POPULAR", "COLORES"]
            
            # Paso 1: Procesar todos los artículos manteniendo el orden original
            articulos_orden_original = []
            for item in self.tree.get_children():
                valores = self.tree.item(item, 'values')
                if not valores or len(valores) < 2:
                    continue
                
                descripcion = valores[1]
                cantidad = valores[0]
                peso = valores[2] if len(valores) > 2 else "0.0"
                
                # Extraer código, nombre base y color (misma lógica que Excel detallado)
                if ' - ' in descripcion:
                    partes = descripcion.split(' - ', 1)
                    codigo = partes[0].strip()
                    nombre_base = partes[1].strip()
                else:
                    codigo = ""
                    nombre_base = descripcion
                
                # Identificar color
                color = "BLANCO"  # Valor por defecto
                for parte in reversed(descripcion.split(' - ')):
                    parte_upper = parte.upper()
                    if any(c in parte_upper for c in colores_conocidos):
                        color = next((c for c in colores_conocidos if c in parte_upper), "BLANCO")
                        break
                
                # Limpiar nombre base de términos de color
                for color_term in colores_conocidos:
                    if nombre_base.upper().endswith(color_term):
                        nombre_base = nombre_base[:-len(color_term)].strip()
                        break
                
                # Guardar artículo con información completa
                articulos_orden_original.append({
                    'descripcion': descripcion,
                    'nombre_base': nombre_base,
                    'color': color,
                    'codigo': codigo,
                    'cantidad': cantidad,
                    'peso': peso,
                    'orden_original': len(articulos_orden_original)  # Mantener orden original
                })

            # Paso 2: Agrupar por nombre_base y color (misma lógica que Excel detallado)
            articulos_agrupados = {}
            for articulo in articulos_orden_original:
                clave = f"{articulo['nombre_base']}_{articulo['color']}"
                if clave not in articulos_agrupados:
                    articulos_agrupados[clave] = []
                articulos_agrupados[clave].append(articulo)

            # Paso 3: Ordenar los grupos por su primera aparición (mantener orden original)
            grupos_ordenados = []
            claves_ordenadas = []
            
            # Primero recorrer todos los artículos para determinar el orden de los grupos
            for articulo in articulos_orden_original:
                clave = f"{articulo['nombre_base']}_{articulo['color']}"
                if clave not in claves_ordenadas:
                    claves_ordenadas.append(clave)
            
            # Ahora ordenar los grupos según el orden de aparición
            for clave in claves_ordenadas:
                if clave in articulos_agrupados:
                    grupos_ordenados.append((clave, articulos_agrupados[clave]))

            # Crear libro de Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "GUIA DE DESPACHO"

            # Configuración de formato general (sin bordes)
            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 60
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12

            # Altura de filas
            for row in range(1, 51):
                ws.row_dimensions[row].height = 15
            
            # Ajustes específicos de altura para filas clave
            ws.row_dimensions[8].height = 18
            ws.row_dimensions[9].height = 18
            ws.row_dimensions[10].height = 18
            ws.row_dimensions[11].height = 18
            ws.row_dimensions[12].height = 18
            ws.row_dimensions[13].height = 20
            ws.row_dimensions[14].height = 20

            # Cabecera sin bordes
            ws['C2'] = "GUARENAS"
            ws['C2'].font = Font(name='Arial', size=11, bold=True)
            
            ws['D2'] = "FECHA:"
            ws['D2'].font = Font(name='Arial', size=10, bold=True)
            ws['D2'].alignment = Alignment(horizontal='right')
            
            ws['D3'] = hoy.strftime("%d/%m/%Y")
            ws['D3'].font = Font(name='Arial', size=10)
            ws['D3'].alignment = Alignment(horizontal='right')

            # Información del cliente sin bordes
            ws['B8'] = "Nombre o Razon Social:"
            ws['B8'].font = Font(name='Arial', size=10, bold=True)
            
            ws['B9'] = f"{self.cliente_actual.get('Nombre', '')}"
            ws['B9'].font = Font(name='Arial', size=10)
            
            rif_text = f"RIF: {self.cliente_actual.get('RIF', '')}"
            ws['B10'] = rif_text
            ws['B10'].font = Font(name='Arial', size=10, bold=True)
            
            ws['B11'] = "DIRECCION:"
            ws['B11'].font = Font(name='Arial', size=10, bold=True)
            
            if self.cliente_actual.get('Dirección', ''):
                ws['B12'] = self.cliente_actual.get('Dirección', '')
                ws['B12'].font = Font(name='Arial', size=10)
                ws.row_dimensions[12].height = 18

            tel_text = f"TEL: {self.cliente_actual.get('Teléfono', '')}"
            ws['C12'] = tel_text
            ws['C12'].font = Font(name='Arial', size=10, bold=True)

            # Tabla de productos (sin bordes)
            encabezados = ["DESCRIPCIÓN", "BULTOS", unidad_texto]
            ws['B13'] = encabezados[0]
            ws['C13'] = encabezados[1]
            ws['D13'] = encabezados[2]
            
            # Estilo para encabezados (sin bordes)
            for col in ['B', 'C', 'D']:
                cell = ws[f'{col}13']
                cell.font = Font(name='Arial', size=10, bold=True)
                cell.alignment = Alignment(horizontal='center')

            # Escribir productos (en el mismo orden que el Excel detallado)
            fila_actual = 14
            total_bultos = 0
            total_peso = 0.0

            # Escribir artículos agrupados (misma lógica que Excel detallado)
            for grupo_nombre, articulos_grupo in grupos_ordenados:
                # Escribir cada artículo del grupo
                for articulo in articulos_grupo:
                    # Escribir datos
                    ws[f'B{fila_actual}'] = articulo['descripcion']
                    ws[f'B{fila_actual}'].font = Font(name='Arial', size=10)
                    
                    # Usar la cantidad real de bultos
                    try:
                        cantidad_int = int(articulo['cantidad']) if articulo['cantidad'] else 1
                    except ValueError:
                        cantidad_int = 1
                        
                    ws[f'C{fila_actual}'] = cantidad_int
                    ws[f'C{fila_actual}'].font = Font(name='Arial', size=10)
                    ws[f'C{fila_actual}'].alignment = Alignment(horizontal='center')
                    
                    # Convertir peso a número
                    try:
                        peso_float = float(articulo['peso'].replace(' kg', '')) if isinstance(articulo['peso'], str) else float(articulo['peso'])
                    except (ValueError, TypeError):
                        peso_float = 0.0
                        
                    ws[f'D{fila_actual}'] = peso_float
                    ws[f'D{fila_actual}'].font = Font(name='Arial', size=10)
                    ws[f'D{fila_actual}'].number_format = '0.00'
                    ws[f'D{fila_actual}'].alignment = Alignment(horizontal='center')
                    
                    total_bultos += cantidad_int
                    total_peso += peso_float
                    fila_actual += 1

            # ----------------------------
            # TOTALES GENERALES - USAR FÓRMULAS DINÁMICAS
            # ----------------------------
            ws['B43'] = "TOTALES"
            ws['B43'].font = Font(name='Arial', size=11, bold=True)
            
            # Fórmula dinámica para sumar bultos (desde fila 14 hasta la última fila con datos)
            ultima_fila = fila_actual - 1
            bultos_formula = f"=SUM(C14:C{ultima_fila})"
            ws['C43'] = bultos_formula
            ws['C43'].font = Font(name='Arial', size=11, bold=True)
            ws['C43'].alignment = Alignment(horizontal='center')
            
            # Fórmula dinámica para sumar peso/metros
            peso_formula = f"=SUM(D14:D{ultima_fila})"
            ws['D43'] = peso_formula
            ws['D43'].font = Font(name='Arial', size=11, bold=True)
            ws['D43'].number_format = '0.00'
            ws['D43'].alignment = Alignment(horizontal='center')

            # ----------------------------
            # MOTIVO DE TRASLADO EN FILA 45 (FIJO)
            # ----------------------------
            ws['B45'] = "MOTIVO DEL TRASLADO:"
            ws['B45'].font = Font(name='Arial', size=10, bold=True)
            
            # Chofer y C.I. (fila 48)
            ws['B48'] = "CHOFER: ____________________"
            ws['B48'].font = Font(name='Arial', size=10)
            ws['C48'] = "C.I: ____________________"
            ws['C48'].font = Font(name='Arial', size=10)
            
            # Vehículo y Placa (fila 49)
            ws['B49'] = "VEHICULO: ____________________"
            ws['B49'].font = Font(name='Arial', size=10)
            ws['C49'] = "PLACA: ____________________"
            ws['C49'].font = Font(name='Arial', size=10)

            # ----------------------------
            # CONFIGURACIÓN DE IMPRESIÓN (HASTA FILA 50)
            # ----------------------------
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
            
            # Márgenes mínimos posibles (en pulgadas)
            ws.page_margins.left = 0.12    # 3.0 mm (reducido de 3.8)
            ws.page_margins.right = 0.12   # 3.0 mm (reducido de 3.8)
            ws.page_margins.top = 0.16     # 4.0 mm (reducido de 5.0)
            ws.page_margins.bottom = 0.12  # 3.0 mm (reducido de 4.5) - Ajuste clave
            ws.page_margins.header = 0.08  # 2.0 mm 
            ws.page_margins.footer = 0.08  # 2.0 mm

            # Área de impresión forzada
            ws.print_area = f"A1:D50"

            # ----------------------------
            # GUARDAR ARCHIVO
            # ----------------------------
            wb.save(filepath)
            
            # Preguntar si abrir el archivo
            respuesta = messagebox.askyesno(
                "Exportación exitosa",
                f"El despacho se exportó correctamente a:\n{os.path.basename(filepath)}\n\n"
                "¿Desea abrir el archivo ahora?")
            
            if respuesta:
                self._abrir_archivo(filepath)
            
            return True

        except Exception as e:
            error_msg = f"No se pudo exportar el despacho:\n{str(e)}"
            self._mostrar_error("Error al exportar", error_msg)
            self._registrar_error(e)
            return False
        
    def _extraer_codigo_seguro(self, descripcion):
        """Extrae código de manera robusta, con manejo de errores"""
        if not descripcion:
            return None
        
        # Estrategia 1: Buscar patrón "código - descripción"
        if ' - ' in descripcion:
            partes = descripcion.split(' - ')
            posible_codigo = partes[0].strip()
            # Validar que sea un código probable (sin espacios, alfanumérico)
            if (isinstance(posible_codigo, str) and  # Asegurar que es string
                len(posible_codigo.split()) == 1 and 
                len(posible_codigo) > 0):
                return posible_codigo
        
        # Estrategia 2: Buscar al inicio de la descripción
        if isinstance(descripcion, str):  # Asegurar que es string
            palabras = descripcion.split()
            if palabras and len(palabras[0]) <= 20:
                return palabras[0]
        
        # Estrategia 3: Buscar en bultos_data por coincidencia parcial
        for codigo_existente in self.bultos_data.keys():
            if isinstance(codigo_existente, str) and codigo_existente in descripcion:
                return codigo_existente
        
        return None

    def _procesar_articulos_detallado(self):
        """Método alias para compatibilidad con código existente"""
        return self._procesar_articulos(mantener_orden_original=True)
        

    def _organizar_bultos_inteligentemente(self, articulos_orden_original):
        """Organiza los bultos de forma inteligente con mejor aprovechamiento del espacio"""
        
        def determinar_estructura_bultos(cantidad_bultos):
            """Determina la estructura óptima de columnas según la cantidad de bultos"""
            if cantidad_bultos <= 40:
                return 40, 1  # 1 columna, 40 filas
            elif cantidad_bultos <= 80:
                return 40, 2  # 2 columnas, 40 filas cada una
            elif cantidad_bultos <= 120:
                return 40, 3  # 3 columnas, 40 filas cada una
            elif cantidad_bultos <= 160:
                return 40, 4  # 4 columnas, 40 filas cada una
            else:
                return 40, 5  # 5 columnas, 40 filas cada una (máximo 200 bultos)

        def organizar_bultos_optimizado(bultos):
            """Organiza los bultos en columnas de forma optimizada"""
            if not bultos:
                return [], 0
            
            cantidad_bultos = len(bultos)
            max_filas, columnas_necesarias = determinar_estructura_bultos(cantidad_bultos)
            
            # Calcular filas reales necesarias por columna
            filas_por_columna = (cantidad_bultos + columnas_necesarias - 1) // columnas_necesarias
            
            # Organizar bultos en columnas
            columnas_organizadas = []
            for i in range(columnas_necesarias):
                inicio = i * filas_por_columna
                fin = min((i + 1) * filas_por_columna, cantidad_bultos)
                columna = bultos[inicio:fin]
                columnas_organizadas.append(columna)
            
            return columnas_organizadas, columnas_necesarias

        # Aplicar organización optimizada a cada artículo
        for articulo in articulos_orden_original:
            if articulo['pesos']:
                # Obtener solo los pesos válidos
                pesos_validos = []
                for peso in articulo['pesos']:
                    if isinstance(peso, (int, float)) and peso > 0:
                        pesos_validos.append(peso)
                
                if pesos_validos:
                    columnas_organizadas, num_columnas = organizar_bultos_optimizado(pesos_validos)
                    articulo['columnas_organizadas'] = columnas_organizadas
                    articulo['num_columnas_necesarias'] = num_columnas
                    articulo['max_filas_por_columna'] = max(len(col) for col in columnas_organizadas) if columnas_organizadas else 0
                else:
                    articulo['columnas_organizadas'] = []
                    articulo['num_columnas_necesarias'] = 0
                    articulo['max_filas_por_columna'] = 0
            else:
                articulo['columnas_organizadas'] = []
                articulo['num_columnas_necesarias'] = 0
                articulo['max_filas_por_columna'] = 0
        
        return articulos_orden_original

    def _optimizar_distribucion_filas(self, grupos_ordenados):
        """Optimiza la distribución de artículos en filas aprovechando mejor el espacio"""
        filas_organizadas = []
        articulos_por_fila = 5  # Máximo de columnas por fila
        current_fila = []
        espacio_actual = 0
        
        for nombre_base, articulos in grupos_ordenados:
            for articulo in articulos:
                # Calcular espacio necesario para este artículo
                espacio_necesario = articulo.get('num_columnas_necesarias', 1)
                
                # Si el artículo cabe en la fila actual, agregarlo
                if espacio_actual + espacio_necesario <= articulos_por_fila:
                    current_fila.append(articulo)
                    espacio_actual += espacio_necesario
                else:
                    # Si no cabe, guardar la fila actual y empezar nueva
                    if current_fila:
                        filas_organizadas.append(current_fila)
                    
                    # Verificar si el artículo es demasiado grande para una fila
                    if espacio_necesario > articulos_por_fila:
                        # Artículo muy grande, va en su propia fila
                        current_fila = [articulo]
                        espacio_actual = espacio_necesario
                    else:
                        # Artículo puede ir en nueva fila
                        current_fila = [articulo]
                        espacio_actual = espacio_necesario
        
        # Agregar la última fila si tiene contenido
        if current_fila:
            filas_organizadas.append(current_fila)
        
        return filas_organizadas
    

    def _procesar_articulos_detallado(self, datos):
        """Procesa artículos para exportación detallada con validaciones robustas"""
        productos = {}
        articulos_procesados = set()  # Para tracking de artículos
        
        # 1. Primero recoger TODOS los artículos del treeview
        articulos_treeview = []
        for item in self.tree.get_children():
            valores = self.tree.item(item)['values']
            if valores and len(valores) > 1:
                descripcion = valores[1]
                articulos_treeview.append((item, descripcion))
        
        # 2. Procesar cada artículo con múltiples estrategias de backup
        for item, descripcion in articulos_treeview:
            valores = self.tree.item(item)['values']
            
            # Estrategia 1: Extraer código de manera robusta
            codigo = self._extraer_codigo_seguro(descripcion)
            
            # Estrategia 2: Si no se encuentra código, usar el ID del item como clave única
            if not codigo:
                codigo = f"item_{item}"  # Clave única basada en ID del treeview
            
            # Estrategia 3: Obtener peso con validación exhaustiva
            try:
                peso_valor = valores[2] if len(valores) > 2 else 0
                if isinstance(peso_valor, (int, float)):
                    peso = float(peso_valor)
                else:
                    peso_str = str(peso_valor).replace(' kg', '').strip()
                    peso = float(peso_str) if peso_str else 0.0
            except (ValueError, TypeError):
                peso = 0.0
                print(f"ADVERTENCIA: Peso inválido para {descripcion}")
            
            # Estrategia 4: Obtener cantidad
            try:
                cantidad = int(valores[0]) if valores[0] else 1
            except (ValueError, TypeError):
                cantidad = 1
            
            # Estrategia 5: Obtener datos de bultos con validación
            bultos_data = {}
            if codigo in self.bultos_data:
                # Filtrar solo datos válidos de bultos
                for bulto_num, peso_bulto in self.bultos_data[codigo].items():
                    if (isinstance(peso_bulto, (int, float)) and 
                        peso_bulto > 0 and bulto_num != '_unidad'):
                        bultos_data[bulto_num] = peso_bulto
            
            # Agrupar por nombre base para consolidación
            nombre_base = self._obtener_nombre_base(descripcion)
            
            if nombre_base not in productos:
                productos[nombre_base] = {
                    'descripciones': [],
                    'pesos': [],
                    'cantidades': [],
                    'bultos_data': [],
                    'items': []  # Track de items originales
                }
            
            # Agregar a la estructura
            productos[nombre_base]['descripciones'].append(descripcion)
            productos[nombre_base]['pesos'].append(peso)
            productos[nombre_base]['cantidades'].append(cantidad)
            productos[nombre_base]['bultos_data'].append(bultos_data)
            productos[nombre_base]['items'].append(item)
            
            articulos_procesados.add(item)
        
        # 3. Validación final - verificar que todos los artículos fueron procesados
        articulos_faltantes = []
        for item, descripcion in articulos_treeview:
            if item not in articulos_procesados:
                articulos_faltantes.append(descripcion)
        
        if articulos_faltantes:
            print(f"ERROR CRÍTICO: Artículos no procesados: {articulos_faltantes}")
            # Forzar inclusión de artículos faltantes
            for item, descripcion in articulos_treeview:
                if item not in articulos_procesados:
                    # Crear entrada mínima para artículo faltante
                    nombre_base = self._obtener_nombre_base(descripcion)
                    if nombre_base not in productos:
                        productos[nombre_base] = {
                            'descripciones': [],
                            'pesos': [],
                            'cantidades': [],
                            'bultos_data': [],
                            'items': []
                        }
                    productos[nombre_base]['descripciones'].append(descripcion)
                    productos[nombre_base]['pesos'].append(0.0)  # Peso por defecto
                    productos[nombre_base]['cantidades'].append(1)  # Cantidad por defecto
                    productos[nombre_base]['bultos_data'].append({})
                    productos[nombre_base]['items'].append(item)
        
        return productos

    def _procesar_articulos_general(self):
        """Procesa artículos de manera simplificada para exportación general"""
        productos = {}
        
        for item in self.tree.get_children():
            valores = self.tree.item(item)['values']
            
            if not valores or len(valores) < 3:
                continue
                
            descripcion = valores[1]
            cantidad = valores[0]
            peso = float(valores[2].replace(' kg', '')) if isinstance(valores[2], str) else float(valores[2])
            
            if descripcion not in productos:
                productos[descripcion] = {
                    'descripciones': [],
                    'cantidades': [],
                    'pesos': []
                }
                
            productos[descripcion]['descripciones'].append(descripcion)
            productos[descripcion]['cantidades'].append(cantidad)
            productos[descripcion]['pesos'].append(peso)
        
        return sorted(productos.items())

    def _extraer_codigo_seguro(self, descripcion):
        """Extrae código de manera robusta con múltiples estrategias"""
        if not descripcion:
            return None
        
        # Estrategia 1: Buscar patrón "código - descripción"
        if ' - ' in descripcion:
            partes = descripcion.split(' - ')
            posible_codigo = partes[0].strip()
            if len(posible_codigo.split()) == 1 and len(posible_codigo) > 0:
                return posible_codigo
        
        # Estrategia 2: Buscar al inicio
        palabras = descripcion.split()
        if palabras and len(palabras[0]) <= 20:
            return palabras[0]
        
        # Estrategia 3: Buscar en bultos_data por coincidencia
        for codigo_existente in self.bultos_data.keys():
            if codigo_existente in descripcion:
                return codigo_existente
        
        return None

    def _obtener_nombre_base(self, descripcion):
        """Obtiene el nombre base del artículo para agrupación"""
        if ' - ' in descripcion:
            return descripcion.split(' - ', 1)[1]
        return descripcion


    def _exportar_despacho_detallado(self):
        """Exporta el despacho en formato detallado con información de bultos"""
        try:
            # Primero validar y guardar el despacho como detallado
            if not self.despacho_guardado:
                if not self._guardar_despacho('detallado'):
                    return False

            # Configurar nombre del archivo con prefijo DETALLADO y verificación de duplicados
            hoy = datetime.now()
            fecha_str = hoy.strftime("%d-%m-%Y")
            cliente_nombre = self.cliente_actual.get('Nombre', 'DESCONOCIDO').replace(' ', '_')
            default_filename = f"DETALLADO_DESPACHO_{cliente_nombre}_{fecha_str}.xlsx"
            
            # Obtener directorio del escritorio para el diálogo de guardado
            desktop_path = Path.home() / "Desktop"
            if not desktop_path.exists():
                desktop_path = Path.home() / "Escritorio"

            # Obtener productos procesados para exportación
            productos = self._procesar_articulos_para_exportacion('detallado')
            
            if not productos:
                self._mostrar_advertencia("No hay artículos para exportar")
                return False
            
            # Generar nombre único que evite sobrescritura
            nombre_unico = self._generar_nombre_archivo_unico(cliente_nombre, Path(default_filename))
            
            filepath = filedialog.asksaveasfilename(
                title="Guardar Despacho Detallado como...",
                defaultextension=".xlsx",
                filetypes=[("Archivo Excel", "*.xlsx")],
                initialdir=str(desktop_path),
                initialfile=nombre_unico.name
            )

            if not filepath:
                return False
            
            # Determinar si todos los artículos son en metros
            solo_metros = True
            for item in self.tree.get_children():
                valores = self.tree.item(item, 'values')
                descripcion = valores[1] if len(valores) > 1 else ""
                codigo = descripcion.split(' - ')[0] if ' - ' in descripcion else ""
                if codigo in self.bultos_data and '_unidad' in self.bultos_data[codigo]:
                    if self.bultos_data[codigo]['_unidad'] != 'mts':
                        solo_metros = False
                        break
                else:
                    solo_metros = False
                    break

            # Configurar texto según unidad de medida
            unidad_texto = "Metros" if solo_metros else "Peso kg"
            total_texto = "TOTAL MTS" if solo_metros else "TOTAL KG"
            total_general_texto = "TOTAL GENERAL METROS" if solo_metros else "TOTAL GENERAL KG"

            wb = Workbook()
            ws = wb.active
            ws.title = "DESPACHO"[:31]  # Limitar a 31 caracteres

            # ------------------------------------------
            # PROCESAMIENTO AVANZADO DE ARTÍCULOS (MANTENIENDO ORDEN DEL TREEVIEW)
            # ------------------------------------------
            colores_conocidos = ["BLANCO", "OSCURO", "PASTEL", "ESPECIAL", "MELANGE", "NEGRO", 
                                "ROJO", "AZUL", "VERDE", "GRIS", "BEIGE", "CREMA", "AMARILLO", 
                                "POPULAR", "COLORES", "BLANCA", "FINA", "GRUESA", "RAYAS", 
                                "ESTAMPADA", "DISEÑO", "1ERA CALIDAD", "2DA CALIDAD", "VARIOS", "NIKE"]
            
            # Paso 1: Procesar todos los artículos manteniendo el orden original
            articulos_orden_original = []
            for item in self.tree.get_children():
                valores = self.tree.item(item, 'values')
                if not valores or len(valores) < 2:
                    continue
                
                descripcion = valores[1]
                codigo = descripcion.split(' - ')[0] if ' - ' in descripcion else ""
                
                # VERIFICACIÓN CRÍTICA: Solo procesar si existe en bultos_data y tiene pesos válidos
                if not codigo or codigo not in self.bultos_data:
                    continue  # Saltar artículos sin datos de bultos
                
                # Extraer código, nombre base y color
                partes = [p.strip() for p in descripcion.split(' - ') if p.strip()]
                
                # 1. Extraer código (si existe)
                codigo = partes[0] if len(partes) > 0 and ' ' not in partes[0] else ""
                
                # 2. Determinar nombre base (ignorando "TEJIDO")
                nombre_base = ""
                if len(partes) >= 2:
                    nombre_candidato = partes[1].upper()
                    nombre_base = nombre_candidato.replace("TEJIDO", "").strip()
                    
                    if not nombre_base and len(partes) >= 3:
                        nombre_base = partes[2].upper()
                else:
                    nombre_base = " ".join(partes[1:]).upper() if len(partes) > 1 else descripcion.upper()
                
                # 3. Determinar color
                color = "BLANCO"  # Valor por defecto
                for parte in reversed(partes):
                    parte_upper = parte.upper()
                    if any(c in parte_upper for c in colores_conocidos):
                        color = next((c for c in colores_conocidos if c in parte_upper), "BLANCO")
                        break
                
                # 4. Limpieza final del nombre base
                for color_term in colores_conocidos:
                    if nombre_base.endswith(color_term):
                        nombre_base = nombre_base[:-len(color_term)].strip()
                        break
                
                # Guardar datos del artículo
                articulo = {
                    'codigo': codigo,
                    'nombre_base': nombre_base,
                    'color': color,
                    'descripcion': descripcion,
                    'pesos': [],
                    'orden_original': len(articulos_orden_original)
                }

                # Procesar pesos individuales de bultos - CON VALIDACIÓN MEJORADA
                if codigo and codigo in self.bultos_data:
                    # Filtrar solo bultos con peso válido > 0 y que no sean marcadores especiales
                    pesos_validos = []
                    for bulto_num, peso in self.bultos_data[codigo].items():
                        if bulto_num == '_unidad':  # Saltar marcador de unidad
                            continue
                        try:
                            peso_float = float(peso)
                            if peso_float > 0:
                                pesos_validos.append(peso_float)
                        except (ValueError, TypeError):
                            continue
                    
                    # Solo agregar si hay pesos válidos
                    if pesos_validos:
                        articulo['pesos'] = pesos_validos
                    else:
                        continue  # Saltar artículos sin pesos válidos

                # Solo agregar a la lista si tiene pesos
                if articulo['pesos']:
                    articulos_orden_original.append(articulo)

            # Aplicar organización inteligente de bultos
            articulos_orden_original = self._organizar_bultos_inteligentemente(articulos_orden_original)

            # Paso 2: Agrupar artículos con el mismo nombre base (manteniendo el orden de primer aparición)
            articulos_agrupados = {}
            orden_aparicion = {}
            
            for idx, articulo in enumerate(articulos_orden_original):
                nombre_base = articulo['nombre_base']
                if nombre_base not in orden_aparicion:
                    orden_aparicion[nombre_base] = idx
                    
                if nombre_base not in articulos_agrupados:
                    articulos_agrupados[nombre_base] = []
                articulos_agrupados[nombre_base].append(articulo)

            # Ordenar los grupos por su primera aparición
            grupos_ordenados = sorted(articulos_agrupados.items(), key=lambda x: orden_aparicion[x[0]])

            # Paso 3: Optimizar distribución de artículos en filas
            filas_organizadas = self._optimizar_distribucion_filas(grupos_ordenados)

            def determinar_estructura_bultos_mejorada(cantidad_bultos):
                """Determina la estructura óptima de columnas con mejor aprovechamiento"""
                # Configuraciones óptimas para diferentes cantidades de bultos
                configuraciones = [
                    (20, 1),   # Hasta 20 bultos: 1 columna
                    (40, 1),   # 21-40 bultos: 1 columna (40 filas)
                    (60, 2),   # 41-60 bultos: 2 columnas (30 filas cada una)
                    (80, 2),   # 61-80 bultos: 2 columnas (40 filas cada una)
                    (100, 3),  # 81-100 bultos: 3 columnas (34 filas cada una)
                    (120, 3),  # 101-120 bultos: 3 columnas (40 filas cada una)
                    (140, 4),  # 121-140 bultos: 4 columnas (35 filas cada una)
                    (160, 4),  # 141-160 bultos: 4 columnas (40 filas cada una)
                    (180, 5),  # 161-180 bultos: 5 columnas (36 filas cada una)
                    (200, 5)   # 181-200 bultos: 5 columnas (40 filas cada una)
                ]
                
                for max_bultos, columnas in configuraciones:
                    if cantidad_bultos <= max_bultos:
                        filas_por_columna = (cantidad_bultos + columnas - 1) // columnas
                        return filas_por_columna, columnas
                
                # Por defecto para más de 200 bultos (no debería ocurrir)
                return 40, 5    

            # ------------------------------------------
            # ESCRITURA EN EXCEL CON ORGANIZACIÓN OPTIMIZADA (VERSIÓN MEJORADA)
            # ------------------------------------------

            current_row = 1

            # Estilos de bordes
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin'))
            
            thick_border = Border(left=Side(style='medium'), right=Side(style='medium'),
                                top=Side(style='medium'), bottom=Side(style='medium'))
            
            # Configuración de fuentes
            header_font = Font(name='Arial', size=14, bold=True)
            normal_font = Font(name='Arial', size=14)
            bold_font = Font(name='Arial', size=14, bold=True)
            title_font = Font(name='Arial', size=14, bold=True)
            number_font = Font(name='Arial', size=16)
            totales_font = Font(name='Arial', size=16, bold=True)
            total_general_font = Font(name='Arial', size=18, bold=True)

            # Cabecera con información del cliente
            ws.cell(row=current_row, column=1, value="FECHA:").font = bold_font
            ws.cell(row=current_row, column=2, value=hoy.strftime("%d/%m/%Y")).font = normal_font
            current_row += 1
                
            ws.cell(row=current_row, column=1, value="CLIENTE:").font = bold_font
            ws.cell(row=current_row, column=2, value=self.cliente_actual.get('Nombre', '')).font = normal_font
            current_row += 1
                
            ws.cell(row=current_row, column=1, value="RIF:").font = bold_font
            ws.cell(row=current_row, column=2, value=self.cliente_actual.get('RIF', '')).font = normal_font
            current_row += 2  # Espacio antes de la tabla

            # Lista para guardar las celdas de totales
            total_kg_cells = []
            total_bultos_cells = []

            # Para cada fila organizada
            for fila in filas_organizadas:
                # Calcular el número total de columnas (considerando artículos con múltiples columnas)
                total_columnas = 0
                for articulo in fila:
                    # Si el artículo tiene columnas organizadas, usamos ese conteo
                    if 'columnas_organizadas' in articulo:
                        total_columnas += len(articulo['columnas_organizadas'])
                    else:
                        # Si no tiene columnas organizadas, cuenta como 1 columna
                        total_columnas += 1
                
                # NOMBRE DEL TEJIDO (con borde grueso)
                start_col = 1
                end_col = total_columnas + 1
                
                # Escribir encabezado de nombre
                ws.cell(row=current_row, column=1, value="NOMBRE DEL TEJIDO:").font = bold_font
                
                col_actual = 2
                for articulo in fila:
                    # Determinar cuántas columnas ocupa este artículo
                    if 'columnas_organizadas' in articulo:
                        num_columnas = len(articulo['columnas_organizadas'])
                    else:
                        num_columnas = 1
                    
                    # Combinar celdas para el nombre del artículo si ocupa múltiples columnas
                    if num_columnas > 1:
                        ws.merge_cells(start_row=current_row, start_column=col_actual, 
                                    end_row=current_row, end_column=col_actual + num_columnas - 1)
                    
                    cell = ws.cell(row=current_row, column=col_actual, value=articulo['nombre_base'])
                    cell.font = bold_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    col_actual += num_columnas
                
                # Aplicar bordes a toda la fila
                for col in range(1, end_col + 1):
                    ws.cell(row=current_row, column=col).border = thick_border
                
                current_row += 1
                
                # SECCIÓN DE COLORES
                ws.cell(row=current_row, column=1, value="COLORES:").font = bold_font
                
                col_actual = 2
                for articulo in fila:
                    if 'columnas_organizadas' in articulo:
                        num_columnas = len(articulo['columnas_organizadas'])
                    else:
                        num_columnas = 1
                    
                    if num_columnas > 1:
                        ws.merge_cells(start_row=current_row, start_column=col_actual, 
                                    end_row=current_row, end_column=col_actual + num_columnas - 1)
                    
                    cell = ws.cell(row=current_row, column=col_actual, value=articulo['color'])
                    cell.font = bold_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill("solid", fgColor="F2F2F2")
                    
                    col_actual += num_columnas
                
                # Aplicar bordes
                for col in range(1, end_col + 1):
                    ws.cell(row=current_row, column=col).border = thick_border
                
                current_row += 1
                
                # CÓDIGOS
                ws.cell(row=current_row, column=1, value="CÓDIGOS:").font = bold_font
                
                col_actual = 2
                for articulo in fila:
                    if 'columnas_organizadas' in articulo:
                        num_columnas = len(articulo['columnas_organizadas'])
                    else:
                        num_columnas = 1
                    
                    if num_columnas > 1:
                        ws.merge_cells(start_row=current_row, start_column=col_actual, 
                                    end_row=current_row, end_column=col_actual + num_columnas - 1)
                    
                    cell = ws.cell(row=current_row, column=col_actual, value=articulo['codigo'])
                    cell.font = bold_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    col_actual += num_columnas
                
                # Aplicar bordes
                for col in range(1, end_col + 1):
                    ws.cell(row=current_row, column=col).border = thick_border
                
                current_row += 1
                
                # ENCABEZADO DE TABLA
                ws.cell(row=current_row, column=1, value="ITEM").font = bold_font
                ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=current_row, column=1).fill = PatternFill("solid", fgColor="F2F2F2")
                
                col_actual = 2
                for articulo in fila:
                    if 'columnas_organizadas' in articulo:
                        num_columnas = len(articulo['columnas_organizadas'])
                    else:
                        num_columnas = 1
                    
                    for i in range(num_columnas):
                        cell = ws.cell(row=current_row, column=col_actual, value=unidad_texto)
                        cell.font = normal_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.fill = PatternFill("solid", fgColor="F2F2F2")
                        col_actual += 1
                
                # Aplicar bordes
                for col in range(1, end_col + 1):
                    ws.cell(row=current_row, column=col).border = thick_border
                
                current_row += 1
                
                # DATOS DE BULTOS - ORGANIZACIÓN INTELIGENTE
                start_data_row = current_row
                
                # Encontrar el máximo número de filas entre todas las columnas de todos los artículos
                max_filas_por_articulo = 0
                for articulo in fila:
                    if 'columnas_organizadas' in articulo:
                        for columna in articulo['columnas_organizadas']:
                            max_filas_por_articulo = max(max_filas_por_articulo, len(columna))
                    else:
                        max_filas_por_articulo = max(max_filas_por_articulo, len(articulo['pesos']))
                
                # Escribir datos de bultos organizados por columnas
                for bulto_idx in range(max_filas_por_articulo):
                    # Número de bulto
                    ws.cell(row=current_row, column=1, value=bulto_idx + 1).font = number_font
                    ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Para cada artículo, escribir los pesos de sus columnas organizadas
                    col_actual = 2
                    for articulo in fila:
                        if 'columnas_organizadas' in articulo:
                            columnas = articulo['columnas_organizadas']
                        else:
                            # Si no tiene columnas organizadas, crear una sola columna con todos los pesos
                            columnas = [articulo['pesos']]
                        
                        for columna in columnas:
                            if bulto_idx < len(columna):
                                peso = columna[bulto_idx]
                                ws.cell(row=current_row, column=col_actual, value=peso).number_format = '0.00'
                            else:
                                ws.cell(row=current_row, column=col_actual, value="")
                            
                            ws.cell(row=current_row, column=col_actual).font = number_font
                            ws.cell(row=current_row, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
                            col_actual += 1
                    
                    # Aplicar bordes a toda la fila
                    for col in range(1, end_col + 1):
                        ws.cell(row=current_row, column=col).border = thin_border
                    
                    current_row += 1
                
                # TOTALES POR ARTÍCULO (KG) - MODIFICADO
                ws.cell(row=current_row, column=1, value="TOTAL KG:").font = bold_font
                ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center', vertical='center')

                col_actual = 2
                for articulo in fila:
                    if 'columnas_organizadas' in articulo:
                        num_columnas = len(articulo['columnas_organizadas'])
                    else:
                        num_columnas = 1
                    
                    # CALCULAR TOTAL POR COLUMNA INDIVIDUALMENTE (NO COMBINAR)
                    for col_idx in range(num_columnas):
                        # Calcular el rango de celdas para esta columna específica
                        col_letter = get_column_letter(col_actual + col_idx)
                        formula = f"=SUM({col_letter}{start_data_row}:{col_letter}{current_row - 1})"
                        
                        cell = ws.cell(row=current_row, column=col_actual + col_idx, value=formula)
                        cell.font = totales_font
                        cell.number_format = '0.00'
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        total_kg_cells.append(cell.coordinate)
                    
                    col_actual += num_columnas

                # Aplicar bordes con línea doble abajo y gruesos arriba/abajo
                for col in range(1, end_col + 1):
                    ws.cell(row=current_row, column=col).border = Border(
                        left=Side(style='thick'), 
                        right=Side(style='thick'),
                        top=Side(style='thick'),      # Borde grueso arriba
                        bottom=Side(style='thick'))  # Borde doble abajo

                current_row += 1

                # TOTAL BULTOS POR ARTÍCULO - MODIFICADO (POR COLUMNA INDIVIDUAL)
                ws.cell(row=current_row, column=1, value="TOTAL BULTOS:").font = Font(name='Arial', size=16, bold=True, color='000000')
                ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
                ws.cell(row=current_row, column=1).border = thick_border

                col_actual = 2
                for articulo in fila:
                    if 'columnas_organizadas' in articulo:
                        columnas = articulo['columnas_organizadas']
                        num_columnas = len(columnas)
                        
                        # CALCULAR BULTOS POR COLUMNA INDIVIDUALMENTE
                        for col_idx, columna in enumerate(columnas):
                            total_bultos = len(columna)  # Total de bultos en esta columna específica
                            
                            cell = ws.cell(row=current_row, column=col_actual + col_idx, value=total_bultos)
                            cell.font = Font(name='Arial', size=18, bold=True, color='000000')
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.border = thick_border
                            total_bultos_cells.append(cell.coordinate)
                    else:
                        num_columnas = 1
                        total_bultos = len(articulo['pesos'])
                        
                        cell = ws.cell(row=current_row, column=col_actual, value=total_bultos)
                        cell.font = Font(name='Arial', size=18, bold=True, color='000000')
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thick_border
                        total_bultos_cells.append(cell.coordinate)
                    
                    col_actual += num_columnas

                current_row += 1  # Espacio adicional entre grupos de artículos
                

                # ------------------------------------------
                # TOTALES POR FILA (para TODAS las filas) - MODIFICADO
                # ------------------------------------------
                current_row += 2  # 2 filas de separación arriba

                # Total kg para esta fila
                ws.cell(row=current_row, column=1, value="TOTAL KG:").font = Font(name='Arial', size=16, bold=True, color='000000')
                ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='right')
                ws.cell(row=current_row, column=1).border = thick_border

                # Suma de kg de la fila TOTAL (2 filas arriba)
                start_col_letter = 'B'
                end_col_letter = get_column_letter(end_col)
                formula_kg = f"=SUM({start_col_letter}{current_row - 4}:{end_col_letter}{current_row - 4})"
                kg_cell = ws.cell(row=current_row, column=2, value=formula_kg)
                kg_cell.font = Font(name='Arial', size=18, bold=True, color='000000')
                kg_cell.number_format = '0.00'
                kg_cell.alignment = Alignment(horizontal='right')  # Alineado a la derecha
                kg_cell.border = thick_border

                # Promedio de peso por bulto (sin enunciado, alineado a la izquierda)
                formula_promedio = f"=IF({get_column_letter(2)}{current_row + 1}<>0, {get_column_letter(2)}{current_row}/{get_column_letter(2)}{current_row + 1}, 0)"
                promedio_cell = ws.cell(row=current_row, column=3, value=formula_promedio)
                promedio_cell.font = Font(name='Arial', size=16, bold=True, color='000000')
                promedio_cell.number_format = '0.00'
                promedio_cell.alignment = Alignment(horizontal='left')  # Alineado a la izquierda
                # Sin borde para la celda de promedio
                
                current_row += 1  # Siguiente fila para Total Bultos

                # Total bultos para esta fila
                ws.cell(row=current_row, column=1, value="TOTAL BULTOS:").font = Font(name='Arial', size=16, bold=True, color='000000')
                ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='right')
                ws.cell(row=current_row, column=1).border = thick_border

                # Suma de bultos de la fila "TOTAL BULTOS" (2 filas arriba)
                formula_bultos = f"=SUM({start_col_letter}{current_row - 4}:{end_col_letter}{current_row - 4})"
                bultos_cell = ws.cell(row=current_row, column=2, value=formula_bultos)
                bultos_cell.font = Font(name='Arial', size=18, bold=True, color='000000')
                bultos_cell.alignment = Alignment(horizontal='right')  # Alineado a la derecha
                bultos_cell.border = thick_border

                current_row += 2  # 2 filas de separación para la siguiente sección

            # ------------------------------------------
            # TOTAL GENERAL FINAL - SOLO SI HAY MÚLTIPLES ARTÍCULOS
            # ------------------------------------------
            if len(articulos_orden_original) > 1:
                # Configurar estilos de fuente
                total_general_font = Font(name='Arial', size=15, bold=True)
                total_label_font = Font(name='Arial', size=12, bold=True)

                # Fila 1: Totales KG y Promedio
                ws.cell(row=current_row, column=1, value=total_general_texto).font = total_label_font
                ws.cell(row=current_row, column=1).border = thick_border
                ws.cell(row=current_row, column=1).alignment = Alignment(vertical='center')

                # Sumar todos los totales individuales de productos
                ws.cell(row=current_row, column=2, value=f"=SUM({','.join(total_kg_cells)})").font = total_general_font
                ws.cell(row=current_row, column=2).border = thick_border
                ws.cell(row=current_row, column=2).number_format = '0.00'
                ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='right', vertical='center')

                # Promedio general (solo valor numérico)
                ws.cell(row=current_row, column=3, 
                    value=f"=IF(SUM({','.join(total_bultos_cells)})>0,SUM({','.join(total_kg_cells)})/SUM({','.join(total_bultos_cells)}),0)")
                ws.cell(row=current_row, column=3).font = total_general_font
                ws.cell(row=current_row, column=3).number_format = '0.00'
                ws.cell(row=current_row, column=3).alignment = Alignment(horizontal='left', vertical='center')

                # Fila 2: Total Bultos
                ws.cell(row=current_row+1, column=1, value="TOTAL GENERAL BULTOS").font = total_label_font
                ws.cell(row=current_row+1, column=1).border = thick_border
                ws.cell(row=current_row+1, column=1).alignment = Alignment(vertical='center')

                # Sumar todos los bultos individuales
                ws.cell(row=current_row+1, column=2, value=f"=SUM({','.join(total_bultos_cells)})").font = total_general_font
                ws.cell(row=current_row+1, column=2).border = thick_border
                ws.cell(row=current_row+1, column=2).alignment = Alignment(horizontal='right', vertical='center')

            # ------------------------------------------
            # AJUSTES FINALES DE FORMATO
            # ------------------------------------------
            # Ajustar anchos de columnas
            ws.column_dimensions['A'].width = 31
            ws.column_dimensions['B'].width = 34
            ws.column_dimensions['C'].width = 34
            ws.column_dimensions['D'].width = 34
            ws.column_dimensions['E'].width = 34
            ws.column_dimensions['F'].width = 34
            
            # Configurar márgenes
            ws.page_margins = PageMargins(
                left=0.5, right=0.5, top=0.5, bottom=0.5,
                header=0.3, footer=0.3
            )
            
            # Configurar para ajustar a una página de ancho
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            
            # Guardar archivo
            wb.save(filepath)
            
            # Preguntar si abrir el archivo
            respuesta = messagebox.askyesno(
                "Exportación exitosa",
                f"El despacho detallado se exportó correctamente a:\n{os.path.basename(filepath)}\n\n"
                "¿Desea abrir el archivo ahora?")
            
            if respuesta:
                self._abrir_archivo(filepath)
            
            for clave, datos in productos.items():
                # Tu lógica para procesar cada artículo...
                descripcion = datos['descripcion']
                cantidad = datos['cantidad']
                peso_total = datos['peso']
                bultos_data = datos.get('bultos_data', {})
                # Procesamiento garantizado para cada artículo
                self._procesar_articulos_detallado(datos)
            
            return True
            
        except Exception as e:
            error_msg = f"Error en exportación detallada: {str(e)}"
            print(f"Traceback: {traceback.format_exc()}")
            self._mostrar_error("Error", error_msg)
            return False
        
    def _generar_nombre_archivo_unico(self, cliente_nombre, base_filename):
        """Genera un nombre de archivo único con sufijos (2), (3), etc. para el mismo cliente"""
        desktop_path = Path.home() / "Desktop"
        if not desktop_path.exists():
            desktop_path = Path.home() / "Escritorio"
        
        # Convertir base_filename a Path si es string
        if isinstance(base_filename, str):
            base_filename = Path(base_filename)
        
        # Buscar archivos existentes del mismo cliente en las últimas 48 horas
        archivos_existentes = []
        ahora = datetime.now()
        
        # Patrón para buscar archivos del mismo cliente (tanto normales como detallados)
        patron_base = f"*DESPACHO_{cliente_nombre}_*.xlsx"
        
        for archivo in desktop_path.glob(patron_base):
            if archivo.is_file():
                # Verificar si el archivo fue creado en las últimas 48 horas
                try:
                    fecha_creacion = datetime.fromtimestamp(archivo.stat().st_ctime)
                    if (ahora - fecha_creacion) <= timedelta(hours=48):
                        archivos_existentes.append(archivo.name)
                except:
                    continue
        
        if not archivos_existentes:
            return base_filename
        
        # Extraer números de sufijo existentes
        numeros_sufijo = [1]  # Siempre incluir 1 para el caso sin sufijo
        
        # Patrón regex para extraer números de sufijo
        fecha_str = datetime.now().strftime("%d-%m-%Y")
        patron_regex = re.compile(rf'(DETALLADO_)?DESPACHO_{re.escape(cliente_nombre)}_({re.escape(fecha_str)})(?:\((\d+)\))?\.xlsx')
        
        for archivo in archivos_existentes:
            match = patron_regex.match(archivo)
            if match:
                sufijo = match.group(3)  # El grupo 3 contiene el número del sufijo
                if sufijo:
                    numeros_sufijo.append(int(sufijo))
                else:
                    # Si no hay sufijo, es el archivo original (1)
                    numeros_sufijo.append(1)
        
        # Encontrar el siguiente número disponible
        siguiente_numero = max(numeros_sufijo) + 1
        
        # Modificar el nombre base para incluir el sufijo
        nombre_base = str(base_filename.stem)
        
        if siguiente_numero == 2:
            # Para el primer sufijo, usar (2)
            nuevo_nombre = f"{nombre_base}({siguiente_numero}){base_filename.suffix}"
        else:
            nuevo_nombre = f"{nombre_base}({siguiente_numero}){base_filename.suffix}"
        
        return Path(nuevo_nombre)


    def _registrar_proceso_exportacion(self, productos):
        """Registra detalles del proceso de exportación para debugging"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_msg = f"\n=== EXPORTACIÓN DETALLADA {timestamp} ===\n"
        
        log_msg += f"Total artículos en treeview: {len(self.tree.get_children())}\n"
        log_msg += f"Total artículos procesados: {len(productos)}\n"
        
        for i, item in enumerate(self.tree.get_children()):
            valores = self.tree.item(item)['values']
            descripcion = valores[1] if len(valores) > 1 else "SIN DESCRIPCIÓN"
            log_msg += f"Artículo {i+1}: {descripcion}\n"
        
        # Guardar log en archivo
        log_path = Path.home() / "exportacion_logs.txt"
        with open(log_path, 'a', encoding='utf-8') as f:
            f.write(log_msg)
        
        print(log_msg)

    
    
    def _add_border(self, ws, cell_range, border_style='thin'):
        """Añade bordes a un rango de celdas"""
        rows = ws[cell_range]
        side = Side(style=border_style)
        
        for row in rows:
            for cell in row:
                cell.border = Border(
                    left=side, right=side,
                    top=side, bottom=side
                )


    def _escribir_info_exportacion(self, sheet):
        """Escribe la información básica del despacho en la hoja Excel"""
        sheet['A1'] = "FECHA:"
        sheet['A1'].style = 'InfoLabel'
        sheet['B1'] = datetime.now().strftime(self.config.get("formato_fecha", "%d/%m/%Y %H:%M"))
        sheet['B1'].style = 'InfoValue'
        
        sheet['A2'] = "NOMBRE DEL CLIENTE:"
        sheet['A2'].style = 'InfoLabel'
        sheet['B2'] = self.cliente_actual.get('Nombre', '')
        sheet['B2'].style = 'InfoValue'
        
        sheet['A3'] = "RIF:"
        sheet['A3'].style = 'InfoLabel'
        sheet['B3'] = self.cliente_actual.get('RIF', '')
        sheet['B3'].style = 'InfoValue'
        
        

    def _escribir_totales_exportacion(self, sheet):
        """Escribe los totales del despacho en la hoja Excel"""
        total_row = 8 + len(self.tree.get_children())
        
        # Escribir línea de totales
        sheet.cell(row=total_row, column=2, value="TOTAL:").style = 'InfoLabel'
        
        # Sumar columna de peso (C)
        peso_cell = sheet.cell(row=total_row, column=3, value=f"=SUM(C8:C{total_row-1})")
        peso_cell.style = 'Total'

        # Escribir resumen final (2 líneas después del total)
        summary_data = [
            ("TOTAL KG:", f"=C{total_row}"),
            ("TOTAL ARTÍCULOS:", f"=COUNT(A8:A{total_row-1})")
        ]
        
        for i, (label, value) in enumerate(summary_data, start=1):
            sheet.cell(row=total_row + 1 + i, column=2, value=label).style = 'InfoLabel'
            sheet.cell(row=total_row + 1 + i, column=3, value=value).style = 'InfoValue'

    def _configurar_estilos_excel(self, wb):
        """Configura los estilos para el archivo Excel exportado"""
        # Estilo para títulos
        title_style = NamedStyle(name="title_style")
        title_style.font = Font(bold=True, size=12, color="FFFFFF")
        title_style.fill = PatternFill("solid", fgColor="4472C4")
        title_style.alignment = Alignment(horizontal="left")
        wb.add_named_style(title_style)

        # Estilo para encabezados de tabla
        header_style = NamedStyle(name="header_style")
        header_style.font = Font(bold=True, color="000000")
        header_style.fill = PatternFill("solid", fgColor="D9E1F2")
        header_style.alignment = Alignment(horizontal="center")
        header_style.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"))
        wb.add_named_style(header_style)

        # Estilo para datos
        data_style = NamedStyle(name="data_style")
        data_style.font = Font(size=11)
        data_style.alignment = Alignment(horizontal="left")
        data_style.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"))
        wb.add_named_style(data_style)

        # Estilo para totales
        total_style = NamedStyle(name="total_style")
        total_style.font = Font(bold=True)
        total_style.fill = PatternFill("solid", fgColor="F2F2F2")
        total_style.border = Border(
            top=Side(style="double"),
            bottom=Side(style="double"))
        wb.add_named_style(total_style)

    def _aplicar_formato_tabla(self, ws):
        """Aplica formato a la tabla de artículos en el Excel"""
        # Aplicar estilos a los títulos
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.style = "title_style"
        
        # Aplicar estilos a los encabezados de la tabla de artículos
        for row in ws.iter_rows(min_row=15, max_row=15):  # Ajusta estos números según tu estructura
            for cell in row:
                cell.style = "header_style"
        
        # Ajustar anchos de columnas
        ws.column_dimensions['A'].width = 50  # Descripción
        ws.column_dimensions['B'].width = 15  # Bultos
        ws.column_dimensions['C'].width = 20  # Kilos/Metros
        
        # Aplicar estilo a los datos
        for row in ws.iter_rows(min_row=16, max_row=ws.max_row-2):  # Ajusta según tu estructura
            for cell in row:
                cell.style = "data_style"
        
        # Aplicar estilo a los totales
        for row in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row):
            for cell in row:
                cell.style = "total_style"


    def _escribir_tabla_articulos(self, sheet):
        """Escribe la tabla de artículos en la hoja (blanco y negro)"""
        for i, item in enumerate(self.tree.get_children(), start=1):
            valores = self.tree.item(item)['values']
            
            # Escribir cantidad
            sheet.cell(row=7+i, column=1, value=int(valores[0])).style = 'TableRow'
            
            # Escribir descripción
            sheet.cell(row=7+i, column=2, value=valores[1]).style = 'TableRow'
            
            # Escribir peso total (extraer solo el valor numérico)
            peso_str = valores[2].replace(' kg', '')  # Eliminar "kg" del string
            try:
                peso = float(peso_str)
                peso_cell = sheet.cell(row=7+i, column=3, value=peso)
                peso_cell.style = 'TableRow'
            except ValueError:
                # Si no se puede convertir a float, escribir 0
                peso_cell = sheet.cell(row=7+i, column=3, value=0.0)
                peso_cell.style = 'TableRow'

    def _configurar_pagina_impresion(self, sheet, ultima_fila=None):
        """Configura los parámetros de impresión para el documento exportado"""
        # Configurar márgenes
        sheet.page_margins = PageMargins(
            left=0.5, right=0.5, top=0.75, bottom=0.75,
            header=0.3, footer=0.3
        )

        # Configurar orientación y ajuste
        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0

        # Configurar área de impresión
        ultima_fila = 8 + len(self.tree.get_children()) + 3
        sheet.print_area = f"A1:F{ultima_fila}"

        # Configurar encabezado y pie de página (versión compatible)
        try:
            # Intenta usar header_footer si está disponible
            sheet.header_footer.center_header.text = "&\"Arial,Bold\"&14DESPACHO DE MATERIAL"
            sheet.header_footer.center_footer.text = "&\"Arial\"&10© 2024 - Sistema de Gestión de Despachos"
        except AttributeError:
            # Si header_footer no está disponible, usa alternativas
            sheet.oddHeader.center.text = "DESPACHO DE MATERIAL"
            sheet.oddHeader.center.size = 14
            sheet.oddHeader.center.font = "Arial,Bold"
            sheet.oddFooter.center.text = "© 2024 - Sistema de Gestión de Despachos"
            sheet.oddFooter.center.size = 10
            sheet.oddFooter.center.font = "Arial"

        # Repetir filas de encabezado
        sheet.print_title_rows = '7:7'
        # Configurar área de impresión si se proporciona ultima_fila
        if ultima_fila:
            sheet.print_area = f"A1:F{ultima_fila}"

    def _abrir_archivo(self, filepath):
        """Abre el archivo generado con el programa predeterminado"""
        try:
            if sys.platform == "win32":
                os.startfile(filepath)
            elif sys.platform == "darwin":
                subprocess.run(["open", filepath])
            else:
                subprocess.run(["xdg-open", filepath])
        except Exception as e:
            self._mostrar_advertencia(
                "Aviso", 
                f"No se pudo abrir el archivo:\n{str(e)}\n"
                f"Puede encontrarlo en:\n{filepath}")

    def _mostrar_configuracion(self):
        """Muestra el diálogo de configuración de la aplicación"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Configuración del Sistema")
        dialog.geometry("700x550")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()

        # Frame principal
        main_frame = ttk.Frame(dialog, padding=20)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # Pestañas
        notebook = ttk.Notebook(main_frame)
        notebook.pack(fill=tk.BOTH, expand=True)

        # Pestaña de Columnas
        columns_tab = ttk.Frame(notebook, padding=10)
        notebook.add(columns_tab, text="Columnas")

        # Configuración de columnas para clientes
        ttk.Label(
            columns_tab, 
            text="Nombres de columnas en hoja 'Clientes':",
            font=('Segoe UI', 10, 'bold')
        ).pack(anchor=tk.W, pady=(5, 2))

        self.columnas_clientes_var = tk.StringVar(
            value=", ".join(self.config.get("columnas_clientes", [])))
        
        columnas_clientes_frame = ttk.Frame(columns_tab)
        columnas_clientes_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Label(columnas_clientes_frame, text="Orden:").pack(side=tk.LEFT)
        self.columnas_clientes_entry = ttk.Entry(
            columnas_clientes_frame, 
            textvariable=self.columnas_clientes_var,
            width=50)
        self.columnas_clientes_entry.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)
        
        ttk.Label(
            columns_tab, 
            text="Ejemplo: Nombre, RIF, Teléfono, Dirección",
            font=('Segoe UI', 8),
            foreground='#666666'
        ).pack(anchor=tk.W, pady=(0, 15))

        # Configuración de columnas para artículos
        ttk.Label(
            columns_tab, 
            text="Nombres de columnas en hoja 'Artículos':",
            font=('Segoe UI', 10, 'bold')
        ).pack(anchor=tk.W, pady=(5, 2))

        self.columnas_articulos_var = tk.StringVar(
            value=", ".join(self.config.get("columnas_articulos", [])))
        
        columnas_articulos_frame = ttk.Frame(columns_tab)
        columnas_articulos_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Label(columnas_articulos_frame, text="Orden:").pack(side=tk.LEFT)
        self.columnas_articulos_entry = ttk.Entry(
            columnas_articulos_frame, 
            textvariable=self.columnas_articulos_var,
            width=50)
        self.columnas_articulos_entry.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)
        
        ttk.Label(
            columns_tab, 
            text="Ejemplo: Código Importado, Descripción Importado, Código Nacional, Descripción Nacional",
            font=('Segoe UI', 8),
            foreground='#666666'
        ).pack(anchor=tk.W, pady=(0, 15))

        # Pestaña de Preferencias
        prefs_tab = ttk.Frame(notebook, padding=10)
        notebook.add(prefs_tab, text="Preferencias")

        # Configuración de tipos de artículos
        ttk.Label(
            prefs_tab, 
            text="Tipos de Artículos:",
            font=('Segoe UI', 10, 'bold')
        ).pack(anchor=tk.W, pady=(5, 2))

        self.tipos_articulos_var = tk.StringVar(
            value=", ".join(self.config.get("tipo_articulos", [])))
        
        tipos_articulos_frame = ttk.Frame(prefs_tab)
        tipos_articulos_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Label(tipos_articulos_frame, text="Valores:").pack(side=tk.LEFT)
        self.tipos_articulos_entry = ttk.Entry(
            tipos_articulos_frame, 
            textvariable=self.tipos_articulos_var,
            width=50)
        self.tipos_articulos_entry.pack(side=tk.LEFT, padx=5, expand=True, fill=tk.X)
        
        ttk.Label(
            prefs_tab, 
            text="Ejemplo: Importado, Nacional, Especial",
            font=('Segoe UI', 8),
            foreground='#666666'
        ).pack(anchor=tk.W, pady=(0, 15))

        # Configuración de formato de fecha
        ttk.Label(
            prefs_tab, 
            text="Formato de Fecha:",
            font=('Segoe UI', 10, 'bold')
        ).pack(anchor=tk.W, pady=(5, 2))

        self.formato_fecha_var = tk.StringVar(
            value=self.config.get("formato_fecha", "%d/%m/%Y"))
        
        formato_fecha_frame = ttk.Frame(prefs_tab)
        formato_fecha_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Label(formato_fecha_frame, text="Formato:").pack(side=tk.LEFT)
        self.formato_fecha_combobox = ttk.Combobox(
            formato_fecha_frame,
            textvariable=self.formato_fecha_var,
            values=[
                "%d/%m/%Y", 
                "%m/%d/%Y", 
                "%Y-%m-%d", 
                "%d-%m-%Y", 
                "%d/%m/%Y %H:%M", 
                "%Y%m%d_%H%M%S"
            ],
            state="readonly",
            width=20)
        self.formato_fecha_combobox.pack(side=tk.LEFT, padx=5)
        
        ttk.Label(
            prefs_tab, 
            text="Ejemplo actual: " + datetime.now().strftime(
                self.config.get("formato_fecha", "%d/%m/%Y")),
            font=('Segoe UI', 8),
            foreground='#666666'
        ).pack(anchor=tk.W, pady=(0, 15))

        # Configuración de valores por defecto
        ttk.Label(
            prefs_tab, 
            text="Valores por Defecto:",
            font=('Segoe UI', 10, 'bold')
        ).pack(anchor=tk.W, pady=(10, 2))

        # Nombre de tejido por defecto
        ttk.Label(
            prefs_tab, 
            text="Nombre de Tejido:",
            font=('Segoe UI', 9)
        ).pack(anchor=tk.W, pady=(5, 2))

        self.tejido_default_var = tk.StringVar(
            value=self.config.get("tejido_default", "SABINA DRY FIT"))
        
        tejido_default_frame = ttk.Frame(prefs_tab)
        tejido_default_frame.pack(fill=tk.X, padx=5, pady=2)
        
        ttk.Entry(
            tejido_default_frame,
            textvariable=self.tejido_default_var,
            width=40).pack(side=tk.LEFT, padx=5)

        # Código de tejido por defecto
        ttk.Label(
            prefs_tab, 
            text="Código de Tejido:",
            font=('Segoe UI', 9)
        ).pack(anchor=tk.W, pady=(5, 2))

        self.codigo_tejido_default_var = tk.StringVar(
            value=self.config.get("codigo_tejido_default", "TN0218"))
        
        codigo_tejido_frame = ttk.Frame(prefs_tab)
        codigo_tejido_frame.pack(fill=tk.X, padx=5, pady=2)
        
        ttk.Entry(
            codigo_tejido_frame,
            textvariable=self.codigo_tejido_default_var,
            width=20).pack(side=tk.LEFT, padx=5)

        # Opción para mostrar pesos individuales
        self.mostrar_pesos_var = tk.BooleanVar(
            value=self.config.get("mostrar_pesos_individuales", True))
        
        ttk.Checkbutton(
            prefs_tab,
            text="Mostrar pesos individuales de bultos",
            variable=self.mostrar_pesos_var
        ).pack(anchor=tk.W, pady=(15, 5))

        # Botones
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=(15, 0))

        ttk.Button(
            button_frame,
            text="Guardar Configuración",
            command=lambda: self._guardar_configuracion_dialogo(dialog),
            style='Accent.TButton'
        ).pack(side=tk.LEFT, padx=5)

        ttk.Button(
            button_frame,
            text="Cancelar",
            command=dialog.destroy,
            style='Secondary.TButton'
        ).pack(side=tk.RIGHT, padx=5)

    def _guardar_configuracion_dialogo(self, dialog):
        """Guarda la configuración desde el diálogo"""
        try:
            # Validar y procesar columnas para clientes
            columnas_clientes = [
                col.strip() for col in self.columnas_clientes_var.get().split(",") 
                if col.strip()
            ]
            if len(columnas_clientes) < 2:
                raise ValueError("Debe especificar al menos 2 columnas para clientes")

            # Validar y procesar columnas para artículos
            columnas_articulos = [
                col.strip() for col in self.columnas_articulos_var.get().split(",") 
                if col.strip()
            ]
            if len(columnas_articulos) < 4:
                raise ValueError("Debe especificar al menos 4 columnas para artículos")

            # Validar y procesar tipos de artículos
            tipos_articulos = [
                tipo.strip() for tipo in self.tipos_articulos_var.get().split(",") 
                if tipo.strip()
            ]
            if len(tipos_articulos) < 2:
                raise ValueError("Debe especificar al menos 2 tipos de artículos")

            # Validar formato de fecha
            try:
                datetime.now().strftime(self.formato_fecha_var.get())
            except ValueError:
                raise ValueError("Formato de fecha inválido")

            # Actualizar configuración
            self.config.update({
                "columnas_clientes": columnas_clientes,
                "columnas_articulos": columnas_articulos,
                "tipo_articulos": tipos_articulos,
                "formato_fecha": self.formato_fecha_var.get(),
                "tejido_default": self.tejido_default_var.get(),
                "codigo_tejido_default": self.codigo_tejido_default_var.get(),
                "mostrar_pesos_individuales": self.mostrar_pesos_var.get()
            })

            # Guardar en archivo
            if self._guardar_configuracion():
                messagebox.showinfo(
                    "Configuración Guardada",
                    "Los cambios se guardaron correctamente.\n\n"
                    "Algunos cambios pueden requerir reiniciar la aplicación.",
                    parent=dialog
                )
                dialog.destroy()
                # Actualizar combobox de tipos de artículos
                self.tipo_combobox['values'] = tipos_articulos
                
        except ValueError as e:
            messagebox.showerror("Error en Configuración", str(e), parent=dialog)
        except Exception as e:
            messagebox.showerror(
                "Error", 
                f"No se pudo guardar la configuración:\n{str(e)}",
                parent=dialog)
            self._registrar_error(e)

    def _mostrar_documentacion(self):
        """Muestra un diálogo con documentación básica del sistema"""
        doc_text = """
        SISTEMA DE GESTIÓN DE DESPACHOS - MANUAL RÁPIDO
        
        1. SELECCIÓN DE CLIENTE
        - Busque clientes por nombre o RIF
        - Seleccione un cliente de la lista
        - Los datos se cargarán automáticamente
        
        2. AGREGAR ARTÍCULOS
        - Seleccione el tipo de artículo
        - Haga clic en 'Agregar Artículo'
        - Busque y seleccione el artículo deseado
        - Ingrese cantidad y peso (use 'Cálculo de Peso' para múltiples bultos)
        - Agregue observaciones si es necesario
        
        3. GESTIÓN DEL DESPACHO
        - Verifique el peso total
        - Elimine artículos si es necesario (Seleccione y click en 'Eliminar')
        - Use 'Nuevo Despacho' para limpiar el formulario
        
        4. GUARDAR Y EXPORTAR
        - 'Guardar Despacho': Guarda en el archivo Excel principal
        - 'Exportar a Excel': Crea un archivo independiente con formato profesional
        
        5. CÁLCULO DE PESO DE BULTOS
        - Para artículos con múltiples bultos:
        1. Ingrese la cantidad de bultos
        2. Click en 'Cálculo de Peso'
        3. Ingrese el peso individual de cada bulto
        4. El sistema calculará automáticamente el total
        
        CONFIGURACIÓN:
        - Personalice columnas, tipos de artículos y formatos en el menú Configuración
        """
        
        dialog = tk.Toplevel(self.root)
        dialog.title("Documentación del Sistema")
        dialog.geometry("700x500")
        
        # Frame principal
        main_frame = ttk.Frame(dialog, padding=15)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Área de texto con scroll
        text_frame = ttk.Frame(main_frame)
        text_frame.pack(fill=tk.BOTH, expand=True)
        
        text_scroll = ttk.Scrollbar(text_frame)
        text_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        doc_text_widget = tk.Text(
            text_frame,
            wrap=tk.WORD,
            yscrollcommand=text_scroll.set,
            font=('Segoe UI', 10),
            padx=10,
            pady=10
        )
        doc_text_widget.pack(fill=tk.BOTH, expand=True)
        text_scroll.config(command=doc_text_widget.yview)
        
        # Insertar texto con formato
        doc_text_widget.insert(tk.END, doc_text)
        doc_text_widget.config(state=tk.DISABLED)
        
        # Botón de cierre
        ttk.Button(
            main_frame,
            text="Cerrar",
            command=dialog.destroy,
            style='Accent.TButton'
        ).pack(pady=10)

    def _mostrar_acerca_de(self):
        """Muestra el diálogo 'Acerca de' con información de la aplicación"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Acerca de Sistema de Despachos")
        dialog.geometry("650x550")  # Aumenté el tamaño a 650x550 para mejor visualización
        dialog.resizable(False, False)
        
        # Frame principal
        main_frame = ttk.Frame(dialog, padding=25)  # Aumenté el padding
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Logo o icono
        ttk.Label(
            main_frame, 
            text="📦",  # Emoji de paquete
            font=('Arial', 48),
            justify=tk.CENTER
        ).pack(pady=15)  # Aumenté el pady
        
        # Información de la aplicación
        info_text = """
        SISTEMA DE GESTIÓN DE DESPACHOS v3.0
        
        Desarrollado por: Diego Borges
        
        © 2025 Todos los derechos reservados
        
        Características principales:
        - Registro completo de despachos
        - Cálculo preciso de pesos
        - Exportación profesional a Excel
        - Configuración personalizable
        
        Contacto:
        Email: diegoborges12082003@gmail.com
        Teléfono: +58412 707 8504
        """
        
        info_label = ttk.Label(
            main_frame, 
            text=info_text,
            justify=tk.LEFT,
            font=('Segoe UI', 11)  # Aumenté ligeramente el tamaño de fuente
        )
        info_label.pack(pady=15, fill=tk.X, padx=20)
        
        # Frame para botones (con más espacio)
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(pady=25, fill=tk.X, padx=50)  # Más espacio y padding horizontal
        
        # Configurar botones con un ancho adecuado
        button_width = 18  # Ancho aumentado para los botones
        
        # Botón de cierre
        ttk.Button(
            button_frame,
            text="Cerrar",
            command=dialog.destroy,
            style='Accent.TButton',
            width=button_width
        ).pack(side=tk.LEFT, expand=True)
        
        # Espaciador entre botones
        ttk.Frame(button_frame, width=20).pack(side=tk.LEFT)
        
        # Botón de documentación
        ttk.Button(
            button_frame,
            text="Documentación",
            command=self._mostrar_documentacion,
            style='Secondary.TButton',
            width=button_width
        ).pack(side=tk.RIGHT, expand=True)

    def _registrar_error(self, error: Exception):
        """Registra errores en un archivo de log"""
        log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'error_log.txt')
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        try:
            with open(log_path, 'a', encoding='utf-8') as f:
                f.write(f"\n[{timestamp}] ERROR:\n")
                f.write(f"Tipo: {type(error).__name__}\n")
                f.write(f"Mensaje: {str(error)}\n")
                
                # Registrar información del despacho actual si existe
                if hasattr(self, 'cliente_actual') and self.cliente_actual:
                    f.write(f"Cliente: {self.cliente_actual.get('Nombre', 'Desconocido')}\n")
                
                if hasattr(self, 'tree') and self.tree.get_children():
                    f.write(f"Artículos en despacho: {len(self.tree.get_children())}\n")
                
                f.write("-" * 50 + "\n")
        except Exception:
            pass  # Si no se puede escribir el log, no hacer nada

    def _mostrar_error(self, titulo: str, mensaje: str):
        """Muestra un mensaje de error"""
        messagebox.showerror(titulo, mensaje, parent=self.root)

    def _mostrar_advertencia(self, mensaje: str):
        """Muestra un mensaje de advertencia"""
        messagebox.showwarning("Advertencia", mensaje, parent=self.root)

    def _mostrar_info(self, titulo: str, mensaje: str):
        """Muestra un mensaje informativo"""
        messagebox.showinfo(titulo, mensaje, parent=self.root)

    def _actualizar_estado(self, mensaje: str):
        """Actualiza la barra de estado"""
        if hasattr(self, 'statusbar'):
            self.statusbar.config(text=mensaje)

def main():
    """Función principal para iniciar la aplicación"""
    try:
        # Configurar manejo de excepciones no capturadas
        sys.excepthook = lambda exc_type, exc_value, exc_traceback: (
            messagebox.showerror(
                "Error Crítico",
                f"Ocurrió un error inesperado:\n\n{''.join(traceback.format_exception(exc_type, exc_value, exc_traceback))}\n\n"
                "La aplicación se cerrará. Revise el archivo error_log.txt para más detalles."),
            sys.exit(1)
        )
        
        # Crear ventana principal
        root = tk.Tk()
        
        # Configurar tema si está disponible
        try:
            import ttkthemes
            style = ttkthemes.ThemedStyle(root)
            style.set_theme("arc")  # Tema moderno
        except ImportError:
            pass
        
        # Crear y ejecutar aplicación
        app = AplicacionDespachos(root)
        root.mainloop()
        
    except Exception as e:
        messagebox.showerror(
            "Error Inicial", 
            f"No se pudo iniciar la aplicación:\n{str(e)}\n\n"
            "Verifique que tenga todas las dependencias instaladas.")
        
        # Registrar error
        with open('error_log.txt', 'a', encoding='utf-8') as f:
            f.write(f"\n[{datetime.now()}] ERROR DE INICIO:\n")
            f.write(traceback.format_exc())
            f.write("\n" + "=" * 80 + "\n")

def main():
    """Función principal de la aplicación"""
    # Configurar manejo de excepciones
    sys.excepthook = excepthook
    
    root = tk.Tk()
    
    # Obtener parámetros desde línea de comandos
    import argparse
    parser = argparse.ArgumentParser(description='Sistema de Gestión de Despachos')
    parser.add_argument('--registro', '-r', default='registro_despachos.json',
                    help='Nombre del archivo de registro (por defecto: registro_despachos.json)')
    parser.add_argument('--combinado', '-c', action='store_true',
                    help='Modo combinado para leer múltiples registros')
    
    # Solo parsear si hay argumentos, para evitar conflictos con otros usos de sys.argv
    if len(sys.argv) > 1 and not sys.argv[1].startswith('-'):  # Para compatibilidad con empaquetadores
        args = parser.parse_args()
    else:
        args = argparse.Namespace(registro='registro_despachos.json', combinado=False)
    
    # Crear la aplicación con los parámetros especificados
    app = AplicacionDespachos(root, nombre_registro=args.registro, modo_combinado=args.combinado)
    
    # Configurar para que la ventana principal se abra maximizada
    root.state('zoomed')
    
    root.mainloop()

if __name__ == "__main__":
    root = tk.Tk()
    app = AplicacionDespachos(root, "registro_despachos.json", modo_combinado=True)  # Usará el nombre predeterminado modificado
    root.mainloop()
